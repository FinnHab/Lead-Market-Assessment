"""
Administrative feasibility index builder
======================================
Equal-width 5 bins on [0, 1] alongside rank-based tiers.
Two indicators normalised separately, 50/50 composite. Z-score robustness.
n=9 reference sectors (C20.1 excluded).
Tier 5 = Highest Feasibility → Tier 1 = Lowest.
Bin = absolute score-based; Tier = relative rank-based.
"""

import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY      = "203864"
HDR_BLUE  = "2E5FA3"
SUB_BLUE  = "E8EEF7"
ROW_A     = "FFFFFF"
ROW_B     = "F2F2F2"
KEY_GREEN = "E2EFDA"

TIER_BG = {1: "FF6B6B", 2: "FFC7CE", 3: "FFEB9C", 4: "92D050", 5: "C6EFCE"}
TIER_FC = {1: "FFFFFF", 2: "9C0006", 3: "7D6608", 4: "276221", 5: "276221"}

TARGET_SECTORS = {
    "Processing and preserving of meat and production of meat products",
    "Manufacture of dairy products",
    "Manufacture of vegetable and animal oils and fats",
    "Manufacture of grain mill products, starches and starch products",
    "Manufacture of prepared animal feeds",
}

def _side():    return Side(style="thin")
def _border():  return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="center", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)

def sc(ws, row, col, val, fmt=None, bg=None, bold=False, sz=10, color="000000",
       halign="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT_NAME, bold=bold, size=sz, color=color)
    c.border    = _border()
    c.alignment = _al(halign, wrap)
    if bg:  c.fill          = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, col, val, bg=HDR_BLUE, sz=10, color="FFFFFF", halign="center"):
    return sc(ws, row, col, val, bg=bg, bold=True, sz=sz, color=color, halign=halign)

def title_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _al()
    c.border    = _border()

def tier_label(t):
    return {1: "Tier 1\n(Low)", 2: "Tier 2", 3: "Tier 3\n(Mid)",
            4: "Tier 4",       5: "Tier 5\n(High)"}[t]

def tier_cell(ws, row, col, t):
    return sc(ws, row, col, tier_label(t), bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])

def assign_tiers_n9(scores):
    """n=9: ranks 1-2→T5, 3-4→T4, 5-6→T3, 7-8→T2, 9→T1."""
    import pandas as pd
    s = pd.Series(scores)
    ranks = s.rank(ascending=False, method="first").astype(int)
    mapping = {1:5, 2:5, 3:4, 4:4, 5:3, 6:3, 7:2, 8:2, 9:1}
    return [mapping[r] for r in ranks.tolist()]

def assign_bins(score):
    """Equal-width 5 bins on [0,1]. Absolute score-based (not rank-based)."""
    if score < 0.20: return 1
    if score < 0.40: return 2
    if score < 0.60: return 3
    if score < 0.80: return 4
    return 5


# ------------------------------------------------------------
# RAW DATA  (Eurostat SBS sbs_sc_ovw, 2022, EU-27)
# ------------------------------------------------------------
RAW = [
    ("Manufacture of vegetable and animal oils and fats",            "C10.4",   42, 54.58),
    ("Manufacture of grain mill products, starches and starch products",         "C10.6",   70, 45.52),
    ("Manufacture of prepared animal feeds",  "C10.9",   89, 42.43),
    ("Manufacture of dairy products",         "C10.5",  270, 73.17),
    ("Processing and preserving of fruit and vegetables",     "C10.3",  210, 56.46),
    ("Manufacture of beverages",              "C11.0",  260, 63.04),
    ("Manufacture of other food products",    "C10.8",  437, 68.00),
    ("Processing and preserving of meat and production of meat products", "C10.1",  526, 62.22),
    ("Manufacture of bakery and farinaceous products",        "C10.7",  600, 44.71),
]

N = len(RAW)
labels   = [r[0] for r in RAW]
nace     = [r[1] for r in RAW]
n_large  = np.array([r[2] for r in RAW], dtype=float)
mkt_sh   = np.array([r[3] for r in RAW], dtype=float)

# ------------------------------------------------------------
# MIN-MAX NORMALIZATION  (two indicators, 50/50)
# ------------------------------------------------------------
mm_nl = (n_large.max() - n_large) / (n_large.max() - n_large.min())
mm_ms = (mkt_sh - mkt_sh.min()) / (mkt_sh.max() - mkt_sh.min())
mm_af = 0.5 * mm_nl + 0.5 * mm_ms

# ------------------------------------------------------------
# Z-SCORE NORMALIZATION  (robustness)
# ------------------------------------------------------------
z_nl_raw = stats.zscore(-n_large)
z_ms_raw = stats.zscore(mkt_sh)

def rescale01(a):
    return (a - a.min()) / (a.max() - a.min())

z_nl = rescale01(z_nl_raw)
z_ms = rescale01(z_ms_raw)
z_af = 0.5 * z_nl + 0.5 * z_ms

# ------------------------------------------------------------
# RANKS, TIERS, BINS
# ------------------------------------------------------------
mm_scores = mm_af.tolist()
z_scores  = z_af.tolist()
tiers_mm  = assign_tiers_n9(mm_scores)
tiers_z   = assign_tiers_n9(z_scores)
bins_mm   = [assign_bins(s) for s in mm_scores]
bins_z    = [assign_bins(s) for s in z_scores]

rank_mm = [0]*N; rank_z = [0]*N
for pos, idx in enumerate(sorted(range(N), key=lambda i: -mm_scores[i])): rank_mm[idx] = pos+1
for pos, idx in enumerate(sorted(range(N), key=lambda i: -z_scores[i])):  rank_z[idx]  = pos+1

rho, p_rho = stats.spearmanr(rank_mm, rank_z)


# ------------------------------------------------------------
# SHEET 1 — AF Comparison
# ------------------------------------------------------------
def build_comparison(wb):
    ws = wb.active
    ws.title = "01_AF_Comparison"
    NC = 16

    title_row(ws, 1, NC,
              "Administrative Feasibility (AF) – Min-Max vs. Z-Score Comparison  "
              f"|  n={N}  |  Tier 5 = Highest Feasibility → Tier 1 = Lowest  |  Spearman ρ = {rho:.3f}")
    ws.row_dimensions[1].height = 22

    spans = [
        (1,  2,  "Sector"),
        (3,  4,  "Raw Indicators"),
        (5,  10, "Min-Max Normalization  (N Large inverted)"),
        (11, 16, "Z-Score Normalization  (N Large inverted, rescaled to 0–1)"),
    ]
    for c1, c2, text in spans:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, text)
        c.font      = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(HDR_BLUE)
        c.alignment = _al()
        c.border    = _border()
    ws.row_dimensions[2].height = 24

    col_hdrs = ["Sector", "",
                "N Large\nEnterpr.", "Market Share\nLarge (%)",
                "N Large\n(inv.)", "Mkt Share\n(norm.)", "AF Score", "Bin", "Rank", "Tier",
                "Z(N Large)\n(inv.)", "Z(Mkt Share)", "AF Score\n(z, rescaled)", "Bin\n(z)", "Rank (z)", "Tier (z)"]
    for j, h in enumerate(col_hdrs):
        hdr(ws, 3, j+1, h, bg=SUB_BLUE, color="000000")
    ws.merge_cells("A3:B3")
    ws.row_dimensions[3].height = 36

    order = sorted(range(N), key=lambda i: rank_mm[i])
    for i, idx in enumerate(order):
        r   = 4 + i
        bg  = ROW_A if i % 2 == 0 else ROW_B
        ist = labels[idx] in TARGET_SECTORS
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 23

        def s(col, val, fmt=None, extra_bold=False):
            sc(ws, r, col, val, fmt=fmt, bg=bg, bold=(ist or extra_bold),
               halign="left" if col == 1 else "center")

        s(1,  labels[idx])
        s(3,  int(n_large[idx]),        "#,##0")
        s(4,  mkt_sh[idx],             "0.00\"%\"")
        s(5,  round(mm_nl[idx], 4),    "0.0000")
        s(6,  round(mm_ms[idx], 4),    "0.0000")
        s(7,  round(mm_scores[idx], 4),"0.0000", extra_bold=True)
        tier_cell(ws, r, 8, bins_mm[idx])
        s(9,  rank_mm[idx])
        tier_cell(ws, r, 10, tiers_mm[idx])
        s(11, round(z_nl[idx], 4),     "0.0000")
        s(12, round(z_ms[idx], 4),     "0.0000")
        s(13, round(z_scores[idx], 4), "0.0000", extra_bold=True)
        tier_cell(ws, r, 14, bins_z[idx])
        s(15, rank_z[idx])
        tier_cell(ws, r, 16, tiers_z[idx])

    r_key = 4 + N + 2
    ws.merge_cells(f"A{r_key}:{get_column_letter(NC)}{r_key}")
    c = ws.cell(r_key, 1,
        f"Key finding: Min-Max and Z-score normalization produce "
        f"{'identical' if rho >= 0.99 else 'highly similar'} sector rankings "
        f"(Spearman ρ = {rho:.3f}). Results are robust to normalization choice.")
    c.font      = Font(name=FONT_NAME, bold=True, size=10, color="000000")
    c.fill      = _fill(KEY_GREEN)
    c.alignment = _al("left")
    c.border    = _border()

    notes = [
        "  1. Data source: Eurostat SBS (sbs_sc_ovw), reference year 2022, EU-27 aggregates.",
        "  2. Two indicators: (a) Number of Large Enterprises (>250 employees) — INVERTED: fewer = higher feasibility.",
        "     (b) Market Share of Large Enterprises in output value (%) — higher = more market covered by few actors = higher feasibility.",
        "  3. Min-Max normalization applied to each indicator separately, then combined with equal weights (50%/50%).",
        "  4. Z-score robustness: z-scores computed per indicator (N Large negated before z), rescaled to [0,1], then combined 50/50.",
        "  5. Tiers: rank-based (relative positioning, n=9: ranks 1–2 → T5, 3–4 → T4, 5–6 → T3, 7–8 → T2, 9 → T1).",
        "  6. Bins: equal-width 5 bins on [0,1] score range (<0.20 = Bin 1, …, ≥0.80 = Bin 5). Absolute score-based, not rank-based.",
        "  7. C20.1 (Basic Chemicals/Fertilisers) excluded from quantitative analysis: NACE 2-digit too broad (fertilisers = 8–15% of C20.1 output).",
        "     Fertiliser production qualitatively highly concentrated (Yara, BASF, Borealis). Admin Feasibility: qualitatively HIGH.",
        "  8. C10.8 (Manufacture of other food products): Production value for 250+ enterprises confidential in Eurostat; derived by subtraction (Total − smaller size classes).",
        "  9. Higher Tier/Bin = better Administrative Feasibility = stronger GLM candidate on this dimension.",
        "  10. Rationale for separate normalization: A single ratio (Market Share / N Large) implicitly overweights firm count,",
        "     penalizing sectors with many large firms despite high collective market control (e.g., Dairy: 73% share, 270 firms).",
    ]
    for k, note in enumerate(notes):
        rn = r_key + 1 + k
        ws.merge_cells(f"A{rn}:{get_column_letter(NC)}{rn}")
        c = ws.cell(rn, 1, note)
        c.font      = Font(name=FONT_NAME, bold=False, size=9, color="000000")
        c.alignment = _al("left")
        c.border    = _border()

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 1
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"


# ------------------------------------------------------------
# SHEET 2 — Ranking
# ------------------------------------------------------------
def build_ranking(wb):
    ws = wb.create_sheet("02_Ranking")
    NC = 11

    title_row(ws, 1, NC,
              f"AF Ranking – Sorted by Score  |  Min-Max & Z-Score "
              f"{'identical' if rho >= 0.99 else 'ρ=' + f'{rho:.3f}'}  "
              f"|  5 Tiers + 5 Bins")

    col_hdrs = ["Rank", "Bin", "Tier", "Sector", "AF Score\n(Min-Max)", "AF Score\n(Z, rescaled)",
                "Bin\n(Z)", "Tier\n(Z)", "N Large\n(inv., MM)", "Mkt Share\n(norm., MM)", "NACE"]
    for j, h in enumerate(col_hdrs):
        hdr(ws, 2, j+1, h)
    ws.row_dimensions[2].height = 36

    order = sorted(range(N), key=lambda i: rank_mm[i])
    for i, idx in enumerate(order):
        r  = 3 + i
        bg = ROW_A if i % 2 == 0 else ROW_B
        ist = labels[idx] in TARGET_SECTORS
        ws.row_dimensions[r].height = 23

        sc(ws, r, 1, rank_mm[idx],                         bg=bg, bold=True)
        tier_cell(ws, r, 2, bins_mm[idx])
        tier_cell(ws, r, 3, tiers_mm[idx])
        sc(ws, r, 4, labels[idx],                          bg=KEY_GREEN if ist else bg, bold=True, halign="left")
        sc(ws, r, 5, round(mm_scores[idx], 4), "0.0000",  bg=bg, bold=True)
        sc(ws, r, 6, round(z_scores[idx], 4),  "0.0000",  bg=bg, bold=True)
        tier_cell(ws, r, 7, bins_z[idx])
        tier_cell(ws, r, 8, tiers_z[idx])
        sc(ws, r, 9, round(mm_nl[idx], 3),     "0.000",   bg=bg)
        sc(ws, r, 10, round(mm_ms[idx], 3),     "0.000",   bg=bg)
        sc(ws, r, 11, nace[idx],                            bg=bg)

    # C201 qualitative note
    r_q = 3 + N + 1
    ws.merge_cells(f"A{r_q}:{get_column_letter(NC)}{r_q}")
    c = ws.cell(r_q, 1,
        "C20.1 (Fertilisers): Excluded from quantitative analysis. NACE C20.1 covers basic chemicals broadly "
        "(dyes, plastics, gases) — fertilisers = 8–15% of output. EU fertiliser production qualitatively highly "
        "concentrated (Yara, BASF, Borealis). Administrative Feasibility: qualitatively HIGH (Tier 5 / Bin 5 equivalent).")
    c.font = Font(name=FONT_NAME, italic=True, size=9, color="7D4E00")
    c.alignment = _al("left"); c.fill = _fill("FFF2CC"); c.border = _border()

    # Robustness note
    r_rob = r_q + 2
    ws.merge_cells(f"A{r_rob}:{get_column_letter(NC)}{r_rob}")
    c = ws.cell(r_rob, 1,
        f"Robustness: Spearman rank correlation between Min-Max and Z-score AF = {rho:.3f}. "
        f"{'All' if rho >= 0.99 else 'Nearly all'} sectors maintain "
        f"{'identical' if rho >= 0.99 else 'highly similar'} rank order. "
        "Bins: equal-width on [0,1] (absolute); Tiers: rank-based (relative). Both stable across normalization methods.")
    c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[r_rob].height = 32

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"


# ------------------------------------------------------------
# BUILD & SAVE
# ------------------------------------------------------------
def main():
    wb = Workbook()
    build_comparison(wb)
    build_ranking(wb)
    out = "../outputs/09_admin_feasibility_index.xlsx"

    # Apply unified visual styling for the replication package (see _styling.py)
    from _styling import apply_unified_styling
    apply_unified_styling(wb)

    wb.save(out)
    print(f"Saved -> {out}")
    print(f"\nSpearman ρ (MM vs Z) = {rho:.3f}")
    print(f"\nRanking (Min-Max composite, 50/50):\n")
    for idx in sorted(range(N), key=lambda i: rank_mm[i]):
        star = " ◄ GLM" if labels[idx] in TARGET_SECTORS else ""
        print(f"  {rank_mm[idx]:2}. {labels[idx]:<28} "
              f"N_Large={int(n_large[idx]):>4}  MktShare={mkt_sh[idx]:>5.1f}%  "
              f"AF={mm_scores[idx]:.4f}  Bin {bins_mm[idx]}  Tier {tiers_mm[idx]}{star}")

if __name__ == "__main__":
    main()
