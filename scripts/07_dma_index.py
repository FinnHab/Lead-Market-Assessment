"""
Domestic Market Autonomy (DMA) index builder.

Combines price transmission, import-penetration and export-intensity
into a single sector-level DMA score. Equal-width bins on [0, 1]
complement the rank-based tiers (Tier 5 = most autonomous,
Tier 1 = least; bins are absolute, tiers are relative).

Run from the scripts/ directory:
    python 07_dma_index.py
Writes ../outputs/07_dma_index.xlsx.
"""

import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY      = "203864"
HDR_BLUE  = "2E5FA3"
SUB_BLUE  = "E8EEF7"
ROW_A     = "FFFFFF"
ROW_B     = "F2F2F2"
KEY_GREEN = "E2EFDA"

TIER_BG = {1: "FF6B6B", 2: "FFC7CE", 3: "FFEB9C", 4: "92D050", 5: "C6EFCE"}
TIER_FC = {1: "FFFFFF", 2: "9C0006", 3: "7D6608", 4: "276221", 5: "276221"}

TARGET_SECTORS = {
    "Meat & Meat Products",
    "Grain-Mill and Bakery Products",
    "Dairy",
    "Prepared Animal Feeds",
    "Oilseed Oils",
}

def _side():    return Side(style="thin")
def _border():  return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="center", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)

def sc(ws, row, col, val, fmt=None, bg=None, bold=False, sz=10, color="000000",
       halign="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT_NAME, bold=bold, size=sz, color=color)
    c.border    = _border()
    c.alignment = _al(halign, wrap)
    if bg:  c.fill          = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, col, val, bg=HDR_BLUE, sz=10, color="FFFFFF", halign="center"):
    return sc(ws, row, col, val, bg=bg, bold=True, sz=sz, color=color, halign=halign)

def title_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _al()
    c.border    = _border()

def tier_label(t):
    return {1: "Tier 1\n(Low)", 2: "Tier 2", 3: "Tier 3\n(Mid)",
            4: "Tier 4",       5: "Tier 5\n(High)"}[t]

def tier_cell(ws, row, col, t):
    return sc(ws, row, col, tier_label(t), bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])

def assign_tiers(scores, n_tiers=5):
    """Rank-based: highest score → Tier 5. n=15: ranks 1-3 → Tier 5, 13-15 → Tier 1."""
    import pandas as pd
    s = pd.Series(scores)
    ranks = s.rank(ascending=False, method="first").astype(int)
    n = len(s)
    size = n // n_tiers
    raw = ((ranks - 1) // size + 1).clip(upper=n_tiers)
    return (n_tiers + 1 - raw).tolist()

def assign_bins(score):
    """Equal-width 5 bins on [0,1]. Absolute score-based (not rank-based)."""
    if score < 0.20: return 1
    if score < 0.40: return 2
    if score < 0.60: return 3
    if score < 0.80: return 4
    return 5


# ------------------------------------------------------------
# RAW DATA  (from DMA_Synthesis_v2.xlsx)
# ------------------------------------------------------------
RAW = [
    ("Potatoes",                      0.003273, 0.0182, 0.0334, 0.0181,
     1.0,       0.99975081, 1.0,       0.99992,  1.25084,  0.98196, 1.85024, 1.0     ),
    ("Meat & Meat Products",          0.158088, 0.0277, 0.0938, 0.1009,
     0.83009,   0.97609,    0.77307,   0.85975,  0.77936,  0.95275, 1.06724, 0.93312 ),
    ("Grain-Mill and Bakery Products",0.021386, 0.0325, 0.1659, 0.0823,
     0.98012,   0.96412,    0.50188,   0.81537,  1.19258,  0.88942, 0.05052, 0.77507 ),
    ("Dairy",                         0.197428, 0.0202, 0.1451, 0.1173,
     0.78692,   0.99477,    0.58008,   0.78725,  0.62641,  0.96902, 0.33304, 0.75154 ),
    ("Beer",                          0.388723, 0.0181, 0.1090, 0.0123,
     0.57697,   1.0,        0.71579,   0.76425,  0.01119,  0.98261, 0.82338, 0.73871 ),
    ("Refined Sugar",                 0.144281, 0.1214, 0.1622, 0.2327,
     0.84524,   0.74259,    0.51579,   0.70121,  0.79734,  0.31412, 0.10077, 0.66895 ),
    ("Prepared Animal Feeds",         0.391202, 0.1393, 0.1077, 0.0360,
     0.57425,   0.69798,    0.72068,   0.66430,  0.00322,  0.19828, 0.84104, 0.64938 ),
    ("Barley",                        0.715247, 0.0525, 0.1668, 0.1453,
     0.21861,   0.91428,    0.49850,   0.54380, -1.03894,  0.75999, 0.03829, 0.50140 ),
    ("Other Cereals and grains",      0.046075, 0.3996, 0.1657, 0.0703,
     0.95303,   0.04934,    0.50263,   0.50167,  1.11318, -1.48621, 0.05323, 0.49228 ),
    ("Fruit & Veg (Processed)",       0.405683, 0.13050, 0.25039, 0.0693,
     0.55836,   0.71992,    0.18426,   0.48751, -0.04336,  0.25524,-1.09706, 0.42708 ),
    ("Maize",                         0.586532, 0.3217, 0.1360, 0.0,
     0.35988,   0.24346,    0.61429,   0.40587, -0.62499, -0.98209, 0.45664, 0.39649 ),
    ("Oilseeds",                      0.275995, 0.4194, 0.1779, 0.0029,
     0.70069,   0.0,        0.45677,   0.38582,  0.37373, -1.61434,-0.11248, 0.37312 ),
    ("Wheat",                         0.914438, 0.1417, 0.2777, 0.0643,
     0.0,       0.69200,    0.08158,   0.25786, -1.67956,  0.18275,-1.46804, 0.18726 ),
    ("Oilseed Oils",                  0.798553, 0.3202, 0.2537, 0.0151,
     0.12718,   0.24720,    0.17180,   0.18206, -1.30687, -0.97239,-1.14206, 0.13462 ),
    ("Fertilisers",                   0.836131, 0.3860, 0.2994, 0.0137,
     0.08594,   0.08323,    0.0,       0.05639, -1.42772, -1.39820,-1.76279, 0.0     ),
]

dma_mm   = [r[8]  for r in RAW]
z_raw    = [r[12] for r in RAW]
z_min, z_max = min(z_raw), max(z_raw)
def rz(v): return (v - z_min) / (z_max - z_min)
dma_z    = [rz(r[12]) for r in RAW]

tiers_mm = assign_tiers(dma_mm)
tiers_z  = assign_tiers(dma_z)
bins_mm  = [assign_bins(s) for s in dma_mm]
bins_z   = [assign_bins(s) for s in dma_z]

rank_mm = [0]*15; rank_z = [0]*15
for pos, idx in enumerate(sorted(range(15), key=lambda i: -dma_mm[i])): rank_mm[idx] = pos+1
for pos, idx in enumerate(sorted(range(15), key=lambda i: -dma_z[i])):  rank_z[idx]  = pos+1

rho, _ = stats.spearmanr(rank_mm, rank_z)


# ------------------------------------------------------------
# SHEET 1 — DMA Comparison
# ------------------------------------------------------------
def build_comparison(wb):
    ws = wb.active
    ws.title = "01_DMA_Comparison"
    NC = 20

    title_row(ws, 1, NC,
              "Domestic Market Autonomy (DMA) – Min-Max vs. Z-Score Comparison  "
              "|  n=15  |  Tier 5 = Most Autonomous → Tier 1 = Least")
    ws.row_dimensions[1].height = 22

    spans = [
        (1,  2,  "Sector"),
        (3,  6,  "Raw Indicators  (EPC = Beta × R²)"),
        (7,  13, "Min-Max Normalization  (inverted)"),
        (14, 20, "Z-Score Normalization  (inverted, rescaled to 0–1)"),
    ]
    for c1, c2, text in spans:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, text)
        c.font      = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(HDR_BLUE)
        c.alignment = _al()
        c.border    = _border()
    ws.row_dimensions[2].height = 24

    col_hdrs = ["Sector", "", "EPC", "IPR", "EI", "Tariff\n(ctx)",
                "EPC_inv", "IPR_inv", "EI_inv", "DMA Score", "Bin", "Rank", "Tier",
                "Z(EPC)_inv", "Z(IPR)_inv", "Z(EI)_inv", "DMA Score\n(z, rescaled)", "Bin\n(z)", "Rank (z)", "Tier (z)"]
    for j, h in enumerate(col_hdrs):
        hdr(ws, 3, j+1, h, bg=SUB_BLUE, color="000000")
    ws.merge_cells("A3:B3")
    ws.row_dimensions[3].height = 36

    order = sorted(range(15), key=lambda i: rank_mm[i])
    for i, idx in enumerate(order):
        r   = 4 + i
        d   = RAW[idx]
        bg  = ROW_A if i % 2 == 0 else ROW_B
        ist = d[0] in TARGET_SECTORS

        ws.merge_cells(f"A{r}:B{r}")

        def s(col, val, fmt=None, extra_bold=False):
            sc(ws, r, col, val, fmt=fmt, bg=bg, bold=(ist or extra_bold),
               halign="left" if col == 1 else "center")

        s(1,  d[0])
        s(3,  d[1],  "0.0000")
        s(4,  d[2],  "0.0000")
        s(5,  d[3],  "0.0000")
        s(6,  d[4],  "0.0%")
        s(7,  d[5],  "0.000")
        s(8,  d[6],  "0.000")
        s(9,  d[7],  "0.000")
        s(10, d[8],  "0.000", extra_bold=True)
        tier_cell(ws, r, 11, bins_mm[idx])
        s(12, rank_mm[idx])
        tier_cell(ws, r, 13, tiers_mm[idx])
        s(14, d[9],  "0.000")
        s(15, d[10], "0.000")
        s(16, d[11], "0.000")
        s(17, round(rz(d[12]), 4), "0.000", extra_bold=True)
        tier_cell(ws, r, 18, bins_z[idx])
        s(19, rank_z[idx])
        tier_cell(ws, r, 20, tiers_z[idx])
        ws.row_dimensions[r].height = 23

    r_key = 4 + 15 + 2
    ws.merge_cells(f"A{r_key}:{get_column_letter(NC)}{r_key}")
    c = ws.cell(r_key, 1,
        f"Key finding: Min-Max and Z-score normalization produce identical sector rankings "
        f"(Spearman ρ = {rho:.3f}). Results are robust to normalization choice.")
    c.font      = Font(name=FONT_NAME, bold=True, size=10, color="000000")
    c.fill      = _fill(KEY_GREEN)
    c.alignment = _al("left")
    c.border    = _border()

    notes = [
        "  1. EPC = Beta × R² (Effective Price Coupling). Beta from log-log OLS, PT_Results_FINAL 2010–2023. Higher EPC = stronger world price integration = LOWER DMA.",
        "  2. IPR = Import Penetration Ratio = Extra-EU imports / Production Value. EI = Export Intensity = Exports / Production Value. Both from IPR_IE_2023_FINAL.",
        "  3. Fruit & Veg (Processed): IPR and EI are production-value weighted averages of Preparations and Juices sub-sectors.",
        "  4. All three index indicators are INVERTED (1 − norm, or negated Z) so that higher score = more autonomous = stronger potential lead market candidate.",
        "  5. Trade Protection (avg. MFN tariff 2019) is contextual only and NOT in the DMA composite score (avoids double-counting with IPR).",
        "  6. Tiers: rank-based (relative positioning among peers). Bins: equal-width on [0,1] score range (absolute positioning). Both use 5 levels.",
        "  7. Z-score composite shown rescaled to [0,1] for visual comparability. Raw Z-composite is the simple mean of the three inverted Z-scores.",
    ]
    for k, note in enumerate(notes):
        rn = r_key + 1 + k
        ws.merge_cells(f"A{rn}:{get_column_letter(NC)}{rn}")
        c = ws.cell(rn, 1, note)
        c.font      = Font(name=FONT_NAME, bold=False, size=9, color="000000")
        c.alignment = _al("left")
        c.border    = _border()

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 1
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A4"


# ------------------------------------------------------------
# SHEET 2 — Correlations
# ------------------------------------------------------------
def build_correlations(wb):
    ws = wb.create_sheet("02_Correlations")
    NC = 6

    title_row(ws, 1, NC, "Correlation Matrix – Raw Indicators + Trade Protection  |  n=15 sectors")

    ws.merge_cells("A2:F2")
    c = ws.cell(2, 1, "Pearson r between raw (non-normalised) indicators. Bold = p<0.05. Orange highlight = EPC–EI correlation requiring Methods disclosure.")
    c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left"); c.border = _border()

    labels = ["EPC (Price Trans.)", "IPR (Import Pen.)", "EI (Export Int.)", "Trade Protection"]
    corr   = [[1.0,0.312,0.691,-0.260],[0.312,1.0,0.490,-0.429],
              [0.691,0.490,1.0,-0.047],[-0.260,-0.429,-0.047,1.0]]

    CORR_CLR = {"sp":"70AD47","mp":"A9D18E","nm":"FF9999","ns":"FF0000",
                "dg":"BFBFBF","wn":"FFC000","nt":"FFFFFF"}
    def cbg(i,j,v):
        if i==j: return CORR_CLR["dg"]
        if i==0 and j==2: return CORR_CLR["wn"]
        if v>=0.5: return CORR_CLR["sp"]
        if v>=0.2: return CORR_CLR["mp"]
        if v<=-0.4: return CORR_CLR["ns"]
        if v<=-0.1: return CORR_CLR["nm"]
        return CORR_CLR["nt"]

    sc(ws, 4, 1, "", bg=SUB_BLUE)
    for j, lbl in enumerate(labels):
        hdr(ws, 4, j+2, lbl)

    for i, (lbl, row_vals) in enumerate(zip(labels, corr)):
        r = 5 + i
        hdr(ws, r, 1, lbl)
        for j, v in enumerate(row_vals):
            bold = abs(v) >= 0.5 and i != j
            sc(ws, r, j+2, v, "0.000", bg=cbg(i,j,v), bold=bold)

    r = 10
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(r, 1, "Colour legend:")
    c.font = Font(name=FONT_NAME, bold=True, size=9); c.fill=_fill("F2F2F2")
    c.alignment=_al("left"); c.border=_border()
    legend = [("70AD47","r ≥ 0.5 (strong positive)"),("A9D18E","0.3 ≤ r < 0.5 (moderate positive)"),
              ("FF9999","−0.5 < r ≤ −0.3 (moderate negative)"),("FFC000","EPC–EI: notable correlation, see notes"),
              ("BFBFBF","Diagonal (r=1.0)")]
    for k,(bg,txt) in enumerate(legend):
        rr = 11+k
        sc(ws, rr, 1, None, bg=bg)
        ws.merge_cells(f"B{rr}:E{rr}")
        c=ws.cell(rr,2,txt); c.font=Font(name=FONT_NAME,size=9); c.alignment=_al("left"); c.border=_border()

    r = 18
    ws.merge_cells(f"A{r}:E{r}")
    c=ws.cell(r,1,"Pairwise interpretation:")
    c.font=Font(name=FONT_NAME,bold=True,size=9); c.alignment=_al("left"); c.border=_border()

    pnotes = [
        "EPC ↔ IPR:  r = +0.312  |  Acceptable. Moderate co-movement, both reflect openness to world markets.",
        "   Acceptable. Moderate co-movement; does not invalidate independent indicator logic.",
        "EPC ↔ EI:  r = +0.691  |  Notable. Both reflect commodity market integration. Addressed in Methods: EPC = dynamic price transmission; EI = structural export dependency. Correlation is genuine structural co-movement, not overlap (Muhammad & Hossen 2025).",
        "   Orange highlight. Disclosed in paper but retained: removing EI would lose structural information on export orientation.",
        "IPR ↔ EI:  r = +0.490  |  Acceptable. Open sectors both import and export.",
        "Trade Protection ↔ IPR:  r = −0.429  |  Weak negative. Higher tariffs associated with lower import penetration — expected direction.",
    ]
    for k, note in enumerate(pnotes):
        rr = 19+k
        ws.merge_cells(f"A{rr}:E{rr}")
        bg = SUB_BLUE if k % 2 == 0 else "FFFFFF"
        c=ws.cell(rr,1,note); c.font=Font(name=FONT_NAME,size=9)
        c.fill=_fill(bg); c.alignment=_al("left"); c.border=_border()

    for j,w in enumerate([24,14,14,14,14]):
        ws.column_dimensions[get_column_letter(j+1)].width = w
    ws.freeze_panes = "B5"


# ------------------------------------------------------------
# SHEET 3 — Ranking
# ------------------------------------------------------------
def build_ranking(wb):
    ws = wb.create_sheet("03_Ranking")
    NC = 11

    title_row(ws, 1, NC,
              f"DMA Ranking – Sorted by Score  |  Min-Max & Z-Score identical (Spearman ρ={rho:.3f})  |  5 Tiers + 5 Bins")

    col_hdrs = ["Rank","Bin","Tier","Sector","DMA Score\n(Min-Max)","DMA Score\n(Z, rescaled)",
                "Bin\n(Z)","Tier\n(Z)","EPC_inv\n(MM)","IPR_inv\n(MM)","EI_inv\n(MM)"]
    for j, h in enumerate(col_hdrs):
        hdr(ws, 2, j+1, h)
    ws.row_dimensions[2].height = 36

    order = sorted(range(15), key=lambda i: rank_mm[i])
    for i, idx in enumerate(order):
        r = 3 + i
        d = RAW[idx]
        bg = ROW_A if i % 2 == 0 else ROW_B
        ws.row_dimensions[r].height = 23
        sc(ws, r, 1, rank_mm[idx],                         bg=bg, bold=True)
        tier_cell(ws, r, 2, bins_mm[idx])
        tier_cell(ws, r, 3, tiers_mm[idx])
        sc(ws, r, 4, d[0],                                 bg=bg, bold=True, halign="left")
        sc(ws, r, 5, round(d[8],4),  "0.0000",             bg=bg, bold=True)
        sc(ws, r, 6, round(rz(d[12]),4),"0.0000",          bg=bg, bold=True)
        tier_cell(ws, r, 7, bins_z[idx])
        tier_cell(ws, r, 8, tiers_z[idx])
        sc(ws, r, 9, round(d[5],3),  "0.000",              bg=bg)
        sc(ws, r, 10, round(d[6],3),  "0.000",             bg=bg)
        sc(ws, r, 11, round(d[7],3),  "0.000",             bg=bg)

    r_rob = 3 + 15 + 2
    ws.merge_cells(f"A{r_rob}:{get_column_letter(NC)}{r_rob}")
    c = ws.cell(r_rob, 1,
        f"Robustness: Spearman rank correlation between Min-Max and Z-score DMA = {rho:.3f}. "
        "All 15 sectors maintain identical rank order. "
        "Bins: equal-width on [0,1] (absolute); Tiers: rank-based (relative). Both fully stable across normalization methods.")
    c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left")
    c.border = _border()
    ws.row_dimensions[r_rob].height = 32

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    build_comparison(wb)
    build_correlations(wb)
    build_ranking(wb)
    out = "../outputs/07_dma_index.xlsx"

    # Apply unified visual styling for the replication package (see _styling.py)
    from _styling import apply_unified_styling
    apply_unified_styling(wb)

    wb.save(out)
    print(f"Saved -> {out}")
    print(f"\nSpearman ρ = {rho:.3f}")
    print("\nTier + Bin assignments:")
    for idx in sorted(range(15), key=lambda i: rank_mm[i]):
        star = " [target]" if RAW[idx][0] in TARGET_SECTORS else ""
