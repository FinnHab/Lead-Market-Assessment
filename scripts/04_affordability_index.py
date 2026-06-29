"""
Affordability index builder.

Log-transformed min-max normalisation, five rank-based tiers, CENTRAL
fossil-reference scenario ($458/t NH3 including 2025 EU ETS at 10%
effective pass-through). Tier algorithm matches the DMA and admin
feasibility indices.

Sheets:
    01_Affordability_Comparison   full normalisation (log MM + Z-score)
    02_Ranking                    sorted ranking
    03_Bins                       five-tier summary
    04_PRODCOM_Comparison         factory-gate vs retail green premium

Run from the scripts/ directory:
    python 04_affordability_index.py
Writes ../outputs/04_affordability_index.xlsx.
"""

import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY      = "203864"
HDR_BLUE  = "2E5FA3"
SUB_BLUE  = "E8EEF7"
ROW_A     = "FFFFFF"
ROW_B     = "F2F2F2"
KEY_GREEN = "E2EFDA"
WARN_ORG  = "FFC000"
N_TIERS   = 5

TIER_BG = {1: "FF6B6B", 2: "FFC7CE", 3: "FFEB9C", 4: "92D050", 5: "C6EFCE"}
TIER_FC = {1: "FFFFFF", 2: "9C0006", 3: "7D6608", 4: "276221", 5: "276221"}

TARGET_SECTORS = {
    "C101 – Meat Processing",
    "C106 – Grain Mill & Starch",
    "C105 – Dairy Processing",
    "C109 – Animal Feed (ICF)",
    "C104 – Oils & Fats",
}

def _side():    return Side(style="thin")
def _border():  return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="center", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)

def sc(ws, row, col, val, fmt=None, bg=None, bold=False, sz=10, color="000000", halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=FONT_NAME, bold=bold, size=sz, color=color)
    c.border = _border(); c.alignment = _al(halign)
    if bg: c.fill = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, col, val, bg=HDR_BLUE, color="FFFFFF"):
    return sc(ws, row, col, val, bg=bg, bold=True, color=color)

def title_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill = _fill(NAVY); c.alignment = _al(); c.border = _border()

def tier_label(t):
    return {1: "Tier 1\n(Low)", 2: "Tier 2", 3: "Tier 3\n(Mid)",
            4: "Tier 4",       5: "Tier 5\n(High)"}[t]

def tier_cell(ws, row, col, t):
    return sc(ws, row, col, tier_label(t), bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])

def assign_tiers(scores, n_tiers=N_TIERS):
    """Rank-based tiers matching the DMA index."""
    import pandas as pd
    ranks = pd.Series(scores).rank(ascending=False, method="first").astype(int)
    n = len(scores)
    size = n / n_tiers
    raw = [min(n_tiers, int((rk - 1) / size) + 1) for rk in ranks]
    return [n_tiers + 1 - t for t in raw]


# ------------------------------------------------------------
# RAW DATA (CENTRAL scenario)
# ------------------------------------------------------------
P = 0.806  # CENTRAL premium (central scenario: fossil $458 incl. ETS)

DATA = [
    ("C104 – Oils & Fats",         "Oilseed Oils",          0.02868, "Approach 1",  "Rapeseed Oil (1L)"),
    ("NACE 10.89 – Eggs",          "—",                     0.00894, "Approach 2",  "Eggs (10)"),
    ("C101 – Meat Processing",     "Meat & Meat Products",  0.01392, "Approach 2",  "Ø Chicken+Pork+Beef (PRODCOM value-wtd)"),
    ("C105 – Dairy Processing",    "Dairy",                 0.00760, "Approach 2",  "Milk (1L)"),
    ("C106 – Grain Mill & Starch", "Grain-Mill & Bakery",   0.00387, "Approach 1",  "Bread-only (NACE 1071)"),
    ("C109 – Animal Feed (ICF)",   "Prepared Animal Feeds", 0.00455, "Approach 2b", "ICF-only on end products (FEFAC wtd)"),
    ("C110 – Beer",                "Beer",                  0.00133, "Approach 1",  "Beer (0.5L)"),
    ("C108 – Sugar",               "Refined Sugar",         0.00050, "Approach 1",  "Carbonated Drink (1L)"),
]

N = len(DATA)
gp_cen = [d[2] for d in DATA]

# ---- LOG + MIN-MAX (inverted) ----
ln_gp  = [np.log(gp) for gp in gp_cen]
ln_min = min(ln_gp); ln_max = max(ln_gp)
mm_scores = [(ln_max - lg) / (ln_max - ln_min) for lg in ln_gp]

# ---- LOG + Z-SCORE (inverted, rescaled) ----
ln_mean = np.mean(ln_gp); ln_std = np.std(ln_gp, ddof=1)
z_raw    = [(lg - ln_mean) / ln_std for lg in ln_gp]
z_inv    = [-z for z in z_raw]
z_rmin   = min(z_inv); z_rmax = max(z_inv)
z_scores = [(z - z_rmin) / (z_rmax - z_rmin) for z in z_inv]

# ---- LINEAR (for comparison) ----
gp_min_lin = min(gp_cen); gp_max_lin = max(gp_cen)
mm_linear = [(gp_max_lin - gp) / (gp_max_lin - gp_min_lin) for gp in gp_cen]

# ---- RANKS & TIERS ----
rank_mm = [0]*N; rank_z = [0]*N
for pos, idx in enumerate(sorted(range(N), key=lambda i: -mm_scores[i])): rank_mm[idx] = pos+1
for pos, idx in enumerate(sorted(range(N), key=lambda i: -z_scores[i])):  rank_z[idx]  = pos+1
tiers_mm = assign_tiers(mm_scores)
tiers_z  = assign_tiers(z_scores)
rho, _ = stats.spearmanr(rank_mm, rank_z)


# ------------------------------------------------------------
# PRODCOM DATA (for Sheet 04)
# ------------------------------------------------------------
PRODCOM = [
    ("C104 – Oils & Fats",     1332,  6532e3,   7.42e9, "10412410+10415600", "Crude+refined rapeseed oil"),
    ("C101 – Meat Processing",  2692, 41210e3, 129.2e9, "1011+1012 primary", "Beef+pork+poultry cuts (excl. offals/fats)"),
    ("C105 – Dairy Processing", 2172, 77380e3, 133.0e9, "1051xxxx",          "All dairy (milk, cheese, yogurt, butter)"),
    ("C106 – Grain Mill",       1058, 96920e3, 159.3e9, "1061+1071+1072+73", "Flour+bread+pasta+biscuits+malt"),
    ("C108 – Sugar",             139, 17910e3,  13.6e9, "10811xxx",          "Raw beet + white sugar"),
    ("C109 – Animal Feed",      1891,159000e3,  62.6e9, "1091xxxx",          "Compound feed (ICF)"),
    ("C110 – Beer",              107, 30302e3,  31.9e9, "11051000",          "Beer"),
]

# Like-for-like retail GPs (matched to PRODCOM product level)
RETAIL_MATCHED = {
    "C104 – Oils & Fats":     (0.02868, "Rapeseed Oil (1L), €2.20/L"),
    "C101 – Meat Processing":  (0.01392, "Ø Chicken+Pork+Beef (retail, yield-corrected)"),
    "C105 – Dairy Processing": (0.00760, "Milk (1L), €1.544"),
    "C106 – Grain Mill":       (0.00387, "Bread-only (retail)"),
    "C108 – Sugar":            (0.00503, "Refined Sugar (1kg), €1.429"),
    "C109 – Animal Feed":      (0.03700, "Compound Feed recipe GP (Approach 3, FEFAC wtd)"),
    "C110 – Beer":             (0.00133, "Beer (0.5L), €1.041"),
}

# Compute PRODCOM GPs
prodcom_rows = []
for name, n_kt, output_t, val_eur, codes, note in PRODCOM:
    price_t = val_eur / output_t
    n_per_t = n_kt * 1e6 / output_t
    gp_p = n_per_t * P / price_t
    gp_r, prod_r = RETAIL_MATCHED[name]
    prodcom_rows.append((name, n_kt, output_t, price_t, n_per_t, gp_p, gp_r, prod_r, codes, note))

NP = len(prodcom_rows)
p_vals = [x[5] for x in prodcom_rows]; r_vals = [x[6] for x in prodcom_rows]
p_rank = [0]*NP; r_rank_p = [0]*NP
for pos, idx in enumerate(sorted(range(NP), key=lambda i: -p_vals[i])): p_rank[idx] = pos+1
for pos, idx in enumerate(sorted(range(NP), key=lambda i: -r_vals[i])): r_rank_p[idx] = pos+1
rho_pr, _ = stats.spearmanr(p_rank, r_rank_p)


# ------------------------------------------------------------
# SHEET 1 — Comparison
# ------------------------------------------------------------
def build_comparison(wb):
    ws = wb.active; ws.title = "01_Affordability_Comparison"
    NC = 16

    title_row(ws, 1, NC,
        "Affordability (Green Premium) – Log-Transformed Min-Max vs. Z-Score  "
        f"|  n={N}  |  Tier 5 = Most Affordable → Tier 1 = Least  |  CENTRAL scenario (+€{P:.3f}/kg N)")
    ws.row_dimensions[1].height = 22

    for c1, c2, text in [
        (1, 3, "Sector & Tier"), (4, 7, "Green Premium CEN (%)"),
        (8, 12, "Log-Transformed Min-Max (inverted)"),
        (13, 16, "Log-Transformed Z-Score (inv., rescaled 0–1)"),
    ]:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, text)
        c.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill = _fill(HDR_BLUE); c.alignment = _al(); c.border = _border()
    ws.row_dimensions[2].height = 24

    for j, h in enumerate([
        "Rank", "Tier", "NACE Sector",
        "GP CEN (%)", "ln(GP)", "Representative Product", "",
        "Score\n(log MM)", "Linear\n(no log)", "Rank", "Tier", "",
        "Z(ln GP)\nraw", "Score\n(Z, resc.)", "Rank\n(Z)", "Tier\n(Z)",
    ]):
        hdr(ws, 3, j+1, h, bg=SUB_BLUE, color="000000")
    ws.row_dimensions[3].height = 44

    order = sorted(range(N), key=lambda i: rank_mm[i])
    for i, idx in enumerate(order):
        r = 4 + i; d = DATA[idx]
        t = tiers_mm[idx]
        # Row background: light tint of tier color
        tier_tint = {
            1: "FFF0F0", 2: "FFF0F0", 3: "FFFBEF",
            4: "F0FAF0", 5: "F0FAF0",
        }[t]
        bg = tier_tint
        ist = d[0] in TARGET_SECTORS
        ws.row_dimensions[r].height = 23
        def s(col, val, fmt=None, extra_bold=False, halign="center"):
            sc(ws, r, col, val, fmt=fmt, bg=bg, bold=(ist or extra_bold), halign=halign)
        # Rank + Tier first
        sc(ws, r, 1, rank_mm[idx], bg=bg, bold=True, sz=11)
        tier_cell(ws, r, 2, t)
        s(3, d[0], halign="left")
        s(4, d[2], "0.00%", extra_bold=True); s(5, ln_gp[idx], "0.000")
        s(6, d[4], halign="left")
        s(8, mm_scores[idx], "0.0000", extra_bold=True); s(9, mm_linear[idx], "0.0000")
        s(10, rank_mm[idx]); tier_cell(ws, r, 11, tiers_mm[idx])
        s(13, z_raw[idx], "0.000")
        s(14, z_scores[idx], "0.0000", extra_bold=True); s(15, rank_z[idx])
        tier_cell(ws, r, 16, tiers_z[idx])

    # -- BINS SUMMARY (integrated) ---
    r_bins = 4 + N + 2
    ws.merge_cells(start_row=r_bins, start_column=1, end_row=r_bins, end_column=NC)
    c = ws.cell(r_bins, 1, "5-TIER BINNED SUMMARY")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill = _fill(NAVY); c.alignment = _al(); c.border = _border()
    ws.row_dimensions[r_bins].height = 22

    r_bins += 1
    for j, h in enumerate(["Tier", "Interpretation", "Sectors", "GP CEN Range",
                            "Score Range\n(log MM)", "n", "GLM Target?"]):
        hdr(ws, r_bins, j+1, h)
    ws.row_dimensions[r_bins].height = 36

    # Build tier data
    tier_data = {}
    for idx in sorted(range(N), key=lambda i: rank_mm[i]):
        t = tiers_mm[idx]
        if t not in tier_data:
            tier_data[t] = {"sectors": [], "gps": [], "scores": [], "targets": 0}
        tier_data[t]["sectors"].append(DATA[idx][0])
        tier_data[t]["gps"].append(DATA[idx][2])
        tier_data[t]["scores"].append(mm_scores[idx])
        if DATA[idx][0] in TARGET_SECTORS:
            tier_data[t]["targets"] += 1

    interp = {5: "Very High\nAffordability", 4: "High\nAffordability", 3: "Medium\nAffordability",
              2: "Low\nAffordability", 1: "Very Low\nAffordability"}

    r_bins += 1
    for t in [5, 4, 3, 2, 1]:
        if t not in tier_data:
            sc(ws, r_bins, 1, f"Tier {t}", bg=TIER_BG[t], bold=True, color=TIER_FC[t])
            sc(ws, r_bins, 2, interp[t], bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])
            for col in range(3, 8): sc(ws, r_bins, col, "—", bg=ROW_B)
            ws.row_dimensions[r_bins].height = 40; r_bins += 1; continue
        td = tier_data[t]; n_t = len(td["sectors"])
        sector_str = "\n".join(td["sectors"])
        gp_lo = min(td["gps"]); gp_hi = max(td["gps"])
        sc_lo = min(td["scores"]); sc_hi = max(td["scores"])
        gp_range = f"{gp_lo*100:.2f}% – {gp_hi*100:.2f}%" if n_t > 1 else f"{gp_lo*100:.2f}%"
        sc_range = f"{sc_lo:.3f} – {sc_hi:.3f}" if n_t > 1 else f"{sc_lo:.3f}"
        target_str = f"{td['targets']}/{n_t}" if td["targets"] > 0 else "—"

        ws.row_dimensions[r_bins].height = max(28, 20 * n_t)
        tier_cell(ws, r_bins, 1, t)
        sc(ws, r_bins, 2, interp[t], bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])
        sc(ws, r_bins, 3, sector_str, bg=ROW_A, halign="left")
        sc(ws, r_bins, 4, gp_range, bg=ROW_A)
        sc(ws, r_bins, 5, sc_range, bg=ROW_A)
        sc(ws, r_bins, 6, n_t, bg=ROW_A, bold=True)
        sc(ws, r_bins, 7, target_str, bg=KEY_GREEN if td["targets"] > 0 else ROW_A, bold=td["targets"] > 0)
        r_bins += 1

    # -- KEY FINDINGS ---
    r_key = r_bins + 1
    ws.merge_cells(f"A{r_key}:{get_column_letter(NC)}{r_key}")
    c = ws.cell(r_key, 1,
        f"Key finding: Log-transformed Min-Max and Z-score produce identical rankings "
        f"(Spearman ρ = {rho:.3f}). GLM target sectors span Tiers 1–4 — "
        f"affordability is a key differentiator across the 5 target nodes.")
    c.font = Font(name=FONT_NAME, bold=True, size=10); c.fill = _fill(KEY_GREEN)
    c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[r_key].height = 32

    for k, note in enumerate([
        "  1. Green Premium = % retail price increase from switching all synthetic N from grey to green ammonia. Lower GP = more affordable = higher score.",
        "  2. Log-transform: GP spans >1 order of magnitude (0.07%–3.80%). Linear Min-Max compresses 6/8 sectors into 0.70–0.97. Log-normalization is standard for right-skewed cost data.",
        "  3. Log Min-Max: score_i = (ln(GP_max) − ln(GP_i)) / (ln(GP_max) − ln(GP_min)). Inverted: lowest GP → score 1.0.",
        f"  4. Tiers: 5 tiers for n={N} (identical algorithm to DMA index). Tier 5 = most affordable. Tier 1 = least. Rows colored by tier.",
        "  5. Regulatory Reach basis for animal products. Rapeseed Oil 100% (no co-product allocation). See GP Calculations workbook.",
        "  6. C109: ICF-only N on animal end products (Approach 2b). Roughage and on-farm feed stay grey.",
        f"  7. Based on 03_affordability_results.xlsx: fossil reference $458/t (incl. 2025 EU ETS at 10% eff. rate). CENTRAL premium = EUR {P:.3f}/kg N.",
    ]):
        rn = r_key + 1 + k
        ws.merge_cells(f"A{rn}:{get_column_letter(NC)}{rn}")
        c = ws.cell(rn, 1, note)
        c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left"); c.border = _border()

    ws.column_dimensions["A"].width = 7; ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 30; ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 10; ws.column_dimensions["F"].width = 34
    ws.column_dimensions["G"].width = 2
    for cl in ["H","I","J","K","L","M","N","O","P"]:
        ws.column_dimensions[cl].width = 12
    ws.freeze_panes = "A4"


# ------------------------------------------------------------
# SHEET 2 — Ranking
# ------------------------------------------------------------
def build_ranking(wb):
    ws = wb.create_sheet("02_Ranking"); NC = 9

    title_row(ws, 1, NC,
        f"Affordability Ranking – Log-Transformed  |  ρ={rho:.3f}  |  5 Tiers")
    for j, h in enumerate(["Rank","Tier","NACE Sector","GP CEN (%)",
        "Afford. Score\n(log MM)","Afford. Score\n(Z, resc.)",
        "Score\nDifference","Linear Score\n(comparison)","Representative Product"]):
        hdr(ws, 2, j+1, h)
    ws.row_dimensions[2].height = 36

    for i, idx in enumerate(sorted(range(N), key=lambda i: rank_mm[i])):
        r = 3+i; d = DATA[idx]; bg = ROW_A if i%2==0 else ROW_B; ist = d[0] in TARGET_SECTORS
        ws.row_dimensions[r].height = 23
        sc(ws,r,1,rank_mm[idx],bg=bg,bold=True); tier_cell(ws,r,2,tiers_mm[idx])
        sc(ws,r,3,d[0],bg=bg,bold=ist,halign="left"); sc(ws,r,4,d[2],"0.00%",bg=bg,bold=True)
        sc(ws,r,5,round(mm_scores[idx],4),"0.0000",bg=bg,bold=True)
        sc(ws,r,6,round(z_scores[idx],4),"0.0000",bg=bg,bold=True)
        sc(ws,r,7,round(abs(mm_scores[idx]-z_scores[idx]),4),"0.0000",bg=bg)
        sc(ws,r,8,round(mm_linear[idx],4),"0.0000",bg=bg)
        sc(ws,r,9,d[4],bg=bg,halign="left")

    r_rob = 3+N+2
    ws.merge_cells(f"A{r_rob}:I{r_rob}")
    c = ws.cell(r_rob, 1,
        f"Robustness: Spearman ρ = {rho:.3f}. All {N} sectors identical rank order. "
        "'Linear Score' shows the compression problem (ranks 2–7 in 0.70–0.97).")
    c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left")
    c.border = _border(); ws.row_dimensions[r_rob].height = 32

    ws.column_dimensions["A"].width = 7;  ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30; ws.column_dimensions["D"].width = 12
    for cl in ["E","F","G","H"]: ws.column_dimensions[cl].width = 14
    ws.column_dimensions["I"].width = 38
    ws.freeze_panes = "A3"


# ------------------------------------------------------------
# SHEET 3 — Bins
# ------------------------------------------------------------
def build_bins(wb):
    ws = wb.create_sheet("03_Bins"); NC = 7

    title_row(ws, 1, NC,
        "Affordability – 5-Tier Binned Index  |  Tier 5 = Most Affordable → Tier 1 = Least")
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(["Tier","Interpretation","Sectors","GP CEN Range",
        "Score Range\n(log MM)","n","GLM Target?"]):
        hdr(ws, 2, j+1, h)
    ws.row_dimensions[2].height = 36

    tier_data = {}
    for idx in sorted(range(N), key=lambda i: rank_mm[i]):
        t = tiers_mm[idx]
        if t not in tier_data:
            tier_data[t] = {"sectors": [], "gps": [], "scores": [], "targets": 0}
        tier_data[t]["sectors"].append(DATA[idx][0])
        tier_data[t]["gps"].append(DATA[idx][2])
        tier_data[t]["scores"].append(mm_scores[idx])
        if DATA[idx][0] in TARGET_SECTORS:
            tier_data[t]["targets"] += 1

    interp = {5:"Very High\nAffordability", 4:"High\nAffordability", 3:"Medium\nAffordability",
              2:"Low\nAffordability", 1:"Very Low\nAffordability"}

    r = 3
    for t in [5, 4, 3, 2, 1]:
        if t not in tier_data:
            sc(ws,r,1,f"Tier {t}",bg=TIER_BG[t],bold=True,color=TIER_FC[t])
            sc(ws,r,2,interp[t],bg=TIER_BG[t],bold=True,sz=9,color=TIER_FC[t])
            for col in range(3,8): sc(ws,r,col,"—",bg=ROW_B)
            ws.row_dimensions[r].height = 40; r += 1; continue

        td = tier_data[t]; n_t = len(td["sectors"])
        sector_str = "\n".join(td["sectors"])
        gp_lo = min(td["gps"]); gp_hi = max(td["gps"])
        sc_lo = min(td["scores"]); sc_hi = max(td["scores"])
        gp_range = f"{gp_lo*100:.2f}% – {gp_hi*100:.2f}%" if n_t > 1 else f"{gp_lo*100:.2f}%"
        sc_range = f"{sc_lo:.3f} – {sc_hi:.3f}" if n_t > 1 else f"{sc_lo:.3f}"
        target_str = f"{td['targets']}/{n_t}" if td["targets"] > 0 else "—"

        ws.row_dimensions[r].height = max(28, 20 * n_t)
        tier_cell(ws, r, 1, t)
        sc(ws,r,2,interp[t],bg=TIER_BG[t],bold=True,sz=9,color=TIER_FC[t])
        sc(ws,r,3,sector_str,bg=ROW_A,halign="left")
        sc(ws,r,4,gp_range,bg=ROW_A)
        sc(ws,r,5,sc_range,bg=ROW_A)
        sc(ws,r,6,n_t,bg=ROW_A,bold=True)
        sc(ws,r,7,target_str,bg=KEY_GREEN if td["targets"]>0 else ROW_A,bold=td["targets"]>0)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    c = ws.cell(r, 1,
        f"Tier algorithm identical to DMA index: ranks divided into {N_TIERS} equal-sized bins. "
        f"For n={N}: bin sizes 2–2–1–2–1 (from Tier 5 to Tier 1).")
    c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[r].height = 28

    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    c = ws.cell(r, 1,
        "GLM target sectors span Tiers 1–4: C109 and C106 (Tier 4), C105 (Tier 3), "
        "C101 (Tier 2), C104 (Tier 1). Affordability is a key differentiator across the 5 target nodes.")
    c.font = Font(name=FONT_NAME, bold=True, size=9); c.fill = _fill(KEY_GREEN)
    c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[r].height = 32

    ws.column_dimensions["A"].width = 12; ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 34; ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18; ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 14
    ws.freeze_panes = "A3"


# ------------------------------------------------------------
# SHEET 4 — PRODCOM Comparison
# ------------------------------------------------------------
def build_prodcom(wb):
    ws = wb.create_sheet("03_PRODCOM_Comparison")
    NC = 15

    title_row(ws, 1, NC,
        f"PRODCOM Factory-Gate GP vs. Retail GP (like-for-like)  |  CENTRAL scenario  |  Spearman ρ = {rho_pr:.3f}")
    ws.row_dimensions[1].height = 22

    for c1, c2, text in [
        (1, 2, "Sector"), (3, 8, "PRODCOM Factory-Gate"),
        (9, 11, "Retail (matched product level)"), (12, 15, "Comparison"),
    ]:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1, text)
        c.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
        c.fill = _fill(HDR_BLUE); c.alignment = _al(); c.border = _border()
    ws.row_dimensions[2].height = 24

    for j, h in enumerate([
        "NACE Sector", "PRODCOM Codes",
        "Emb. N (kt)", "Output (Mt)", "Price (€/t)", "N/t (kg)", "GP PRODCOM (%)", "Rank (P)",
        "GP Retail (%)", "Rank (R)", "Product (Retail, matched)",
        "Ratio\nP / R", "Rank\nShift", "Driver of difference", "",
    ]):
        hdr(ws, 3, j+1, h, bg=SUB_BLUE, color="000000")
    ws.row_dimensions[3].height = 44

    drivers = {
        "C104 – Oils & Fats":     "Crude oil (1,136€/t) vs retail (2,418€/t). Both measure rapeseed oil.",
        "C101 – Meat Processing":  "PRODCOM avg all cuts (3,135€/t) vs retail premium cuts (~10,875€/t).",
        "C105 – Dairy Processing": "PRODCOM avg all dairy (1,719€/t) incl. cheese; retail = fresh milk only.",
        "C106 – Grain Mill":       "PRODCOM incl. high-value bakery (1,644€/t); retail = bread-only.",
        "C108 – Sugar":            "Both = refined sugar. PRODCOM 762€/t vs retail 1,429€/t. Pure markup.",
        "C109 – Animal Feed":      "Both = compound feed level. PRODCOM 394€/t vs recipe-weighted cost.",
        "C110 – Beer":             "PRODCOM 1,054€/t vs retail ~2,082€/t. Standard markup.",
    }

    order = sorted(range(NP), key=lambda i: p_rank[i])
    for i, idx in enumerate(order):
        r = 4 + i
        name, n_kt, output_t, price_t, n_per_t, gp_p, gp_r, prod_r, codes, note = prodcom_rows[idx]
        bg = ROW_A if i % 2 == 0 else ROW_B
        ws.row_dimensions[r].height = 28

        ratio = gp_p / gp_r if gp_r > 0 else 0
        shift = r_rank_p[idx] - p_rank[idx]
        shift_str = f"{'↑' if shift > 0 else '↓'}{abs(shift)}" if shift != 0 else "="
        is_shift = abs(shift) >= 2

        sc(ws, r, 1, name, bg=bg, bold=True, halign="left")
        sc(ws, r, 2, codes, bg=bg, halign="left", sz=9)
        sc(ws, r, 3, n_kt, "#,##0", bg=bg)
        sc(ws, r, 4, output_t / 1e6, "0.0", bg=bg)
        sc(ws, r, 5, price_t, "#,##0", bg=bg, bold=True)
        sc(ws, r, 6, n_per_t, "0.0", bg=bg)
        sc(ws, r, 7, gp_p, "0.00%", bg=bg, bold=True)
        sc(ws, r, 8, p_rank[idx], bg=bg, bold=True)
        sc(ws, r, 9, gp_r, "0.00%", bg=bg, bold=True)
        sc(ws, r, 10, r_rank_p[idx], bg=bg, bold=True)
        sc(ws, r, 11, prod_r, bg=bg, halign="left", sz=9)
        sc(ws, r, 12, ratio, "0.0×", bg=bg)
        sc(ws, r, 13, shift_str, bg=WARN_ORG if is_shift else bg, bold=is_shift)
        sc(ws, r, 14, drivers.get(name, ""), bg=bg, halign="left", sz=9)

    # Key finding
    r = 4 + NP + 1
    ws.merge_cells(f"A{r}:O{r}")
    c = ws.cell(r, 1,
        f"Spearman ρ = {rho_pr:.3f} between PRODCOM and retail rankings (like-for-like). "
        "5/7 sectors identical rank. Remaining shifts (±1) reflect factory-gate vs. retail price differences, not product mismatch.")
    c.font = Font(name=FONT_NAME, bold=True, size=10); c.fill = _fill(KEY_GREEN)
    c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[r].height = 36

    # Reference: end-product GPs
    r += 2
    ws.merge_cells(f"A{r}:O{r}")
    c = ws.cell(r, 1, "REFERENCE: End-product retail GPs used in the primary Affordability index (Sheets 01–03)")
    c.font = Font(name=FONT_NAME, bold=True, size=10, color=HDR_BLUE)
    c.alignment = _al("left"); c.border = _border()
    r += 1
    sc(ws, r, 1, "C108 – Sugar", bg=SUB_BLUE, halign="left", bold=True)
    sc(ws, r, 2, "Carbonated Drink (1L)", bg=SUB_BLUE, halign="left")
    sc(ws, r, 3, 0.00050, "0.00%", bg=SUB_BLUE, bold=True)
    sc(ws, r, 4, "← used in index (end product)", bg=SUB_BLUE, halign="left", sz=9)
    ws.merge_cells(f"D{r}:O{r}")
    r += 1
    sc(ws, r, 1, "C109 – Animal Feed", bg=SUB_BLUE, halign="left", bold=True)
    sc(ws, r, 2, "ICF-only on end products", bg=SUB_BLUE, halign="left")
    sc(ws, r, 3, 0.00455, "0.00%", bg=SUB_BLUE, bold=True)
    sc(ws, r, 4, "← used in index (Approach 2b: ICF N on animal end products)", bg=SUB_BLUE, halign="left", sz=9)
    ws.merge_cells(f"D{r}:O{r}")

    # Notes
    r += 2
    for k, note in enumerate([
        "  1. Like-for-like: C108 = Refined Sugar GP (1.00%) in both columns. C109 = Compound Feed recipe GP (5.37%, Approach 3). Matched product levels for fair comparison.",
        "  2. Primary index (Sheets 01–03) uses end-product retail GPs: Carbonated Drink for C108, ICF-only on end products for C109. These are the consumer-facing measures.",
        "  3. PRODCOM prices are factory-gate (ex-works), excluding retail margins, distribution, packaging, VAT.",
        "  4. C106 is the only sector where PRODCOM GP < Retail GP: PRODCOM includes high-value bakery in the NACE 106 aggregate.",
        "  5. Eggs (NACE 10.89) excluded: shell eggs = NACE 01.47 (primary agriculture), not in PRODCOM manufacturing.",
        "  6. PRODCOM GP = direct cost pressure at regulated node. Retail GP = consumer-facing affordability. High ρ confirms consistent rankings.",
    ]):
        rn = r + k
        ws.merge_cells(f"A{rn}:O{rn}")
        c = ws.cell(rn, 1, note)
        c.font = Font(name=FONT_NAME, size=9); c.alignment = _al("left"); c.border = _border()

    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10; ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 11; ws.column_dimensions["F"].width = 9
    ws.column_dimensions["G"].width = 13; ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 12; ws.column_dimensions["J"].width = 8
    ws.column_dimensions["K"].width = 32; ws.column_dimensions["L"].width = 9
    ws.column_dimensions["M"].width = 8; ws.column_dimensions["N"].width = 46
    ws.column_dimensions["O"].width = 2
    ws.freeze_panes = "A4"


# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
def main():
    wb = Workbook()
    build_comparison(wb)
    build_ranking(wb)
    # Bins now integrated into Sheet 01 — no standalone bins sheet
    build_prodcom(wb)
    out = "../outputs/04_affordability_index.xlsx"

    # Apply unified visual styling for the replication package (see _styling.py)
    from _styling import apply_unified_styling
    apply_unified_styling(wb)

    wb.save(out)
    print(f"Saved -> {out}")
    print(f"\nSpearman ρ (log MM vs Z) = {rho:.3f}")
    print(f"Spearman ρ (PRODCOM vs Retail, like-for-like) = {rho_pr:.3f}")
    print(f"\nAffordability Ranking (log-transformed, CENTRAL):")
    print(f"{'Rk':<4} {'Tier':<5} {'Sector':<30} {'GP CEN':>8} {'Log MM':>9} {'Linear':>9}")
    print("-" * 68)
    for idx in sorted(range(N), key=lambda i: rank_mm[i]):
        star = " ◄" if DATA[idx][0] in TARGET_SECTORS else ""
        print(f"  {rank_mm[idx]:<2}  T{tiers_mm[idx]}   {DATA[idx][0]:<30} {gp_cen[idx]*100:>7.2f}%"
              f"  {mm_scores[idx]:.4f}   {mm_linear[idx]:.4f}{star}")
    print(f"\nPRODCOM vs Retail (like-for-like):")
    print(f"{'Sector':<24} {'PRODCOM':>9} {'Rk':>3}  {'Retail':>9} {'Rk':>3}  {'Shift':>5}")
    print("-" * 58)
    for idx in sorted(range(NP), key=lambda i: p_rank[i]):
        name = prodcom_rows[idx][0]; gp_p = prodcom_rows[idx][5]; gp_r = prodcom_rows[idx][6]
        shift = r_rank_p[idx] - p_rank[idx]
        arrow = f"{'↑' if shift>0 else '↓'}{abs(shift)}" if shift != 0 else "="
        print(f"  {name:<22} {gp_p*100:>7.2f}%  {p_rank[idx]:>2}   {gp_r*100:>7.2f}%  {r_rank_p[idx]:>2}   {arrow:>4}")

if __name__ == "__main__":
    main()
