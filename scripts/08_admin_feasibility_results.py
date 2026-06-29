"""
SBS Market Concentration – Formatted Raw Data
==============================================
Reformats the raw Eurostat SBS data into GLM project design.
Includes C10.8 (Manufacture of other food products) with derived values.
C20.1 flagged as qualitative-only.
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY      = "203864"
HDR_BLUE  = "2E5FA3"
SUB_BLUE  = "E8EEF7"
ROW_A     = "FFFFFF"
ROW_B     = "F2F2F2"
KEY_GREEN = "E2EFDA"
WARN_BG   = "FFF2CC"

def _side():    return Side(style="thin")
def _border():  return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="center", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)

def sc(ws, row, col, val, fmt=None, bg=None, bold=False, sz=10, color="000000",
       halign="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT_NAME, bold=bold, size=sz, color=color)
    c.border    = _border()
    c.alignment = _al(halign, wrap)
    if bg:  c.fill          = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, col, val, bg=HDR_BLUE, sz=10, color="FFFFFF", halign="center"):
    return sc(ws, row, col, val, bg=bg, bold=True, sz=sz, color=color, halign=halign)

def title_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _al()
    c.border    = _border()

TARGET_SECTORS = {
    "Processing and preserving of meat and production of meat products",
    "Manufacture of dairy products",
    "Manufacture of vegetable and animal oils and fats",
    "Manufacture of grain mill products, starches and starch products",
    "Manufacture of prepared animal feeds",
}

# ------------------------------------------------------------
# RAW DATA  (Eurostat SBS sbs_sc_ovw, 2022, EU-27)
# ------------------------------------------------------------
# (sector, nace, year, n_total, n_large, pv_total_meur, pv_large_meur,
#  share_count_pct, mkt_share_pct, persons_total, persons_large, note)
DATA = [
    ("Processing and preserving of meat and production of meat products", "C10.1", 2022, 33086,  526, 252853.75, 157336.60,  1.59, 62.22, None, None, ""),
    ("Processing and preserving of fruit and vegetables",     "C10.3", 2022, 14000,  210,  83285.30,  47022.61,  1.50, 56.46, None, None, ""),
    ("Manufacture of vegetable and animal oils and fats",            "C10.4", 2022,  8400,   42,  84021.29,  45856.03,  0.50, 54.58, None, None, ""),
    ("Manufacture of dairy products",         "C10.5", 2022, 13050,  270, 200000.00, 146342.50,  2.07, 73.17, None, None, ""),
    ("Manufacture of grain mill products, starches and starch products",         "C10.6", 2022,  6001,   70,  51389.83,  23391.01,  1.17, 45.52, None, None, ""),
    ("Manufacture of bakery and farinaceous products",        "C10.7", 2022,152617,  600, 136183.99,  60883.43,  0.39, 44.71, None, None, ""),
    ("Manufacture of other food products",    "C10.8", 2022, 37571,  437, 199000.00, 135316.40,  1.16, 68.00, 644000, 332795,
     "PV 250+ derived by subtraction (confidential in SBS). Persons employed available."),
    ("Manufacture of prepared animal feeds",  "C10.9", 2022,  6000,   89,  90000.00,  38187.20,  1.48, 42.43, None, None, ""),
    ("Manufacture of beverages",              "C11.0", 2022, 35202,  260, 174020.40, 109699.39,  0.74, 63.04, None, None, ""),
    ("Manufacture of basic chemicals, fertilisers and nitrogen compounds, plastics and synthetic rubber in primary forms","C20.1", 2022,  8410,  360, 432367.94, 334312.37,  4.28, 77.32, None, None,
     "NACE 2-digit too broad: fertilisers = 8–15% of C20.1 output. Excluded from quantitative AF index."),
]

N = len(DATA)

wb = Workbook()
NC = 12

# ------------------------------------------------------------
# SHEET 1 — FULL DATA TABLE
# ------------------------------------------------------------
ws = wb.active
ws.title = "SBS_Data"
ws.sheet_properties.tabColor = "2E5FA3"

title_row(ws, 1, NC,
          "SBS Market Concentration – Raw Data  |  Eurostat sbs_sc_ovw  |  EU-27  |  Reference Year 2022")
ws.row_dimensions[1].height = 22

# Section spans
spans = [
    (1, 3, "Sector"),
    (4, 5, "Enterprise Count"),
    (6, 7, "Production Value (M€)"),
    (8, 9, "Concentration Metrics"),
    (10, 12, "Context"),
]
for c1, c2, text in spans:
    ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
    c = ws.cell(2, c1, text)
    c.font      = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    c.fill      = _fill(HDR_BLUE)
    c.alignment = _al()
    c.border    = _border()
ws.row_dimensions[2].height = 24

# Column headers
col_hdrs = [
    "Sector", "NACE\nCode", "Year",
    "Enterprises\n(Total)", "Large Enterpr.\n(≥250 emp.)",
    "PV Total\n(M€)", "PV Large\n(M€)",
    "Share Large\nin Count (%)", "Market Share\nLarge in PV (%)",
    "Persons Empl.\n(Total)", "Persons Empl.\n(Large)", "Notes"
]
for ci, h in enumerate(col_hdrs, 1):
    hdr(ws, 3, ci, h, bg=SUB_BLUE, color="000000")
ws.row_dimensions[3].height = 38

# Data rows
for idx in range(N):
    r  = 4 + idx
    d  = DATA[idx]
    bg = ROW_A if idx % 2 == 0 else ROW_B
    ist = d[0] in TARGET_SECTORS
    is_c201 = d[1] == "C20.1"
    row_bg = WARN_BG if is_c201 else (KEY_GREEN if ist else bg)

    sc(ws, r, 1,  d[0],   bg=row_bg, bold=ist, halign="left")
    sc(ws, r, 2,  d[1],   bg=row_bg, bold=ist)
    sc(ws, r, 3,  d[2],   bg=row_bg, bold=ist)
    sc(ws, r, 4,  d[3],   bg=row_bg, bold=ist, fmt="#,##0")
    sc(ws, r, 5,  d[4],   bg=row_bg, bold=ist, fmt="#,##0")
    sc(ws, r, 6,  d[5],   bg=row_bg, bold=ist, fmt="#,##0.00")
    sc(ws, r, 7,  d[6],   bg=row_bg, bold=ist, fmt="#,##0.00")
    sc(ws, r, 8,  d[7],   bg=row_bg, bold=ist, fmt="0.00\"%\"")
    sc(ws, r, 9,  d[8],   bg=row_bg, bold=ist, fmt="0.00\"%\"")
    sc(ws, r, 10, d[9] if d[9] else "–",  bg=row_bg, bold=ist, fmt="#,##0" if d[9] else None)
    sc(ws, r, 11, d[10] if d[10] else "–", bg=row_bg, bold=ist, fmt="#,##0" if d[10] else None)
    sc(ws, r, 12, d[11] if d[11] else "",  bg=row_bg, bold=False, halign="left", sz=9,
       color="7D4E00" if is_c201 else "000000")
    ws.row_dimensions[r].height = 23

# Source & notes section
r_n = 4 + N + 1
notes = [
    "DATA SOURCE & NOTES",
    "  Source: Eurostat, Enterprise statistics by size class and NACE Rev. 2 activity [sbs_sc_ovw], extracted 13/11/2025.",
    "  Coverage: EU-27 (from 2020), reference year 2022. Some 2022 values flagged as estimated (e) in source.",
    "  Size class: \"Large\" = 250 or more persons employed.",
    "  Production Value (PV): Value of output in million euro. Market Share = PV Large / PV Total × 100.",
    "  Share in Count: N Large / N Total × 100.",
    "",
    "  C10.8 (Manufacture of other food products): PV for 250+ enterprises is confidential (suppressed) in Eurostat.",
    "     Derived by subtraction: PV(250+) = PV(Total) − PV(0–9) − PV(10–19) − PV(20–49) − PV(50–249).",
    "     = 199,000 − 6,383.6 − 5,000 − 12,400 − 39,900 = 135,316.4 M€.  Market Share = 68.0%.",
    "     N Large (437) is an estimate (Eurostat flag 'e'). Persons employed data available for C10.8.",
    "",
    "  C20.1 (Manufacture of basic chemicals, fertilisers and nitrogen compounds, plastics and synthetic rubber in primary forms): NACE C20.1 covers basic chemicals, fertilisers AND nitrogen compounds,",
    "     plastics in primary forms, synthetic rubber, industrial gases, dyes — fertilisers represent only 8–15%",
    "     of sector output. SBS values therefore describe the broader chemical industry, not fertilisers specifically.",
    "     NACE C20.15 (Fertilisers & N compounds) exists at 4-digit level but has insufficient EU-wide SBS coverage.",
    "     → Excluded from quantitative Administrative Feasibility index. Treated qualitatively in the paper.",
    "",
    "  Green highlight = GLM target sector (5 processing nodes). Yellow highlight = excluded from AF quantification.",
]
for i, note in enumerate(notes):
    rn = r_n + i
    ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=NC)
    c = ws.cell(rn, 1, note)
    c.font = Font(name=FONT_NAME, bold=(i == 0), size=9, color=NAVY)
    c.alignment = _al("left", wrap=False)

# Column widths
widths = [55, 9, 7, 13, 13, 14, 14, 12, 13, 13, 13, 45]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A4"

# ------------------------------------------------------------
# SHEET 2 — SUMMARY (compact, sorted by Mkt Share)
# ------------------------------------------------------------
ws2 = wb.create_sheet("Summary_by_MktShare")
ws2.sheet_properties.tabColor = "E2EFDA"
NC2 = 6

title_row(ws2, 1, NC2, "Market Concentration Summary – Sorted by Market Share of Large Enterprises (descending)")

hdrs2 = ["Rank", "Sector", "NACE", "N Large\n(≥250)", "Mkt Share\nLarge (%)", "Status"]
for ci, h in enumerate(hdrs2, 1):
    hdr(ws2, 2, ci, h, bg=SUB_BLUE, color="000000")
ws2.row_dimensions[2].height = 36

# Sort by market share descending
order = sorted(range(N), key=lambda i: -DATA[i][8])
for pos, idx in enumerate(order):
    r  = 3 + pos
    d  = DATA[idx]
    bg = ROW_A if pos % 2 == 0 else ROW_B
    ist = d[0] in TARGET_SECTORS
    is_c201 = d[1] == "C20.1"

    sc(ws2, r, 1, pos + 1, bg=bg, bold=True)
    row_bg = WARN_BG if is_c201 else (KEY_GREEN if ist else bg)
    sc(ws2, r, 2, d[0], bg=row_bg, bold=ist, halign="left")
    sc(ws2, r, 3, d[1], bg=bg)
    sc(ws2, r, 4, d[4], bg=bg, bold=ist, fmt="#,##0")
    sc(ws2, r, 5, d[8], bg=bg, bold=ist, fmt="0.00\"%\"")

    if is_c201:
        status = "Excluded (qualitative)"
        s_col = "7D4E00"
    elif ist:
        status = "GLM target sector"
        s_col = "276221"
    else:
        status = "Reference sector"
        s_col = "808080"
    sc(ws2, r, 6, status, bg=bg, color=s_col, sz=9, halign="left")
    ws2.row_dimensions[r].height = 23

widths2 = [7, 55, 9, 11, 12, 22]
for ci, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = "A3"

# ------------------------------------------------------------
# SAVE
# ------------------------------------------------------------
OUT = "../outputs/08_admin_feasibility_results.xlsx"

# Apply unified visual styling for the replication package (see _styling.py)
from _styling import apply_unified_styling
apply_unified_styling(wb)

wb.save(OUT)
print(f"Saved -> {OUT}")
print(f"\n{N} sectors, sorted by Market Share Large:")
for pos, idx in enumerate(order):
    d = DATA[idx]
    star = " [GLM]" if d[0] in TARGET_SECTORS else (" [excl]" if d[1] == "C20.1" else "")
    print(f"  {pos+1:2}. {d[0]:<28} {d[1]:>6}  N_Large={d[4]:>4}  MktShare={d[8]:>5.1f}%{star}")
