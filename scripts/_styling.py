"""Shared styling for the replication-package result workbooks.

Call apply_unified_styling(wb) immediately before wb.save(...). The function is
non-destructive with respect to numerical data, formulas, and sheet structure;
it only normalises visual presentation:

  - Dark-blue (#1F4E79) section banners with white text for all bold header rows
  - Single-cell long-text rows (titles, sub-section labels, notes) merged across
    the used column range
  - Thin grey (#D9D9D9) borders on every cell in the used range
  - Sensible column widths based on data content (capped at 35 characters)
  - Wrap-text only where a cell's value would otherwise overflow its column
  - Tab colours cleared
  - Stray '=' prefixes on non-formula strings neutralised

Dependencies: openpyxl >= 3.0.

The styling palette is deliberately conservative: a single dark-blue tone for
all header levels, black text on white for data, italic grey for notes.
"""

from __future__ import annotations

import re
from copy import copy

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
DATA_FONT    = Font(name="Calibri", size=11, bold=False, color="000000")
EMPH_FONT    = Font(name="Calibri", size=11, bold=True, color="000000")
NOTE_FONT    = Font(name="Calibri", size=10, italic=True, color="595959")
BLANK_FILL   = PatternFill(fill_type=None)

_THIN_SIDE = Side(border_style="thin", color="D9D9D9")
DATA_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                     top=_THIN_SIDE, bottom=_THIN_SIDE)

_SECTION_SOURCE_FILLS = {
    "FF2C5282", "002C5282", "FF1F4E79", "001F4E79",
    "FF0D2137", "000D2137", "002E5FA3", "FF2E5FA3",
    "00203864", 
}
_SUBHDR_SOURCE_FILLS = {
    "FFD6E4F7", "00D6E4F7", "FFEEF3FA", "00EEF3FA",
    "FFF5F8FF", "00F5F8FF", "FFC1D6F0", "00C1D6F0",
    "00DEEAF1", "00E8EEF7", "00D9E1F2", "FFD9E1F2",
    "00DDEBF7", "FFDDEBF7", "00D6E4F0", "00D9E2F3",
}
_DIM_FONT_COLORS = {
    "FF555555", "00555555", "FF999999", "00999999",
    "FF808080", "00808080", "FF444444", "00444444",
}

_SECTION_BANNER_PATTERN = re.compile(r"^[A-Z]\d?\s*[|:]\s*.+")


def _get_color(color_obj) -> str | None:
    if color_obj is None:
        return None
    try:
        value = color_obj.value
    except Exception:
        return None
    if not isinstance(value, str):
        return None
    return value.upper()


def _cell_has_fill(cell) -> bool:
    color = _get_color(cell.fill.fgColor) if cell.fill else None
    if color is None:
        return False
    return color not in ("00000000", "FFFFFFFF", "00FFFFFF")


def _is_real_formula(value) -> bool:
    if not isinstance(value, str) or not value.startswith("="):
        return False
    body = value[1:].strip()
    if any(ch in value for ch in "÷×²"):
        return False
    if re.search(r"\$\d", body):
        return False
    if not body:
        return False
    if re.match(r"^[A-Z][A-Z0-9_.]*\s*\(.*\)\s*$", body, re.IGNORECASE):
        return True
    if re.match(r"^[\s\d.,+\-*/()A-Z$:!]+$", body, re.IGNORECASE):
        return True
    return False


def _neutralise_formula_strings(ws):
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and not _is_real_formula(v):
                cell.value = v.lstrip("=").strip() or "-"


def _restyle_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None and not _cell_has_fill(cell):
                continue
            fc = _get_color(cell.fill.fgColor) if cell.fill else None
            font_c = _get_color(cell.font.color) if (cell.font and cell.font.color) else None
            is_bold = bool(cell.font and cell.font.bold)
            is_italic = bool(cell.font and cell.font.italic)
            size = cell.font.size if (cell.font and cell.font.size) else 11

            if fc in _SECTION_SOURCE_FILLS:
                if is_bold:
                    cell.font = copy(SECTION_FONT)
                    cell.fill = copy(SECTION_FILL)
                else:
                    cell.font = copy(DATA_FONT)
                    cell.fill = copy(BLANK_FILL)
                continue
            if fc in _SUBHDR_SOURCE_FILLS:
                # Sub-header source fills used for single-cell highlights are NOT section headers
                if is_bold:
                    cell.font = copy(EMPH_FONT)
                else:
                    cell.font = copy(DATA_FONT)
                cell.fill = copy(BLANK_FILL)
                continue
            if is_italic and (font_c in _DIM_FONT_COLORS or (size and size <= 10)):
                cell.font = copy(NOTE_FONT)
                cell.fill = copy(BLANK_FILL)
                continue
            if is_bold:
                cell.font = copy(EMPH_FONT)
                cell.fill = copy(BLANK_FILL)
                continue
            cell.font = copy(DATA_FONT)
            cell.fill = copy(BLANK_FILL)


def _merge_banner_rows(ws):
    max_col = ws.max_column or 1
    already_merged = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            already_merged.add(r)
    for r in range(1, (ws.max_row or 1) + 1):
        if r in already_merged:
            continue
        cells_with_val = [(c, ws.cell(r, c).value)
                          for c in range(1, max_col + 1)
                          if ws.cell(r, c).value is not None]
        if len(cells_with_val) != 1 or cells_with_val[0][0] != 1:
            continue
        v = str(cells_with_val[0][1])
        longest = max((len(line) for line in v.split("\n")), default=0)
        cell = ws.cell(r, 1)
        is_bold = bool(cell.font and cell.font.bold)
        is_italic = bool(cell.font and cell.font.italic)
        if longest >= 18 or is_bold or is_italic or _SECTION_BANNER_PATTERN.match(v.strip() or ""):
            try:
                ws.merge_cells(start_row=r, start_column=1,
                               end_row=r, end_column=max_col)
                cur = cell.alignment
                cell.alignment = Alignment(horizontal="left",
                                           vertical="center",
                                           wrap_text=False,
                                           indent=cur.indent)
            except Exception:
                pass


def _apply_section_to_full_width_merges(ws):
    max_col = ws.max_column or 1
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row != mr.max_row:
            continue
        if mr.max_col < max_col:
            continue
        r = mr.min_row
        tl = ws.cell(r, 1)
        is_bold = bool(tl.font and tl.font.bold)
        is_italic = bool(tl.font and tl.font.italic)
        if is_bold:
            for c in range(mr.min_col, mr.max_col + 1):
                cell = ws.cell(r, c)
                cell.font = copy(SECTION_FONT)
                cell.fill = copy(SECTION_FILL)
            cur = tl.alignment
            tl.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=False, indent=cur.indent)
        elif is_italic:
            tl.font = copy(NOTE_FONT)


def _apply_column_widths(ws):
    max_col = ws.max_column or 1
    banner_rows = set()
    for mr in ws.merged_cells.ranges:
        if mr.min_row == mr.max_row and mr.max_col == max_col:
            banner_rows.add(mr.min_row)
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.row in banner_rows:
                continue
            v = str(cell.value)
            col = cell.column
            longest = max((len(line) for line in v.split("\n")), default=0)
            if longest > widths.get(col, 0):
                widths[col] = longest
    for c in range(1, max_col + 1):
        L = widths.get(c, 10)
        ws.column_dimensions[get_column_letter(c)].width = min(35, max(10, L + 2))


def _apply_wrap_rules(ws):
    max_col = ws.max_column or 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fc = _get_color(cell.fill.fgColor) if cell.fill else None
            cur = cell.alignment
            if fc == "001F4E79":
                cell.alignment = Alignment(horizontal=cur.horizontal or "left",
                                           vertical="center",
                                           wrap_text=False, indent=cur.indent)
                continue
            cl = get_column_letter(cell.column)
            cw = ws.column_dimensions[cl].width or 10
            v = str(cell.value)
            longest = max((len(line) for line in v.split("\n")), default=0)
            needs_wrap = longest > cw - 1 or "\n" in v
            cell.alignment = Alignment(horizontal=cur.horizontal,
                                       vertical=cur.vertical,
                                       wrap_text=needs_wrap, indent=cur.indent)


def _apply_borders(ws):
    mr_max = ws.max_row or 1
    mc_max = ws.max_column or 1
    for rng in ws.merged_cells.ranges:
        if rng.max_row > mr_max:
            mr_max = rng.max_row
        if rng.max_col > mc_max:
            mc_max = rng.max_col
    for r in range(1, mr_max + 1):
        for c in range(1, mc_max + 1):
            ws.cell(r, c).border = copy(DATA_BORDER)


def _reset_row_heights(ws):
    for r_idx in list(ws.row_dimensions.keys()):
        ws.row_dimensions[r_idx].height = None


def apply_unified_styling(wb):
    """Apply the unified styling palette across all sheets of an openpyxl workbook.

    Safe to call multiple times; idempotent on the output.
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            ws.sheet_properties.tabColor = None
        except Exception:
            pass
        _neutralise_formula_strings(ws)
        _restyle_cells(ws)
        _merge_banner_rows(ws)
        _apply_section_to_full_width_merges(ws)
        _reset_row_heights(ws)
        _apply_column_widths(ws)
        _apply_wrap_rules(ws)
        _apply_borders(ws)
    return wb
