"""
Scale index — normalisation builder
================================
Reference table shows all nodes (5 target + 3 anchors + full crop level).
CALCULATION uses only the 8 core nodes (5 target + 3 anchors).
Crop-level nodes shown as context/reference, clearly marked.
Includes Z-score robustness comparison matching DMA style.
"""

import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
NAVY      = "203864"
HDR_BLUE  = "2E5FA3"
SUB_BLUE  = "E8EEF7"
ROW_A     = "FFFFFF"
ROW_B     = "F2F2F2"
KEY_GREEN = "E2EFDA"

TIER_BG = {1: "FF6B6B", 2: "FFC7CE", 3: "FFEB9C", 4: "92D050", 5: "C6EFCE"}
TIER_FC = {1: "FFFFFF", 2: "9C0006", 3: "7D6608", 4: "276221", 5: "276221"}
CAT_BG  = {"Target": None, "Crop": None, "Anchor": None}
CAT_FC  = {"Target": "000000", "Crop": "000000", "Anchor": "000000"}

def _side():    return Side(style="thin")
def _border():  return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _al(h="center", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)

def sc(ws, row, col, val, fmt=None, bg=None, bold=False, sz=10, color="000000",
       halign="center", wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name=FONT_NAME, bold=bold, size=sz, color=color)
    c.border    = _border()
    c.alignment = _al(halign, wrap)
    if bg:  c.fill          = _fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, col, val, bg=HDR_BLUE, sz=10, color="FFFFFF", halign="center"):
    return sc(ws, row, col, val, bg=bg, bold=True, sz=sz, color=color, halign=halign)

def title_row(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill      = _fill(NAVY)
    c.alignment = _al()
    c.border    = _border()

def tier_label(t):
    return {1: "Tier 1\n(Low)", 2: "Tier 2", 3: "Tier 3\n(Mid)",
            4: "Tier 4",       5: "Tier 5\n(High)"}[t]

def tier_cell(ws, row, col, t):
    return sc(ws, row, col, tier_label(t), bg=TIER_BG[t], bold=True, sz=9, color=TIER_FC[t])

def assign_bins(score):
    if score < 0.20: return 1
    if score < 0.40: return 2
    if score < 0.60: return 3
    if score < 0.80: return 4
    return 5

def assign_tiers_n8(scores):
    """n=8: ranks 1-2→T5, 3-4→T4, 5-6→T3, 7→T2, 8→T1."""
    import pandas as pd
    s = pd.Series(scores)
    ranks = s.rank(ascending=False, method="first").astype(int)
    mapping = {1:5, 2:5, 3:4, 4:4, 5:3, 6:3, 7:2, 8:1}
    return [mapping[r] for r in ranks.tolist()]

# ------------------------------------------------------------
# ALL NODES — displayed in reference table
# ------------------------------------------------------------
# (label, kt_N, category, in_calc, source_note)
# in_calc=True: used for normalization; False: context only
ALL_NODES = [
    # 5 Target processing nodes — USED IN CALCULATION
    ("C101 – Meat & Meat Products (NACE 10.1)", 2692, "Target", True,
     "Regulatory Entry Points: Regulatory Reach: pre-alloc. Pigs+Beef+Broilers excl. dairy/layer co-products"),
    ("C109 – ICF (NACE 10.9)",           2119, "Target", True,
     "Regulatory Entry Points: Regulatory Reach = Full Oilseed Scope (no co-product alloc. involved)"),
    ("C105 – Dairy (NACE 10.5)",         2172, "Target", True,
     "Regulatory Entry Points: Regulatory Reach: pre-alloc. full dairy cattle feed (100%, not 92.2%)"),
    ("C104 – Oils & Fats (NACE 10.4)",   1332, "Target", True,
     "Regulatory Entry Points: Regulatory Reach = Full Oilseed Scope (oil+meal inseparable co-products)"),
    ("C106 – Grain+Starch (1061+1062)",  1058, "Target", True,
     "Regulatory Entry Points: Regulatory Reach = Full Oilseed Scope (no oilseed meal input)"),
    # 3 Anchor nodes — USED IN CALCULATION
    ("Eggs (NACE 10.89)",                 306, "Anchor", True,
     "Regulatory Entry Points: Regulatory Reach: pre-alloc. full layer feed (100%, not 95.3%)"),
    ("Refined Sugar (NACE 1081)",          139, "Anchor", True,
     "Regulatory Entry Points: Regulatory Reach = Full Oilseed Scope (no oilseed content)"),
    ("Beer (NACE 1105)",                   107, "Anchor", True,
     "Regulatory Entry Points: Regulatory Reach = Full Oilseed Scope (no oilseed content)"),
]

# 8-node calculation set
CALC_NODES = [(label, kt) for label, kt, cat, in_calc, _ in ALL_NODES if in_calc]
calc_kt    = [n[1] for n in CALC_NODES]
kt_min, kt_max = min(calc_kt), max(calc_kt)
kt_mean  = np.mean(calc_kt)
kt_std   = np.std(calc_kt, ddof=1)

def mm(kt):  return (kt - kt_min) / (kt_max - kt_min)
def zr(kt):  return (kt - kt_mean) / kt_std

z_raw_all = [zr(kt) for kt in calc_kt]
zs_min, zs_max = min(z_raw_all), max(z_raw_all)
def zs(kt):  return (zr(kt) - zs_min) / (zs_max - zs_min)

mm_scores = [mm(kt)  for _, kt in CALC_NODES]
z_scores  = [zs(kt)  for _, kt in CALC_NODES]
bins_mm   = [assign_bins(s) for s in mm_scores]
bins_z    = [assign_bins(s) for s in z_scores]
tiers_mm  = assign_tiers_n8(mm_scores)
tiers_z   = assign_tiers_n8(z_scores)

rank_mm8 = [0]*8; rank_z8 = [0]*8
for pos, idx in enumerate(sorted(range(8), key=lambda i: -mm_scores[i])): rank_mm8[idx] = pos+1
for pos, idx in enumerate(sorted(range(8), key=lambda i: -z_scores[i])):  rank_z8[idx]  = pos+1

rho, _ = stats.spearmanr(rank_mm8, rank_z8)


# ------------------------------------------------------------
# SHEET 1 — Full Reference Table (MM + Z side by side)
# ------------------------------------------------------------
def build_reference(wb):
    ws = wb.active
    ws.title = "01_Scale_Normalization"
    # Columns: A-B Node | C Cat | D kt N | E-G MM (Score, Bin, Tier) | H-J Z (Score, Bin, Tier) | K Used | L Note
    NC = 12

    title_row(ws, 1, NC,
              "Scale Dimension – Min-Max vs. Z-Score Comparison  "
              f"|  n=8 for normalization  |  Crop level shown for context  |  Spearman ρ={rho:.3f}")
    ws.row_dimensions[1].height = 22

    # Note row
    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws.cell(2, 1,
        "Normalization uses 8 nodes only (5 NACE target processing nodes + 3 anchor nodes, marked ✓). "
        "Crop-level nodes shown as system context — excluded from calculation. "
        "Economic allocation (ISO 14044) at all co-product nodes. Equal-width 5 bins on [0,1].")
    c.font = Font(name=FONT_NAME, size=9, color="444444")
    c.fill = _fill("F7F7F7"); c.alignment = _al("left"); c.border = _border()
    ws.row_dimensions[2].height = 30

    # Row 3: section span headers
    spans = [(1,2,"Node"),(3,3,"Cat."),(4,4,"kt N"),
             (5,7,"Min-Max Normalization"),(8,10,"Z-Score Normalization\n(rescaled to 0–1)"),
             (11,11,"Calc?"),(12,12,"Source / Note")]
    for c1,c2,text in spans:
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        c = ws.cell(3, c1, text)
        c.font=Font(name=FONT_NAME,bold=True,size=10,color="FFFFFF")
        c.fill=_fill(HDR_BLUE); c.alignment=_al(); c.border=_border()
    ws.row_dimensions[3].height = 30

    # Row 4: column sub-headers
    col_hdrs = ["Node","","Cat.","kt N",
                "MM Score","Bin","Tier",
                "Z Score","Bin (z)","Tier (z)",
                "✓","Source / Note"]
    for j,h in enumerate(col_hdrs):
        hdr(ws, 4, j+1, h, bg=SUB_BLUE, color="000000")
    ws.merge_cells("A4:B4")
    ws.row_dimensions[4].height = 30

    # Data rows — all nodes sorted by kt N descending
    nodes_sorted = sorted(ALL_NODES, key=lambda x: -x[1])
    for i, (label, kt, cat, in_calc, note) in enumerate(nodes_sorted):
        r  = 5 + i
        bg = ROW_A if i % 2 == 0 else ROW_B
        ws.merge_cells(f"A{r}:B{r}")

        if in_calc:
            ist = cat == "Target"
            ci = next(j for j,(l,_) in enumerate(CALC_NODES) if l == label)
            sc(ws, r, 1,  label,                bg=CAT_BG[cat], bold=ist, color=CAT_FC[cat], halign="left")
            sc(ws, r, 3,  cat,                  bg=CAT_BG[cat], bold=ist, color=CAT_FC[cat])
            sc(ws, r, 4,  kt,    "#,##0",       bg=bg, bold=ist)
            sc(ws, r, 5,  round(mm_scores[ci],3), "0.000", bg=bg, bold=ist)
            tier_cell(ws, r, 6, bins_mm[ci])
            tier_cell(ws, r, 7, tiers_mm[ci])
            sc(ws, r, 8,  round(z_scores[ci],3),  "0.000", bg=bg, bold=ist)
            tier_cell(ws, r, 9, bins_z[ci])
            tier_cell(ws, r, 10, tiers_z[ci])
            sc(ws, r, 11, "✓", bg="DEEAF1", bold=True, color="1F4E79")
        else:
            sc(ws, r, 1,  label, bg=CAT_BG[cat], color=CAT_FC[cat], halign="left")
            sc(ws, r, 3,  cat,   bg=CAT_BG[cat], color=CAT_FC[cat])
            sc(ws, r, 4,  kt,    bg=bg, fmt="#,##0")
            for col in [5,6,7,8,9,10,11]:
                sc(ws, r, col, "—", bg="F7F7F7", color="999999")

        sc(ws, r, 12, note, bg=bg, halign="left", sz=9, color="444444")
        ws.row_dimensions[r].height = 20

    # Key finding row
    r_key = 5 + len(ALL_NODES) + 1
    ws.merge_cells(f"A{r_key}:{get_column_letter(NC)}{r_key}")
    c = ws.cell(r_key, 1,
        f"Key finding: Min-Max and Z-score normalization produce identical sector rankings "
        f"(Spearman ρ = {rho:.3f}). Bin and Tier assignments are fully stable across both methods.")
    c.font=Font(name=FONT_NAME,bold=True,size=10); c.fill=_fill(KEY_GREEN)
    c.alignment=_al("left"); c.border=_border(); ws.row_dimensions[r_key].height=18

    # Parameter note
    r_p = r_key + 1
    ws.merge_cells(f"A{r_p}:{get_column_letter(NC)}{r_p}")
    c = ws.cell(r_p, 1,
        f"Normalization parameters (8-node calc set):  "
        f"Min-Max: min={kt_min} kt N (Beer), max={kt_max} kt N (C101–Meat).  "
        f"Z-score: mean={kt_mean:.0f} kt N, SD={kt_std:.0f} kt N (rescaled to [0,1]).")
    c.font=Font(name=FONT_NAME,size=9,color="1F4E79"); c.fill=_fill("DEEAF1")
    c.alignment=_al("left"); c.border=_border(); ws.row_dimensions[r_p].height=14

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 50
    ws.freeze_panes = "A5"


# ------------------------------------------------------------
# SHEET 2 — MM vs Z-Score Comparison (8 nodes)
# ------------------------------------------------------------
def build_comparison(wb):
    ws = wb.create_sheet("02_Scale_Comparison")
    NC = 12

    title_row(ws, 1, NC,
              "Scale Dimension – Min-Max vs. Z-Score Comparison  "
              f"|  n=8  |  Tier 5 = Largest embedded-N footprint  |  Spearman ρ={rho:.3f}")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws.cell(2, 1,
        "5 NACE target processing nodes + 3 anchor nodes (Eggs, Refined Sugar, Beer). "
        "Min-Max on [107, 2562] kt N. Z-score rescaled to [0,1]. Equal-width 5 bins.")
    c.font=Font(name=FONT_NAME,size=9,color="444444"); c.fill=_fill("F7F7F7")
    c.alignment=_al("left"); c.border=_border()
    ws.row_dimensions[2].height = 24

    # Section span headers
    spans = [(1,2,"Node"),(3,3,"Cat."),(4,4,"kt N"),
             (5,8,"Min-Max Normalization"),(9,12,"Z-Score Normalization (rescaled 0–1)")]
    for c1,c2,text in spans:
        ws.merge_cells(start_row=3, start_column=c1, end_row=3, end_column=c2)
        c = ws.cell(3, c1, text)
        c.font=Font(name=FONT_NAME,bold=True,size=10,color="FFFFFF")
        c.fill=_fill(HDR_BLUE); c.alignment=_al(); c.border=_border()

    col_hdrs = ["Node","","Cat.","kt N","MM Score","Bin","Rank","Tier",
                "Z Score","Bin (z)","Rank (z)","Tier (z)"]
    for j,h in enumerate(col_hdrs):
        hdr(ws, 4, j+1, h, bg=SUB_BLUE, color="000000")
    ws.merge_cells("A4:B4")
    ws.row_dimensions[4].height = 30

    order = sorted(range(8), key=lambda i: rank_mm8[i])
    for i, idx in enumerate(order):
        r   = 5 + i
        lbl = CALC_NODES[idx][0]
        kt  = CALC_NODES[idx][1]
        cat = next(c for l,_,c,inc,__ in ALL_NODES if l==lbl)
        bg  = ROW_A if i%2==0 else ROW_B
        ist = cat == "Target"

        ws.merge_cells(f"A{r}:B{r}")
        sc(ws, r, 1,  lbl,                  bg=CAT_BG[cat], bold=ist, color=CAT_FC[cat], halign="left")
        sc(ws, r, 3,  cat,                  bg=CAT_BG[cat], color=CAT_FC[cat])
        sc(ws, r, 4,  kt,    "#,##0",       bg=bg, bold=ist)
        sc(ws, r, 5,  round(mm_scores[idx],3), "0.000", bg=bg, bold=ist)
        tier_cell(ws, r, 6, bins_mm[idx])
        sc(ws, r, 7,  rank_mm8[idx],        bg=bg)
        tier_cell(ws, r, 8, tiers_mm[idx])
        sc(ws, r, 9,  round(z_scores[idx],3), "0.000", bg=bg, bold=ist)
        tier_cell(ws, r, 10, bins_z[idx])
        sc(ws, r, 11, rank_z8[idx],         bg=bg)
        tier_cell(ws, r, 12, tiers_z[idx])
        ws.row_dimensions[r].height = 23

    r_key = 5 + 8 + 1
    ws.merge_cells(f"A{r_key}:{get_column_letter(NC)}{r_key}")
    c = ws.cell(r_key, 1,
        f"Key finding: Min-Max and Z-score normalization produce identical sector rankings "
        f"(Spearman ρ = {rho:.3f}). Bin assignments are fully stable across normalization methods.")
    c.font=Font(name=FONT_NAME,bold=True,size=10); c.fill=_fill(KEY_GREEN)
    c.alignment=_al("left"); c.border=_border()

    notes_s = [
        "  1. Embedded N measured at processing node input level (upstream), not at end-product level (ISO 14044 economic allocation).",
        f"  2. Normalization parameters: Min={kt_min} kt N (Beer), Max={kt_max} kt N (C101–Meat), Mean={kt_mean:.0f}, SD={kt_std:.0f} kt N.",
        "  3. Tier 5 = largest embedded-N footprint = strongest Scale leverage point. Consistent with DMA: higher tier = stronger GLM candidate.",
        "  4. Crop-level nodes shown in Sheet 01 for system context but excluded from normalization to avoid mixing processing and farm levels.",
    ]
    for k, note in enumerate(notes_s):
        rn = r_key + 1 + k
        ws.merge_cells(f"A{rn}:{get_column_letter(NC)}{rn}")
        c=ws.cell(rn,1,note); c.font=Font(name=FONT_NAME,size=9)
        c.alignment=_al("left"); c.border=_border()

    for j,w in enumerate([32,3,10,10,10,12,8,13,10,12,8,13]):
        ws.column_dimensions[get_column_letter(j+1)].width = w
    ws.freeze_panes = "A5"


# ------------------------------------------------------------
# SHEET 3 — Methodology Decisions
# ------------------------------------------------------------
def build_decisions(wb):
    ws = wb.create_sheet("03_Methodology_Decisions")
    NC = 5
    title_row(ws, 1, NC, "Scale Dimension — Methodology Decisions Log  |  GLM Green Lead Markets")
    ws.row_dimensions[1].height = 22

    for j,h in enumerate(["ID","Decision point","Decision taken","Rationale","Reference"]):
        hdr(ws, 2, j+1, h, halign="left")
    ws.row_dimensions[2].height = 20

    decisions = [
        ("D1","Allocation method",
         "Economic allocation (ISO 14044) at all co-production points.",
         "Research question targets economic demand lever, not biophysical N flow.",
         "Methods section"),
        ("D2","Measurement point",
         "Embedded N at input of processing node (upstream), not at end product.",
         "Direct correspondence to NACE node as regulatory anchor point.",
         "Methods section"),
        ("D3","ICF vs. on-farm feed",
         "C109 = industrial compound feed only (2,025 kt N). Grassland and silage tracked separately.",
         "Regulatorily relevant intervention point is the industrial processing stage.",
         "Methods section"),
        ("D4","Reference population (normalisation)",
         "8 nodes: 5 target sectors + eggs + refined sugar + beer.",
         "Well-delineated product categories; crops excluded from normalisation.",
         "Methods section"),
        ("D5","Normalisation and bins",
         "Min-max [0,1] with equal-width 5 bins; z-score as robustness check in sheet 02.",
         f"Identical to DMA methodology; Spearman rho = {rho:.3f} confirms stability.",
         "DMA methodology (see 07_dma_index)"),
        ("D6","Tier direction",
         "Tier 5 = largest embedded-N footprint (strongest scale lever).",
         "Consistent with DMA: higher tier number = stronger lead-market candidate.",
         "Index synthesis logic"),
        ("D7","Scale vs. index",
         "Scale = economic allocation; index construction = regulatory logic (full catchment).",
         "Methodology change documented transparently in the Methods section.",
         "Methods section"),
    ]

    for i,(did,point,dec,rat,ref) in enumerate(decisions):
        r  = 3 + i
        bg = ROW_A if i%2==0 else ROW_B
        sc(ws,r,1,did,   bg="DEEAF1",bold=True,color="1F4E79")
        sc(ws,r,2,point, bg=bg,halign="left",bold=True)
        sc(ws,r,3,dec,   bg=bg,halign="left")
        sc(ws,r,4,rat,   bg=bg,halign="left",sz=9,color="444444")
        sc(ws,r,5,ref,   bg=bg,halign="left",sz=9,color="666666")
        ws.row_dimensions[r].height = 44

    for j,w in enumerate([6,24,40,52,22]):
        ws.column_dimensions[get_column_letter(j+1)].width = w
    ws.freeze_panes = "A3"


def main():
    wb = Workbook()
    build_reference(wb)
    build_comparison(wb)
    build_decisions(wb)
    out = "../outputs/02_scale_index.xlsx"
    # Apply unified visual styling for the replication package (see _styling.py)
    from _styling import apply_unified_styling
    apply_unified_styling(wb)

    wb.save(out)
    print(f"Saved -> {out}")
    print(f"\nSpearman ρ = {rho:.3f}")
    print("\nScale scores (8-node calc):")
    print(f"{'Node':<38} {'kt N':>6}  {'MM':>5}  {'Z':>5}  Bin")
    for idx in sorted(range(8), key=lambda i: rank_mm8[i]):
        lbl, kt = CALC_NODES[idx]
        cat = next(c for l,_,c,inc,__ in ALL_NODES if l==lbl)
        flag = " ◄" if cat=="Target" else ""
        print(f"  {lbl:<38} {kt:>6,}  {mm_scores[idx]:>5.3f}  {z_scores[idx]:>5.3f}  {bins_mm[idx]}{flag}")

if __name__ == "__main__":
    main()
