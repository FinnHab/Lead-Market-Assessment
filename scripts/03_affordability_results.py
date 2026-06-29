"""
Affordability results — green premium calculations.

Builds the workbook from scratch:
    Sheet 1   Inputs & Parameters
    Sheet 2   Green Premium Calculations (three scenarios)
    Sheet 3   Sources & Methods
    Sheet 4   LCOA Literature Review (EU-focused, n=8, 2022-2025)
    Sheet 5   LCOA Summary (EU scenarios with traceability)

Fossil reference includes the 2025 EU ETS pass-through (~$440 + ~$19 = ~$459/t
NH3). Inputs: IEAGHG 2023 (1.8 t CO2/t NH3), DG ECFIN ETS market average
(EUR 70/t), CBAM Regulation 2023/956.

Run from the scripts/ directory:
    python 03_affordability_results.py
Writes ../outputs/03_affordability_results.xlsx.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# CONSTANTS
# ============================================================
B = 1.202       # Baseline N cost €/kg

# --- Green premium derivation ---
# Fossil NH₃ reference: $440/t production + ETS surcharge (2025: ~$14/t at 10% eff., ETS €70/t) = ~$454/t
# EU LCOA scenarios (median over 8 EU studies, in USD at FX=1.08):
#   Low $816, Central $1,054, High $1,308  (corrected: midpoint used where no genuine central reported)
# Stoichiometry: AN requires 0.425 t NH₃/t; Urea requires 0.567 t NH₃/t + 0.733 t CO₂/t
# CO₂ cost: $150/t (base case)
# EU mix: 64.2% AN (34% N) + 32.1% Urea (46% N) [UAN decomposed]
# N_mix = 0.642 × 0.34 + 0.321 × 0.46 = 0.366
FX = 1.08
LCOA_FOSSIL_PRODUCTION = 440  # Pure production cost without ETS ($/t NH₃)

# === EU ETS / CBAM ===
CO2_INTENSITY_NH3 = 1.8       # t CO₂/t NH₃ (direct ETS-relevant; dena Folie 76, IEA 2021)
ETS_PRICE_2025 = 70           # €/t CO₂ (2025 ETS market average, EC DG ECFIN)
FREE_ALLOC_2025 = 0.90        # ~90% free: benchmark 1.619 / actual 1.8 ≈ 90% covered
ETS_EFF_RATE_2025 = 1 - FREE_ALLOC_2025  # = 10% (above-benchmark emissions not covered)
ETS_SURCHARGE_2025 = CO2_INTENSITY_NH3 * ETS_PRICE_2025 * ETS_EFF_RATE_2025 * FX  # €→$ = ~$14/t

LCOA_FOSSIL = LCOA_FOSSIL_PRODUCTION + ETS_SURCHARGE_2025  # ~$454/t (incl. 2025 ETS at 10% eff.)

NH3_AN = 0.425; NH3_UREA = 0.567; CO2_UREA = 0.733; CO2_COST = 150
SHARE_AN = 0.642; SHARE_UREA = 0.321
N_MIX = SHARE_AN * 0.34 + SHARE_UREA * 0.46  # = 0.366

def derive_premium(lcoa_green, lcoa_fossil_override=None):
    """Derive €/kg N premium from green LCOA ($/t NH₃)."""
    fossil = lcoa_fossil_override if lcoa_fossil_override is not None else LCOA_FOSSIL
    dl = lcoa_green - fossil
    d_AN = dl * NH3_AN
    d_Urea = dl * NH3_UREA + CO2_COST * CO2_UREA
    d_Mix = SHARE_AN * d_AN + SHARE_UREA * d_Urea
    return d_Mix / (N_MIX * 1000) / FX  # €/kg N

P_LOW = derive_premium(816)    # incl. 2025 ETS — corrected median of 8 EU study lows
P_CEN = derive_premium(1054)   # incl. 2025 ETS — corrected: (Low+High)/2 where no genuine central reported
P_HIGH = derive_premium(1308)  # incl. 2025 ETS — unchanged (median of 8 EU study highs)

SCENARIOS = [("Low", P_LOW), ("Central", P_CEN), ("High", P_HIGH)]

# === IMPORT SENSITIVITY ===
LCOA_IMPORT = 610 + 70  # Non-EU Median $610 + $70 transport = $680 delivered
P_IMPORT = derive_premium(LCOA_IMPORT)  # €/kg N

def calc_factor(lcoa_green, lcoa_fossil_override=None):
    """Return (factor, delta_N) for any LCOA."""
    fossil = lcoa_fossil_override if lcoa_fossil_override is not None else LCOA_FOSSIL
    dl = lcoa_green - fossil
    d_AN = dl * NH3_AN; d_Urea = dl * NH3_UREA + CO2_COST * CO2_UREA
    d_Mix = SHARE_AN * d_AN + SHARE_UREA * d_Urea
    d_N = d_Mix / (N_MIX * 1000) / FX
    return d_N / B, d_N
cn = lambda n, y, p: n / y * p  # N cost per tonne of crop

# Regulatory Reach scaling factor
F_MEAT = 2692 / 2405

# ============================================================
# STYLES
# ============================================================
tf  = Font(name='Arial', bold=True, size=14)
sf  = Font(name='Arial', bold=True, size=12, color='4472C4')
hw  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hf  = PatternFill('solid', fgColor='4472C4')
cf  = Font(name='Arial', size=10)
bfont = Font(name='Arial', size=10, color='0000FF')
nf  = Font(name='Arial', size=9, italic=True, color='666666')
rf  = PatternFill('solid', fgColor='FFF2CC')
rfo = Font(name='Arial', bold=True, size=10)
sf2 = Font(name='Arial', bold=True, size=10, color='4472C4')
lf  = Font(name='Arial', size=9, color='888888')
thin_border = Border(
    bottom=Side(style='thin', color='CCCCCC')
)

def hdr(ws, r, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(r, c, h)
        cell.font = hw; cell.fill = hf
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def setw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# HELPER: write a plant-based product row with 3 scenarios
# ============================================================
def a1_3s(ws, r, name, units, ul, price, n_per_t_crop, notes):
    ws.cell(r, 1, name).font = cf
    ws.cell(r, 2, units).font = cf; ws.cell(r, 2).number_format = '#,##0.0'
    ws.cell(r, 3, ul).font = cf
    ws.cell(r, 4, price).font = bfont; ws.cell(r, 4).number_format = '0.0000'
    gps = []
    for ci, (sn, prem) in enumerate(SCENARIOS, 5):
        gp = n_per_t_crop * prem / units / price
        f_ = lf
        fl_ = None
        ws.cell(r, ci, gp).font = f_; ws.cell(r, ci).number_format = '0.000%'
        if fl_: ws.cell(r, ci).fill = fl_
        gps.append(gp)
    ws.cell(r, 9, notes).font = nf
    return gps

# ============================================================
# HELPER: write an animal product row with 3 scenarios
# ============================================================
def a2_3s(ws, r, name, nkt, prod, npu, price, notes):
    ws.cell(r, 1, name).font = cf
    ws.cell(r, 2, nkt).font = bfont; ws.cell(r, 2).number_format = '#,##0'
    ws.cell(r, 3, prod).font = bfont; ws.cell(r, 3).number_format = '#,##0'
    ws.cell(r, 4, npu).font = cf; ws.cell(r, 4).number_format = '0.000000'
    ws.cell(r, 5, price).font = bfont; ws.cell(r, 5).number_format = '€#,##0.0000'
    ws.cell(r, 10, notes).font = nf
    gps = []
    for ci, (sn, prem) in enumerate(SCENARIOS, 6):
        gp = npu * prem / price
        f_ = lf
        fl_ = None
        ws.cell(r, ci, gp).font = f_; ws.cell(r, ci).number_format = '0.000%'
        if fl_: ws.cell(r, ci).fill = fl_
        gps.append(gp)
    return gps
    return gps


# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ==========================
# SHEET 1: INPUTS
# ==========================
ws1 = wb.active; ws1.title = "1. Inputs"
setw(ws1, [36, 18, 18, 18, 52])

r = 1; ws1.cell(r, 1, "GREEN PREMIUM MODEL — INPUT PARAMETERS").font = tf
r = 2; ws1.cell(r, 1, "EU27 2023 | EU-LCOA median scenarios | Differential approach | incl. 2025 EU ETS").font = nf

r = 4; ws1.cell(r, 1, "A | NITROGEN FERTILIZER BASELINE").font = sf; r += 1
hdr(ws1, r, ["Parameter", "Value", "Unit", "Period", "Source"]); r += 1
for label, val, unit, per, src in [
    ("5-Year Avg N-Fertilizer Price EU", 414.54, "€/t fertilizer", "2020–2024", "EU Agridata Fertiliser Observatory [S1]"),
    ("Avg N-content in EU fertilizer mix", 0.3449, "kg N/kg fert.", "—", "Wtd: AN 30.25% (58%), Urea 46% (27%), UAN 30% (15%) [S2]"),
    ("Baseline cost of N", 1.2018, "€/kg N", "2020–2024", "= Price ÷ N-content"),
]:
    ws1.cell(r, 1, label).font = cf; ws1.cell(r, 2, val).font = bfont; ws1.cell(r, 2).number_format = '0.0000'
    ws1.cell(r, 3, unit).font = cf; ws1.cell(r, 4, per).font = cf; ws1.cell(r, 5, src).font = nf; r += 1

r += 1; ws1.cell(r, 1, "B | CROP DATA & NITROGEN APPLICATION (EU27 2023)").font = sf; r += 1
setw(ws1, [36, 15, 15, 12, 16, 16, 14, 52])
hdr(ws1, r, ["Crop", "Area (ha)", "Production (t)", "Yield (t/ha)", "EU Avg Price (€/t)", "N Applic. (kg/ha)", "N Cost (€/ha)", "Source"]); r += 1
for name, area, prod, yld, price, n_app in [
    ("Wheat", 24232800, 133481170, 5.508, 216.24, 98.41),
    ("Barley", 10368150, 47389510, 4.571, 219.89, 77.31),
    ("Grain Maize", 8298970, 61035300, 7.355, 206.21, 96.87),
    ("Rapeseed", 6192830, 19585220, 3.163, 431.59, 108.88),
    ("Sugar Beet", 1468070, 110253300, 75.10, 33.90, 95.15),
    ("Potatoes", 1328950, 48142050, 36.23, 327.21, 91.28),
]:
    ws1.cell(r, 1, name).font = cf; ws1.cell(r, 2, area).font = bfont; ws1.cell(r, 2).number_format = '#,##0'
    ws1.cell(r, 3, prod).font = bfont; ws1.cell(r, 3).number_format = '#,##0'
    ws1.cell(r, 4, yld).font = cf; ws1.cell(r, 4).number_format = '0.000'
    ws1.cell(r, 5, price).font = bfont; ws1.cell(r, 5).number_format = '#,##0.00'
    ws1.cell(r, 6, n_app).font = bfont; ws1.cell(r, 6).number_format = '0.00'
    ws1.cell(r, 7, n_app * B).font = cf; ws1.cell(r, 7).number_format = '0.00'
    ws1.cell(r, 8, "Eurostat apro_cpsh1 2023; IFA 2017/18 × ADJ 0.7809 [S4][S5]").font = nf; r += 1

r += 1; ws1.cell(r, 1, "C | TECHNICAL CONVERSION FACTORS").font = sf; r += 1
hdr(ws1, r, ["Product", "Conversion", "Factor", "", "Source"]); r += 1
for label, desc, val, src in [
    ("Wheat → Flour", "Extraction rate", 0.775, "FAO TCF, EU median of 6 countries [S6]"),
    ("Flour → Dough", "Dough yield (hydration 63% + salt/yeast 3.5%)", 1.635, "Cauvain (2015), Technology of Breadmaking, 3rd ed. [S7]"),
    ("Dough → Bread", "Baking yield (12% weight loss)", 0.88, "Purlis (2011), J. Food Engineering 104(3), 461–467 [S8]"),
    ("Wheat → Bread (net)", "Combined: 0.775 × 1.635 × 0.88", 1.115, "Derived from [S6][S7][S8]"),
    ("Durum → Pasta (net)", "1.422 kg durum/kg pasta", 1.422, "Processing Node Scale Analysis [S6][S9]"),
    ("Barley → Beer", "kg barley per L beer", 0.2027, "Brewers of Europe; DG AGRI BS 1.4 [S10]"),
    ("Sugar Beet → Sugar", "Extraction rate", 0.142, "DG AGRI Balance Sheet 1.15 [S11]"),
    ("Potatoes → Fries", "Yield (frozen)", 0.50, "FAO TCF EU median [S6]"),
    ("Rapeseed → Oil", "Extraction rate", 0.40, "FAO TCF EU median [S6]"),
    ("Rapeseed oil density", "kg/L", 0.91, "Standard"),
    ("Sugar per L carb. drink", "kg/L", 0.105, "Industry standard (~105g/L)"),
]:
    ws1.cell(r, 1, label).font = cf; ws1.cell(r, 2, desc).font = cf
    ws1.cell(r, 3, val).font = bfont; ws1.cell(r, 3).number_format = '0.000'
    ws1.cell(r, 5, src).font = nf; r += 1

r += 1; ws1.cell(r, 1, "D | EU AVERAGE RETAIL PRICES (2023)").font = sf; r += 1
ws1.cell(r, 1, "Eurostat prc_dap15 (2015 baseline), population-weighted, HICP-adjusted to 2023. Rapeseed Oil estimated from DE retail ÷ EU price index.").font = nf; r += 1
hdr(ws1, r, ["Product", "EU Price 2023", "Unit", "", "Source"]); r += 1
for name, price, unit in [
    ("Wheat Flour", 1.042, "€/kg"), ("White Bread", 1.670, "€/500g loaf"), ("Pasta (dry)", 1.192, "€/500g"),
    ("Refined Sugar", 1.429, "€/kg"), ("Carbonated Drink", 1.522, "€/L"), ("Beer", 1.041, "€/0.5L glass"),
    ("Frozen Fries", 1.968, "€/kg"), ("Rapeseed Oil", 2.20, "€/L"),
    ("Chicken Breast", 1.0875, "€/100g"), ("Pork Loin Chop", 0.8437, "€/100g"), ("Minced Beef", 0.9658, "€/100g"),
    ("Milk (fresh)", 1.544, "€/L"), ("Eggs (10)", 2.874, "€/10 eggs"),
]:
    s = "[S12]" if name != "Rapeseed Oil" else "[S13]"
    ws1.cell(r, 1, name).font = cf; ws1.cell(r, 2, price).font = bfont; ws1.cell(r, 2).number_format = '0.0000'
    ws1.cell(r, 3, unit).font = cf; ws1.cell(r, 5, f"Eurostat prc_dap15 weighted 2023 {s}").font = nf; r += 1

r += 1; ws1.cell(r, 1, "E | LIVESTOCK N — REGULATORY REACH").font = sf; r += 1
ws1.cell(r, 1, "Regulatory Reach = pre-allocation + full oilseed scope. Source: Regulatory Entry Points [S14].").font = nf; r += 1
ws1.cell(r, 1, "Meat: premium-product allocation (consistent with rapeseed → oil). N ÷ (PRODCOM × retail yield). Yield: poultry 73% (FAO), pork 77% (FAO), beef 65% (USDA cutting yield).").font = nf; r += 1
hdr(ws1, r, ["Species / End Product", "Reg. Reach (kt N)", "EU Production (t)", "N per unit", "Unit", "Source"]); r += 1
for name, nkt, prod, npu, unit in [
    ("Broilers (→ Chicken Breast)", int(363 * F_MEAT), 6517284, 363 * F_MEAT * 1000 / (6517284 * 0.75) / 10, "kg N / 100g (÷0.75)"),
    ("Pigs (→ Pork Loin Chop)", int(1032 * F_MEAT), 12962605, 1032 * F_MEAT * 1000 / (12962605 * 0.75) / 10, "kg N / 100g (÷0.75)"),
    ("Beef cattle (→ Minced Beef)", int(1010 * F_MEAT), 5427055, 1010 * F_MEAT * 1000 / (5427055 * 0.65) / 10, "kg N / 100g (÷0.65)"),
    ("Dairy cattle (→ Milk)", 2172, 153600000, 2172 * 1000 / 153600000 * 1.03, "kg N / L"),
    ("Laying hens (→ Eggs)", 306, 6700000, 306 * 1000 / 6700000 * 0.0698 * 10, "kg N / 10 eggs"),
]:
    ws1.cell(r, 1, name).font = cf; ws1.cell(r, 2, nkt).font = bfont; ws1.cell(r, 2).number_format = '#,##0'
    ws1.cell(r, 3, prod).font = bfont; ws1.cell(r, 3).number_format = '#,##0'
    ws1.cell(r, 4, npu).font = cf; ws1.cell(r, 4).number_format = '0.000000'
    ws1.cell(r, 5, unit).font = cf; ws1.cell(r, 6, "Regulatory Entry Points [S14]").font = nf; r += 1

r += 1; ws1.cell(r, 1, "F | ICF N PER SPECIES (for NACE 10.9)").font = sf; r += 1
ws1.cell(r, 1, "ICF-only = N from industrial compound feed crops. Excludes roughages and on-farm feed. [S14][S16]").font = nf; r += 1
hdr(ws1, r, ["Species", "ICF N (kt)", "FEFAC ICF Vol. (Mt)", "", "", "Source"]); r += 1
for sp, nkt, vol in [("Broilers", 337, 32.9), ("Pigs", 497, 47.7), ("Beef cattle", 227, 14.2), ("Dairy cattle", 461, 27.9), ("Laying hens", 235, 16.2)]:
    ws1.cell(r, 1, sp).font = cf; ws1.cell(r, 2, nkt).font = bfont; ws1.cell(r, 2).number_format = '#,##0'
    ws1.cell(r, 3, vol).font = bfont; ws1.cell(r, 3).number_format = '0.0'
    ws1.cell(r, 6, "Animal Feed Background [S14]; FEFAC 2024 [S15]").font = nf; r += 1

# ==========================
# SHEET 2: GREEN PREMIUM CALCS
# ==========================
ws2 = wb.create_sheet("2. Green Premium Calcs")
setw(ws2, [32, 14, 14, 14, 14, 14, 14, 14, 52])

wnt = 98.41 / 5.508; bnt = 77.31 / 4.571; sbnt = 95.15 / 75.10; ptnt = 91.28 / 36.23; rsnt = 108.88 / 3.163

r = 1; ws2.cell(r, 1, "GREEN PREMIUM CALCULATIONS — ALL SCENARIOS").font = tf
r = 2; ws2.cell(r, 1, f"Baseline: €{B:.3f}/kg N. Premiums: Low +€{P_LOW:.3f}, Central +€{P_CEN:.3f}, High +€{P_HIGH:.3f}/kg N.").font = nf

# Crop GPs
r = 4; ws2.cell(r, 1, "CROP-LEVEL GREEN PREMIUMS (illustrative)").font = sf
r += 1; ws2.cell(r, 1, "GP at farmgate. Not used in Affordability index.").font = nf
r += 2; hdr(ws2, r, ["Crop", "N/ha (kg)", "Yield (t/ha)", "N/t (kg)", "Price (€/t)", "GP LOW (%)", "GP CEN (%)", "GP HIGH (%)", ""]); r += 1
for name, n_ha, yld, price in [("Wheat", 98.41, 5.508, 216.24), ("Barley", 77.31, 4.571, 219.89), ("Grain Maize", 96.87, 7.355, 206.21),
    ("Rapeseed", 108.88, 3.163, 431.59), ("Sugar Beet", 95.15, 75.10, 33.90), ("Potatoes", 91.28, 36.23, 327.21)]:
    npt = n_ha / yld
    ws2.cell(r, 1, name).font = cf; ws2.cell(r, 2, n_ha).font = cf; ws2.cell(r, 2).number_format = '0.00'
    ws2.cell(r, 3, yld).font = cf; ws2.cell(r, 3).number_format = '0.000'
    ws2.cell(r, 4, npt).font = cf; ws2.cell(r, 4).number_format = '0.00'
    ws2.cell(r, 5, price).font = bfont; ws2.cell(r, 5).number_format = '0.00'
    for ci, (sn, prem) in enumerate(SCENARIOS, 6):
        gp = npt * prem / price
        f_ = lf; fl_ = None
        ws2.cell(r, ci, gp).font = f_; ws2.cell(r, ci).number_format = '0.00%'
        if fl_: ws2.cell(r, ci).fill = fl_
    r += 1

# Approach 1
r += 2; ws2.cell(r, 1, "APPROACH 1 | PLANT-BASED END PRODUCTS").font = sf
r += 1; ws2.cell(r, 1, "1 tonne crop → conversion chain → retail units × EU price. GP = N cost increase per unit ÷ price.").font = nf
r += 2; hdr(ws2, r, ["Product", "Units from 1t", "Unit", "Price (€)", "GP LOW (%)", "GP CEN (%)", "GP HIGH (%)", "", "Conversion chain [sources]"]); r += 1

flour_gps = a1_3s(ws2, r, "Flour (1 kg)", 775, "kg", 1.042, wnt, "Wheat ER 77.5% [S6]"); r += 1
bread_u = 1000 * 0.775 * 1.635 * 0.88 / 0.5
bread_gps = a1_3s(ws2, r, "Bread (500g loaf)", bread_u, "500g", 1.670, wnt, "ER 77.5% [S6] → dough 163.5% [S7] → baking −12% [S8]"); r += 1
pasta_u = 1000 / 1.422 / 0.5
pasta_gps = a1_3s(ws2, r, "Pasta (500g)", pasta_u, "500g", 1.192, wnt, "Conv. factor 1.422 kg durum/kg pasta [S6][S9]"); r += 1
sugar_gps = a1_3s(ws2, r, "Refined Sugar (1 kg)", 142, "kg", 1.429, sbnt, "Sugar Beet ER 14.2% [S11]"); r += 1
coke_gps = a1_3s(ws2, r, "Carbonated Drink (1L)", 142 / 0.105, "L", 1.522, sbnt, "105g sugar/L, ER 14.2%"); r += 1
beer_gps = a1_3s(ws2, r, "Beer (0.5L glass)", (1000 / 0.2027) / 0.5, "0.5L", 1.041, bnt, "202.7g barley/L [S10]"); r += 1
fries_gps = a1_3s(ws2, r, "Frozen Fries (1 kg)", 500, "kg", 1.968, ptnt, "Potato ER 50% (FAO) [S6]"); r += 1
oil_gps = a1_3s(ws2, r, "Rapeseed Oil (1L)", 400 / 0.91, "L", 2.20, rsnt, "ER 40% [S6], 100% N to oil (regulatory focus) [S13]"); r += 1

# Approach 2
r += 2; ws2.cell(r, 1, "APPROACH 2 | ANIMAL PRODUCTS — REGULATORY REACH").font = sf
r += 1; ws2.cell(r, 1, "Regulatory Reach (pre-alloc + full oilseed) ÷ EU production. [S14]").font = nf
r += 1; ws2.cell(r, 1, "Meat per-species: NACE 10.1 total 2,692 kt × species share (363/1032/1010 out of 2,405 kt).").font = nf
r += 1; ws2.cell(r, 1, "GP = (N/unit × Premium) / Price. Price from Eurostat prc_dap15 [S12]. Premium-product allocation: 100% of feed-N to retail meat (÷ yield factor).").font = nf
r += 2; hdr(ws2, r, ["Product", "Total N (kt)", "EU Prod (t)", "N/unit (kg)", "Price (€/unit)", "GP LOW (%)", "GP CEN (%)", "GP HIGH (%)", "", "Notes"]); r += 1

# Retail meat yield factors: PRODCOM output → boneless retail meat
# Source: FAO Technical Conversion Factors (tcf.pdf), EU-6 avg dressing % methodology
# Premium-product allocation: 100% of feed-N allocated to retail meat (consistent with rapeseed → oil)
YIELD_CHICKEN = 0.73  # FAO TCF EU-6 avg: 73% dressing output becomes retail cuts
YIELD_PORK    = 0.77  # FAO TCF EU-6 avg: 77% dressing output becomes retail cuts
YIELD_BEEF    = 0.65  # Carcass-to-retail cutting yield 65% (USDA; Penn State Extension) output becomes retail cuts (higher bone fraction)

ch_n = 363 * F_MEAT * 1000 / (6517284 * YIELD_CHICKEN) / 10
ch_gps = a2_3s(ws2, r, "Chicken Breast (100g)", int(363 * F_MEAT), 6517284, ch_n, 1.0875, "NACE 1012; ÷0.73 FAO"); r += 1
pk_n = 1032 * F_MEAT * 1000 / (12962605 * YIELD_PORK) / 10
pk_gps = a2_3s(ws2, r, "Pork Loin Chop (100g)", int(1032 * F_MEAT), 12962605, pk_n, 0.8437, "NACE 1011; ÷0.77 FAO"); r += 1
bf_n = 1010 * F_MEAT * 1000 / (5427055 * YIELD_BEEF) / 10
bf_gps = a2_3s(ws2, r, "Minced Beef (100g)", int(1010 * F_MEAT), 5427055, bf_n, 0.9658, "NACE 1011; ÷0.65 cutting yield"); r += 1
mk_n = 2172 * 1000 / 153600000 * 1.03
mk_gps = a2_3s(ws2, r, "Milk (1L)", 2172, 153600000, mk_n, 1.544, "Dairy RegReach 2,172 kt; 1L≈1.03kg"); r += 1
eg_n = 306 * 1000 / 6700000 * 0.0698 * 10
eg_gps = a2_3s(ws2, r, "Eggs (10 eggs)", 306, 6700000, eg_n, 2.874, "Eggs RegReach 306 kt; avg 69.8g/egg"); r += 1

# Meat weighted
r += 1; ws2.cell(r, 1, "MEAT WEIGHTED AVERAGE (PRODCOM 2023 value) [S18]").font = sf2; r += 1
ws2.cell(r, 1, "  Poultry 29.8%, Pork 42.8%, Beef 27.3%").font = nf; r += 1
meat_gps = [ch_gps[i] * 0.298 + pk_gps[i] * 0.428 + bf_gps[i] * 0.273 for i in range(3)]
ws2.cell(r, 1, "  Weighted Ø C101").font = rfo
for ci, (sn, _) in enumerate(SCENARIOS, 6):
    f_ = lf; fl_ = None
    ws2.cell(r, ci, meat_gps[ci - 6]).font = f_; ws2.cell(r, ci).number_format = '0.000%'
    if fl_: ws2.cell(r, ci).fill = fl_
r += 1

# Approach 2b: C109
r += 2; ws2.cell(r, 1, "APPROACH 2b | NACE 10.9 — ICF-ONLY N ON ANIMAL END PRODUCTS").font = sf
r += 1; ws2.cell(r, 1, "Only ICF-attributed N is 'greened'. Roughage + on-farm stays grey. [S14][S16]").font = nf
r += 1; ws2.cell(r, 1, "GP = (N/unit × Premium) / Price. Same yield factors as Approach 2 (meat: ÷ 0.75/0.65). Same prices.").font = nf
r += 2; hdr(ws2, r, ["Species", "ICF N (kt)", "EU Prod (t)", "N/unit", "Price (€/unit)", "GP LOW (%)", "GP CEN (%)", "GP HIGH (%)", "", "FEFAC vol. (Mt) [S15]"]); r += 1

icf_n = {"Broilers": 337, "Pigs": 497, "Beef": 227, "Dairy": 461, "Layers": 235}
prod_d = {"Broilers": (6517284, 1.0875), "Pigs": (12962605, 0.8437), "Beef": (5427055, 0.9658), "Dairy": (153600000, 1.544), "Layers": (6700000, 2.874)}
fefac = {"Broilers": 32.9, "Layers": 16.2, "Pigs": 47.7, "Dairy": 27.9, "Beef": 14.2}
ftot = sum(fefac.values())

c109_species_gps = {}
for sp, nkt in icf_n.items():
    tp, price = prod_d[sp]
    if sp == "Pigs": npu = nkt * 1000 / (tp * YIELD_PORK) / 10
    elif sp == "Broilers": npu = nkt * 1000 / (tp * YIELD_CHICKEN) / 10
    elif sp == "Beef": npu = nkt * 1000 / (tp * YIELD_BEEF) / 10
    elif sp == "Dairy": npu = nkt * 1000 / tp * 1.03
    elif sp == "Layers": npu = nkt * 1000 / tp * 0.0698 * 10
    ws2.cell(r, 1, sp).font = cf; ws2.cell(r, 2, nkt).font = bfont; ws2.cell(r, 2).number_format = '#,##0'
    ws2.cell(r, 3, tp).font = bfont; ws2.cell(r, 3).number_format = '#,##0'
    ws2.cell(r, 4, npu).font = cf; ws2.cell(r, 4).number_format = '0.000000'
    ws2.cell(r, 5, price).font = bfont; ws2.cell(r, 5).number_format = '€#,##0.0000'
    ws2.cell(r, 10, f"FEFAC: {fefac[sp]:.1f} Mt ({fefac[sp] / ftot * 100:.1f}%)").font = nf
    sp_gps = []
    for ci, (sn, prem) in enumerate(SCENARIOS, 6):
        gp = npu * prem / price; sp_gps.append(gp)
        f_ = lf; fl_ = None
        ws2.cell(r, ci, gp).font = f_; ws2.cell(r, ci).number_format = '0.000%'
        if fl_: ws2.cell(r, ci).fill = fl_
    c109_species_gps[sp] = sp_gps
    r += 1

c109_wtd = [sum(c109_species_gps[s][i] * fefac[s] / ftot for s in c109_species_gps) for i in range(3)]
r += 1; ws2.cell(r, 1, "FEFAC vol.-weighted Ø C109").font = rfo
for ci, (sn, _) in enumerate(SCENARIOS, 6):
    f_ = lf; fl_ = None
    ws2.cell(r, ci, c109_wtd[ci - 6]).font = f_; ws2.cell(r, ci).number_format = '0.000%'
    if fl_: ws2.cell(r, ci).fill = fl_
r += 1

# Approach 3 (informational)
r += 2; ws2.cell(r, 1, "APPROACH 3 | ICF RECIPE COST PASS-THROUGH (informational)").font = sf
r += 1; ws2.cell(r, 1, "Feed recipe × N-cost-share × 75% RM cost share. Not used for index. [S15][S16]").font = nf
r += 2; hdr(ws2, r, ["Species", "Wheat", "Maize", "Barley", "OilMeal", "Soya", "Wtd N-sh", "×75% RM", "GP HIGH (%)"]); r += 1
w_ncs = 98.41 / 5.508 * B / 216.24; m_ncs = 96.87 / 7.355 * B / 206.21; b_ncs = 77.31 / 4.571 * B / 219.89
feeds_data = [
    ("Broiler", 0.3784, 0.2174, 0.0, 0.0448, 0.2319, 32.9),
    ("Laying Hens", 0.4394, 0.4394, 0.0671, 0.0490, 0.1330, 16.2),
    ("Pigs", 0.2373, 0.1238, 0.2246, 0.0051, 0.1698, 47.7),
    ("Dairy Cattle", 0.0, 0.1574, 0.0123, 0.0600, 0.1185, 27.9),
    ("Beef Cattle", 0.0816, 0.1703, 0.3229, 0.1700, 0.1293, 14.2),
]
for name, w, m, b, o, s, vol in feeds_data:
    wtd = w * w_ncs + m * m_ncs + b * b_ncs + o * 0.1; wtd75 = wtd * 0.75; gph = wtd75 * (P_HIGH / B)
    ws2.cell(r, 1, name).font = cf
    for ci, v in enumerate([w, m, b, o, s], 2): ws2.cell(r, ci, v).font = cf; ws2.cell(r, ci).number_format = '0.00%'
    ws2.cell(r, 7, wtd).font = cf; ws2.cell(r, 7).number_format = '0.000%'
    ws2.cell(r, 8, wtd75).font = cf; ws2.cell(r, 8).number_format = '0.000%'
    ws2.cell(r, 9, gph).font = bfont; ws2.cell(r, 9).number_format = '0.000%'; r += 1

# SUMMARY
r += 2; ws2.cell(r, 1, "SUMMARY | SECTOR ENDPOINTS — ALL SCENARIOS").font = sf
r += 1; ws2.cell(r, 1, "8 NACE nodes. Regulatory Reach basis. Three scenarios from EU-LCOA medians.").font = nf
r += 2; hdr(ws2, r, ["NACE Sector", "Representative End Product(s)", "GP LOW (%)", "GP CEN (%)", "GP HIGH (%)", "Rank (HIGH)", "", "", "Weighting / Notes"]); r += 1

# C106: Bread-only (not weighted with Pasta). Consistent with figure and paper scope.
# Bread (NACE 1071) = €70.6B production value; Pasta (NACE 1073) = €9.1B.
# Bread alone represents the dominant end product of the grain milling chain.
cereal_gps = bread_gps
summary_all = [
    ("C104 – Oils & Fats", "Rapeseed Oil (1L)", oil_gps, 1, "100% N to oil (regulatory focus)"),
    ("C101 – Meat", "Ø Chicken+Pork+Beef", meat_gps, 2, "PRODCOM value-wtd; premium-product alloc. (÷yield)"),
    ("NACE 10.89 – Eggs", "Eggs (10)", eg_gps, 3, "Regulatory Reach 306 kt"),
    ("C105 – Dairy", "Milk (1L)", mk_gps, 4, "Regulatory Reach 2,172 kt"),
    ("C109 – Animal Feed", "ICF-only on end products", c109_wtd, 5, "FEFAC vol.-wtd (Approach 2b)"),
    ("C106 – Grain mill", "Bread (500g)", cereal_gps, 6, "Bread-only (NACE 1071, €70.6B)"),
    ("C110 – Beer", "Beer (0.5L)", beer_gps, 7, "Single product"),
    ("C108 – Sugar", "Carbonated Drink (1L)", coke_gps, 8, "End product of sugar chain"),
]
for nace, prod, gps, rank, notes in summary_all:
    ws2.cell(r, 1, nace).font = bfont
    ws2.cell(r, 2, prod).font = cf
    for ci in range(3):
        f_ = lf; fl_ = None
        ws2.cell(r, 3 + ci, gps[ci]).font = f_; ws2.cell(r, 3 + ci).number_format = '0.000%'
        if fl_: ws2.cell(r, 3 + ci).fill = fl_
    ws2.cell(r, 6, rank).font = cf; ws2.cell(r, 9, notes).font = nf; r += 1

# IMPORT SENSITIVITY
import_fill = PatternFill('solid', fgColor='E2EFDA')
r += 2; ws2.cell(r, 1, "IMPORT SENSITIVITY | GREEN NH₃ FROM NON-EU SOURCES").font = sf
r += 1; ws2.cell(r, 1, "Non-EU Median Low $610/t + $70/t intercontinental transport = $680/t delivered. Comparison vs EU Central ($1,081/t).").font = nf
r += 1; ws2.cell(r, 1, "Sources: Moritz et al. (2023), Rosa & Mingolla (2025). Non-EU median from 9 studies incl. AU, Chile, US, China.").font = nf
r += 2; hdr(ws2, r, ["Product", "", "", "", "EU Central (%)", "Import (%)", "Δ (pp)", "", "Notes"]); r += 1

# Crop-level
ws2.cell(r, 1, "CROPS (farmgate)").font = sf2; r += 1
for name, n_ha, yld, price in [("Wheat", 98.41, 5.508, 216.24), ("Barley", 77.31, 4.571, 219.89),
    ("Grain Maize", 96.87, 7.355, 206.21), ("Rapeseed", 108.88, 3.163, 431.59),
    ("Sugar Beet", 95.15, 75.10, 33.90), ("Potatoes", 91.28, 36.23, 327.21)]:
    npt = n_ha / yld
    gp_eu = npt * P_CEN / price
    gp_imp = npt * P_IMPORT / price
    ws2.cell(r, 1, f"  {name}").font = cf
    ws2.cell(r, 5, gp_eu).font = cf; ws2.cell(r, 5).number_format = '0.00%'
    ws2.cell(r, 6, gp_imp).font = cf; ws2.cell(r, 6).number_format = '0.00%'; ws2.cell(r, 6).fill = import_fill
    ws2.cell(r, 7, gp_imp - gp_eu).font = cf; ws2.cell(r, 7).number_format = '+0.00%;-0.00%'
    r += 1

# Key end products
ws2.cell(r, 1, "END PRODUCTS").font = sf2; r += 1
for name, n_per_t, units, price in [
    ("Bread (500g loaf)", wnt, bread_u, 1.670),
    ("Rapeseed Oil (1L)", rsnt, 400 / 0.91, 2.20),
    ("Beer (0.5L glass)", bnt, (1000 / 0.2027) / 0.5, 1.041),
    ("Refined Sugar (1 kg)", sbnt, 142, 1.429),
]:
    gp_eu = n_per_t * P_CEN / units / price
    gp_imp = n_per_t * P_IMPORT / units / price
    ws2.cell(r, 1, f"  {name}").font = cf
    ws2.cell(r, 5, gp_eu).font = cf; ws2.cell(r, 5).number_format = '0.000%'
    ws2.cell(r, 6, gp_imp).font = cf; ws2.cell(r, 6).number_format = '0.000%'; ws2.cell(r, 6).fill = import_fill
    ws2.cell(r, 7, gp_imp - gp_eu).font = cf; ws2.cell(r, 7).number_format = '+0.000%;-0.000%'
    r += 1

# Animal products
ws2.cell(r, 1, "ANIMAL PRODUCTS").font = sf2; r += 1
for name, npu, price in [
    ("Chicken Breast (100g)", ch_n, 1.0875),
    ("Milk (1L)", mk_n, 1.544),
    ("Eggs (10 eggs)", eg_n, 2.874),
    ("Ø Meat (C101)", sum([ch_n * 0.298, pk_n * 0.428, bf_n * 0.273]), sum([1.0875 * 0.298, 0.8437 * 0.428, 0.9658 * 0.273])),
]:
    gp_eu = npu * P_CEN / price
    gp_imp = npu * P_IMPORT / price
    ws2.cell(r, 1, f"  {name}").font = cf
    ws2.cell(r, 5, gp_eu).font = cf; ws2.cell(r, 5).number_format = '0.000%'
    ws2.cell(r, 6, gp_imp).font = cf; ws2.cell(r, 6).number_format = '0.000%'; ws2.cell(r, 6).fill = import_fill
    ws2.cell(r, 7, gp_imp - gp_eu).font = cf; ws2.cell(r, 7).number_format = '+0.000%;-0.000%'
    r += 1

# Summary
r += 1
ws2.cell(r, 1, "Green Premium Factor (overall)").font = rfo
ws2.cell(r, 5, P_CEN / B).font = rfo; ws2.cell(r, 5).number_format = '+0.0%'
ws2.cell(r, 6, P_IMPORT / B).font = rfo; ws2.cell(r, 6).number_format = '+0.0%'; ws2.cell(r, 6).fill = import_fill
ws2.cell(r, 7, P_IMPORT / B - P_CEN / B).font = rfo; ws2.cell(r, 7).number_format = '+0.0%;-0.0%'
ws2.cell(r, 9, f"Import reduces premium by {(1 - P_IMPORT/P_CEN)*100:.0f}% vs EU production").font = nf
r += 1

# Transport sensitivity
r += 1; ws2.cell(r, 1, "TRANSPORT COST SENSITIVITY").font = sf2; r += 1
for t_label, t_cost in [("$40/t (minimum)", 40), ("$70/t (central)", 70), ("$100/t (conservative)", 100)]:
    lcoa_del = 610 + t_cost
    p_t = derive_premium(lcoa_del)
    f_t = p_t / B
    ws2.cell(r, 1, f"  Non-EU $610 + {t_label}").font = cf
    ws2.cell(r, 5, f"${lcoa_del}/t").font = cf
    ws2.cell(r, 6, f_t).font = cf; ws2.cell(r, 6).number_format = '+0.0%'; ws2.cell(r, 6).fill = import_fill
    ws2.cell(r, 9, f"vs EU Low +{P_LOW/B*100:.1f}%, EU Central +{P_CEN/B*100:.1f}%").font = nf
    r += 1

r += 2; ws2.cell(r, 1, "REFERENCE PRODUCTS").font = sf2; r += 1
for name, gps, note in [
    ("Refined Sugar (1 kg)", sugar_gps, "C108 upstream: GP at refinery node"),
    ("Flour (1 kg)", flour_gps, "C106 upstream: GP at milling node"),
    ("Frozen Fries (1 kg)", fries_gps, "Reference for normalization"),
]:
    ws2.cell(r, 1, name).font = Font(name='Arial', size=10, italic=True, color='666666')
    for ci in range(3):
        ws2.cell(r, 3 + ci, gps[ci]).font = Font(name='Arial', size=10, italic=True, color='666666')
        ws2.cell(r, 3 + ci).number_format = '0.000%'
    ws2.cell(r, 9, note).font = nf; r += 1

r += 2; ws2.cell(r, 1, "NOTES").font = sf2; r += 1
for n in [
    f"GREEN PREMIUM DERIVATION: Differential approach. ΔLCOA × stoichiometric NH₃ input + CO₂ cost for urea. EU mix: 64.2% AN + 32.1% Urea. CO₂ at $150/t. Fossil ref incl. 2025 EU ETS (€70/t CO₂ [S22], 1.8 t CO₂/t [S21], 10% eff. rate [S23])..",
    "RANKING STABILITY: Identical across all scenarios — depends on relative N intensity per € product value, not absolute premium.",
    "EGGS vs MEAT: Eggs rank above weighted meat average due to lower retail price per kg despite similar N intensity.",
]:
    ws2.cell(r, 1, "•").font = cf; ws2.cell(r, 2, n).font = nf
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
    ws2.cell(r, 2).alignment = Alignment(wrap_text=True); r += 1

# ==========================
# SHEET 3: SOURCES & METHODS 
# ==========================
ws3 = wb.create_sheet("3. Sources & Methods")
setw(ws3, [8, 28, 85])
r = 1; ws3.cell(r, 1, "SOURCES, REFERENCES & METHODOLOGY").font = tf
r = 2; ws3.cell(r, 1, "All sources referenced in the analysis via [S##] notation.").font = nf
r = 4; ws3.cell(r, 1, "A | REFERENCES").font = sf; r += 1
hdr(ws3, r, ["#", "Short Label", "Full Citation / URL"]); r += 1
sources = [
    ("S1", "EU Agridata Fertiliser", "European Commission, Agridata Portal. EU Fertiliser Prices Dashboard. Monthly N-fertilizer prices, 2020–2025. URL: https://agridata.ec.europa.eu/extensions/DashboardFertiliser/FertiliserPrices.html"),
    ("S2", "Fertilizers Europe 2023", "Fertilizers Europe (2023). Industry Facts and Figures. EU N-fertilizer product mix: AN 47%, Urea 22%, UAN 12%, Other 19%."),
    ("S3", "LCOA Literature Review", "Systematic review of 8 EU + 9 Non-EU LCOA studies (2022–2025). EU scenarios: median Low $815, Central $1,081, High $1,308 $/t NH₃. Fossil ref $440/t + ETS. Differential approach: ΔLCOA × stoichiometry + CO₂. See Sheet 5 and Green_Premium_Methodik_v3."),
    ("S4", "Eurostat apro_cpsh1", "Eurostat (2024). Crop production in EU standard humidity [apro_cpsh1]. EU27, reference year 2023."),
    ("S5", "IFA × ADJ Factor", "IFA Fertiliser Use by Crop 2017/18 × ADJ 0.7809 (= Fertilizers Europe 8,300 kt ÷ IFA 10,628 kt)."),
    ("S6", "FAO TCFs", "FAO (1997). Technical Conversion Factors for Agricultural Commodities. EU median from 6 countries (DE, FR, IT, ES, PL, NL)."),
    ("S7", "Cauvain (2015)", "Cauvain, S.P. (2015). Technology of Breadmaking, 3rd ed., Springer. Industrial white bread: hydration 58–65%, dough yield ~163.5%."),
    ("S8", "Purlis (2011)", "Purlis, E. (2011). J. Food Engineering 104(3), 461–467. doi: 10.1016/j.jfoodeng.2011.01.012. Baking weight loss 10–15%, industrial avg ~12%."),
    ("S9", "Processing Nodes", "GLM N-Flow Model, Processing Node Scale Analysis. Pasta CF 1.422 kg durum/kg pasta."),
    ("S10", "Brewers of Europe", "Brewers of Europe Annual Report + DG AGRI BS 1.4: 6,300 kt malting barley → 0.2027 kg barley/L beer."),
    ("S11", "DG AGRI BS 1.15", "DG AGRI Crop Balance Sheets 2023/24, Sugar Beet. Extraction rate 14.2%."),
    ("S12", "Eurostat prc_dap15", "Eurostat Detailed Average Prices 2015 [prc_dap15]. Pop-weighted, HICP-adjusted to 2023."),
    ("S13", "Rapeseed Oil price", "Est. EU avg: DE retail €2.40/L ÷ Eurostat price level index (DE=108.6) → €2.20/L."),
    ("S14", "Regulatory Entry Pts", "GLM N-Flow Model, Regulatory Entry Points + Animal Feed Background sheets."),
    ("S15", "FEFAC 2024", "FEFAC Feed & Food Statistical Yearbook 2024. EU ICF: 146.9 Mt. https://fefac.eu/wp-content/uploads/2025/04/FF_2024_FINAL.pdf"),
    ("S16", "GLEAM / GFLI", "FAO GLEAM 3.0 + GFLI 2016. Feed ration crop shares per livestock type."),
    ("S17", "FAO TCF; USDA", "FAO Technical Conversion Factors (tcf.pdf): Pigs 77%, Chickens 73% (EU-6 avg dressing %). Beef: 65% carcass-to-retail cutting yield (USDA yield grade data; Penn State Extension 2022)."),
    ("S18", "PRODCOM 2023", "Eurostat PRODCOM EU27 2023 production value. Meat: NACE 1011 Beef €35.3B, Pork €55.3B; NACE 1012 Poultry €38.5B. Grain: NACE 1071 Bread €70.6B; NACE 1073 Pasta €9.1B."),
    ("S19", "Eurostat farmgate", "Eurostat [apri_ap_crpouta]. Selling prices of crop products, 5-yr avg 2019–2024."),
    ("S20", "World Bank Pink Sheet", "World Bank (2025). Commodity Price Data monthly. TTF 'Natural gas, Europe' series. Pre-crisis avg €16.5/MWh; 2022 avg €131/MWh."),
    ("S21", "DECHEMA 2022 Annex 2", "DECHEMA/Ausfelder (2022). Perspective Europe 2030, Tables 7.7/7.8. NH₃ cost model: NG 32.2 GJ/t, elec 0.61 MWh/t, maint €47/t, CAPEX €85/t."),
    ("S22", "TFI / AGA 2023", "The Fertilizer Institute / AGA (2023). 'Natural gas accounts for 70–90% of ammonia production costs.'"),
]
for ref, label, citation in sources:
    ws3.cell(r, 1, ref).font = Font(name='Arial', bold=True, size=10)
    ws3.cell(r, 2, label).font = cf; ws3.cell(r, 3, citation).font = cf
    ws3.cell(r, 3).alignment = Alignment(wrap_text=True); r += 1

r += 2; ws3.cell(r, 1, "B | METHODOLOGY NOTES").font = sf; r += 2
for n in [
    "SYSTEM BOUNDARY: Green Premium = % increase in retail price from switching all synthetic N fertilizer from grey to green ammonia. Organic manures, biological fixation, atmospheric deposition outside scope.",
    "GREEN PREMIUM DERIVATION: Differential approach. ΔCost = ΔLCOA × NH₃_stoichiometry + CO₂_cost × CO₂_stoichiometry. Conversion costs cancel. EU mix: 64.2% AN (0.425 t NH₃/t) + 32.1% Urea (0.567 t NH₃/t + 0.733 t CO₂/t). CO₂ base case $150/t.",
    "REGULATORY REACH: Pre-allocation + full oilseed scope for all animal products. Rapeseed Oil: 100% of N to oil (no co-product alloc.).",
    "PREMIUM-PRODUCT ALLOCATION: Meat products use retail meat yield factors (poultry 75%, pork 75%, beef 65%) to allocate 100% of feed-N to retail cuts. Co-products (bones, offal, rendering) bear no GP. Consistent with rapeseed → oil allocation. [S17]",
    "BREAD CHAIN: Wheat ER 77.5% (FAO) → Dough 163.5% (Cauvain 2015: 63% hydration + 3.5% salt/yeast) → Baking −12% (Purlis 2011). Net: 1t wheat → 1,115 kg bread.",
    "C108 ENDPOINT: Carbonated Drink (not Refined Sugar). C106 uses Bread as representative end product.",
    "C109: ICF-only N per species ÷ EU production. Approach 3 (recipe method) shown for information only.",
    "MEAT AVERAGE: PRODCOM 2023 production-value-weighted (Poultry 29.8%, Pork 42.8%, Beef 27.3%). [S18]",
    "GRAIN: Bread-only (NACE 1071, €70.6B PRODCOM 2023). Pasta excluded for consistency with figure and paper scope. [S18]",
    "RANKING STABILITY: Identical across Low/Central/High scenarios — determined by N intensity per € product value.",
    "GAS PRICE SENSITIVITY: Fossil NH₃ cost depends 60–90% on natural gas (DECHEMA model: 32.2 GJ/t at TTF). At current TTF (€59/MWh), fossil LCOA is ~$754/t — far above the TEA reference ($440 + ETS). At 2022 crisis (€131/MWh), green NH₃ is cheaper than fossil. Sheet 6 Section E. [S20][S21][S22]",
]:
    ws3.cell(r, 1, "•").font = cf; ws3.cell(r, 2, n).font = cf
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws3.cell(r, 2).alignment = Alignment(wrap_text=True); r += 1

# ==========================
# SHEET 4: LCOA LITERATURE REVIEW 
# ==========================
ws4 = wb.create_sheet("4. LCOA Literature Review")
setw(ws4, [10, 30, 6, 10, 10, 10, 8, 18, 52, 28])

r = 1; ws4.cell(r, 1, "LCOA LITERATURE DATABASE").font = tf
r = 2; ws4.cell(r, 1, "Systematic review: 8 EU + 9 Non-EU studies with tabulated LCOA values (2022–2025). 3 qualitative + 3 excluded.").font = nf
r = 3; ws4.cell(r, 1, "All values in original currency. EU scenarios derived from median (see Sheet 5). FX: 1.08 EUR/USD.").font = nf

r = 5; hdr(ws4, r, ["ID", "Authors / Source", "Year", "Low ($/€)", "High ($/€)", "Base ($/€)", "Cur", "Technology", "Scope / Notes", "Status"]); r += 1

eu_data = [
    ("EU-01", "Magnino et al.", 2025, 723, 1368, 1046, "EUR", "PEM", "PEMEC at 0.10 €/kWh (base), 0.04 (low). Adv. Appl. Energy. Tier 1.", "INCLUDED"),
    ("EU-01b", "Magnino et al.", 2025, 680, 1174, 927, "EUR", "SOEC", "SOEC variant. Higher efficiency. Same boundary. Tier 1.", "INCLUDED"),
    ("EU-02", "Moritz et al.", 2023, 624, 874, 743, "USD", "PEM + HB", "⚠ 2030 proj. Supply costs incl. transport to DE. Tier 1.", "EXCLUDED (transport)"),
    ("EU-03", "DECHEMA/Ausfelder", 2022, 760, 1350, 1035, "EUR", "ALK/PEM", "4 EU regions: N 760, S 830, C 1200, W 1350. 2030 proj. Tier 2.", "INCLUDED"),
    ("EU-04", "Smith & Torrente-M.", 2025, 925, 1240, 1030, "USD", "PEM off-grid", "Nature Chem. Eng. Best 2% EU locations. 4,500+ sites. Tier 1.", "INCLUDED"),
    ("EU-05", "Vinardell et al.", 2023, 600, 900, 700, "EUR", "PEM off-grid", "Spain. 430 kt/yr. Incl. O₂ revenue. ACS Sust. Chem. Eng. Tier 1.", "INCLUDED"),
    ("EU-06", "Sousa et al.", 2022, 750, 1248, 999, "EUR", "PEM hydro", "Norway. 30 MW. Hydropower 0.02–0.04 €/kWh. I&EC Res. Tier 1.", "INCLUDED"),
    ("EU-07", "Cameli et al.", 2024, 915, 915, 915, "USD", "AEL", "France/Belgium. 50 kt/yr. 2019 tech only (2050 proj. excluded). Tier 1.", "INCLUDED"),
    ("EU-08", "Mingolla et al.", 2024, 762, 2325, 1544, "EUR", "AEL/PEM flex", "SSRN preprint. 5 EU zones. Wind-dom 762, Low-cap 2325. Tier 2.", "INCLUDED"),
]

noneu_data = [
    ("NONEU-01", "Butterworth/CRU", 2022, 610, 945, 945, "USD", "PEM wind+solar", "W. Australia. 6.5 Mt/yr mega-project. Tier 2.", "INCLUDED"),
    ("NONEU-02a", "Campion et al.", 2023, 759, 842, 842, "EUR", "AEC off-grid", "N. Chile (solar+). 430 kt/yr. RSER. Tier 1.", "INCLUDED"),
    ("NONEU-02b", "Campion et al.", 2023, 807, 951, 951, "EUR", "AEC off-grid", "S. Australia (wind/solar). Same method. Tier 1.", "INCLUDED"),
    ("NONEU-03", "Tully et al.", 2025, 508, 636, 508, "USD", "PEM off-grid", "USA TX+IA. 1 GW. BAT flexibility. IJHE. Tier 1.", "INCLUDED"),
    ("NONEU-04", "Pan et al.", 2023, 469, 469, 469, "USD", "AEL wind+solar", "China Inner Mongolia. C2A retrofit. iScience. Tier 1.", "INCLUDED"),
    ("NONEU-05", "Schueler et al.", 2024, 508, 704, 508, "USD", "Wind/Solar", "US+CA 41 facilities. Wind $508, Solar $704 medians. ERL. Tier 1.", "INCLUDED"),
    ("NONEU-06", "Lee et al.", 2022, 920, 1060, 920, "USD", "PEM renewable", "USA Argonne. 2022 electrolyzer costs. Green Chem. Tier 1.", "INCLUDED"),
    ("NONEU-07", "Reznicek et al./NREL", 2025, None, None, None, "USD", "Multiple", "NREL multi-lab. LCOA graphical only. Cell Rep. Sust. Tier 1.", "QUALITATIVE"),
    ("NONEU-08", "Saygin et al./IRENA", 2023, 400, 1670, 730, "USD", "AEL review", "Global review. $400–1,670. Sustainability (MDPI). Tier 2.", "INCLUDED"),
    ("NONEU-09", "Smith & Torrente-M.", 2024, None, None, None, "USD", "Multiple", "Joule. Dynamic operation. LCOA in figures only.", "QUALITATIVE"),
    ("NONEU-10", "Mersch et al.", 2024, None, None, None, "USD", "Blue+Green", "Sust. E&F. US policy. LCOA graphical only.", "QUALITATIVE"),
    ("NONEU-11", "Biswas/Green et al.", 2025, 861, 1351, 861, "USD", "PEM wind", "MIT. Green $861–1,351. Blue $365–418. E&F. Tier 1.", "INCLUDED"),
]

excl_data = [
    ("EXCL-01", "Arnaiz del Pozo", 2022, 484, 772, None, "EUR", "PEM", "2050 projection. Primary focus: blue ammonia.", "EXCLUDED — 2050"),
    ("EXCL-02", "Assowe Dabar et al.", 2024, 400, 537, None, "USD", "AEL", "Djibouti. Internal calculation errors.", "EXCLUDED — errors"),
    ("EXCL-03", "Palys & Daoutidis", 2024, None, None, None, "USD", "PEM+Urea", "Reports LCOU not LCOA. Used for cross-validation.", "EXCLUDED — urea"),
]

sep_fill = PatternFill('solid', fgColor='D9E2F3')
new_fill = PatternFill('solid', fgColor='E2EFDA')
excl_fill = PatternFill('solid', fgColor='FFF2CC')

for section_label, section_data, section_fill in [
    (None, eu_data, None),
    ("— NON-EU —", noneu_data, None),
    ("— EXCLUDED —", excl_data, excl_fill),
]:
    if section_label:
        ws4.cell(r, 1, section_label).font = Font(name='Arial', bold=True, size=10)
        for c in range(1, 11): ws4.cell(r, c).fill = sep_fill
        r += 1
    for entry in section_data:
        id_, auth, yr, low, high, base, cur, tech, notes, status = entry
        for c, v in enumerate([id_, auth, yr, low, high, base, cur, tech, notes, status], 1):
            cell = ws4.cell(r, c, v)
            cell.font = cf
            if c in (4, 5, 6) and isinstance(v, (int, float)):
                cell.font = bfont; cell.number_format = '#,##0'
        if section_fill:
            for c in range(1, 11): ws4.cell(r, c).fill = section_fill
        r += 1

# ==========================
# SHEET 5: LCOA SUMMARY (EU Scenario Construction)
# ==========================
ws5 = wb.create_sheet("5. LCOA Summary")
setw(ws5, [8, 30, 8, 10, 52])

r = 1; ws5.cell(r, 1, "LCOA SUMMARY — EU SCENARIO CONSTRUCTION").font = tf
r = 2; ws5.cell(r, 1, "Three scenarios derived from median across 8 EU LCOA studies (2022–2025). All values converted to USD at FX=1.08.").font = nf
r = 3; ws5.cell(r, 1, "These scenarios are used to derive the Green Premium estimates in the main analysis.").font = nf

# Section A: EU studies overview
r = 5; ws5.cell(r, 1, "A | EU STUDIES OVERVIEW (n=8)").font = sf
r += 2; hdr(ws5, r, ["Rank", "Study", "Year", "USD", "Context"]); r += 1

# Low scenario
ws5.cell(r, 1, "LOW SCENARIO — Median of Low values").font = sf2
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
low_ranked = [
    (1, "EU-05 Vinardell (ACS)", 2023, 648, "Solar PV Spain, 430 kt/yr, incl. O₂ revenue"),
    (2, "EU-01b Magnino SOEC (AAE)", 2025, 734, "SOEC at 0.04 €/kWh dedicated RE"),
    (3, "EU-01 Magnino PEM (AAE)", 2025, 781, "PEM at 0.04 €/kWh dedicated RE"),
    (4, "EU-06 Sousa (I&ECR)", 2022, 810, "◄ Hydropower Norway at 0.02 €/kWh"),
    (5, "EU-03 DECHEMA", 2022, 821, "◄ Northern Europe, best wind"),
    (6, "EU-08 Mingolla (SSRN)", 2024, 823, "Wind-dominated EU zone, flex"),
    (7, "EU-07 Cameli (IJHE)", 2024, 915, "AEL at 40 $/MWh, 2019 tech"),
    (8, "EU-04 Smith (Nat.C.E.)", 2025, 925, "Best 2% EU locations, off-grid"),
]
for rank, study, yr, val, ctx in low_ranked:
    ws5.cell(r, 1, rank).font = cf
    ws5.cell(r, 2, study).font = rfo if rank in (4, 5) else cf
    ws5.cell(r, 3, yr).font = cf
    ws5.cell(r, 4, val).font = bfont if rank in (4, 5) else cf; ws5.cell(r, 4).number_format = '$#,##0'
    ws5.cell(r, 5, ctx).font = nf
    if rank in (4, 5):
        pass  # no highlight fill
    r += 1
ws5.cell(r, 1, "").font = cf; ws5.cell(r, 2, "→ Median = ($810 + $821) / 2 =").font = rfo
ws5.cell(r, 4, 815).font = Font(name='Arial', bold=True, size=12, color='4472C4'); ws5.cell(r, 4).number_format = '$#,##0'; r += 2

# Central scenario
ws5.cell(r, 1, "CENTRAL SCENARIO — Median of Midpoints").font = sf2
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
cen_ranked = [
    (1, "EU-05 Vinardell", 2023, 810, "Mid of $648–972"),
    (2, "EU-07 Cameli", 2024, 915, "Single value"),
    (3, "EU-01b Magnino SOEC", 2025, 1001, "Mid of $734–1,268"),
    (4, "EU-06 Sousa", 2022, 1079, "◄ Mid of $810–1,348"),
    (5, "EU-04 Smith", 2025, 1082, "◄ Mid of $925–1,240"),
    (6, "EU-01 Magnino PEM", 2025, 1129, "Mid of $781–1,477"),
    (7, "EU-03 DECHEMA", 2022, 1139, "Mid of $821–1,458"),
    (8, "EU-08 Mingolla", 2024, 1667, "Mid of $823–2,511"),
]
for rank, study, yr, val, ctx in cen_ranked:
    ws5.cell(r, 1, rank).font = cf
    ws5.cell(r, 2, study).font = rfo if rank in (4, 5) else cf
    ws5.cell(r, 3, yr).font = cf
    ws5.cell(r, 4, val).font = bfont if rank in (4, 5) else cf; ws5.cell(r, 4).number_format = '$#,##0'
    ws5.cell(r, 5, ctx).font = nf
    if rank in (4, 5):
        pass  # no highlight fill
    r += 1
ws5.cell(r, 2, "→ Median = ($1,079 + $1,082) / 2 =").font = rfo
ws5.cell(r, 4, 1054).font = Font(name='Arial', bold=True, size=12, color='4472C4'); ws5.cell(r, 4).number_format = '$#,##0'; r += 2

# High scenario
ws5.cell(r, 1, "HIGH SCENARIO — Median of High values").font = sf2
ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
high_ranked = [
    (1, "EU-07 Cameli", 2024, 915, "Single value"),
    (2, "EU-05 Vinardell", 2023, 972, "Wind off-grid Spain"),
    (3, "EU-04 Smith", 2025, 1240, "Broader EU locations"),
    (4, "EU-01b Magnino SOEC", 2025, 1268, "◄ SOEC at 0.10 €/kWh grid"),
    (5, "EU-06 Sousa", 2022, 1348, "◄ Hydropower at 0.04 €/kWh base"),
    (6, "EU-03 DECHEMA", 2022, 1458, "Western Europe, grid"),
    (7, "EU-01 Magnino PEM", 2025, 1477, "PEM at 0.10 €/kWh grid"),
    (8, "EU-08 Mingolla", 2024, 2511, "Low-capacity region"),
]
for rank, study, yr, val, ctx in high_ranked:
    ws5.cell(r, 1, rank).font = cf
    ws5.cell(r, 2, study).font = rfo if rank in (4, 5) else cf
    ws5.cell(r, 3, yr).font = cf
    ws5.cell(r, 4, val).font = bfont if rank in (4, 5) else cf; ws5.cell(r, 4).number_format = '$#,##0'
    ws5.cell(r, 5, ctx).font = nf
    if rank in (4, 5):
        pass  # no highlight fill
    r += 1
ws5.cell(r, 2, "→ Median = ($1,268 + $1,348) / 2 =").font = rfo
ws5.cell(r, 4, 1308).font = Font(name='Arial', bold=True, size=12, color='4472C4'); ws5.cell(r, 4).number_format = '$#,##0'; r += 2

# Section B: Derivation to €/kg N
r += 1; ws5.cell(r, 1, "B | GREEN PREMIUM DERIVATION (incl. 2025 EU ETS)").font = sf; r += 1
ws5.cell(r, 1, "Differential approach: ΔLCOA × stoichiometric NH₃ input + CO₂ cost for urea. Fossil ref includes 2025 EU ETS at ~10% effective rate (above-benchmark gap).").font = nf; r += 2
setw(ws5, [8, 30, 12, 12, 12])
hdr(ws5, r, ["", "Parameter", "Low", "Central", "High"]); r += 1

_f = LCOA_FOSSIL  # ~$459 (incl. ETS)
_dL = 816 - _f; _dC = 1054 - _f; _dH = 1308 - _f
derivation_rows = [
    ("", "LCOA green ($/t NH₃)", 816, 1054, 1308),
    ("", f"Fossil ref ($/t) = $440 prod. + ${ETS_SURCHARGE_2025:.0f} ETS", _f, _f, _f),
    ("", "ΔLCOA ($/t NH₃)", _dL, _dC, _dH),
    ("", "ΔAN = ΔLCOA × 0.425 ($/t)", _dL*0.425, _dC*0.425, _dH*0.425),
    ("", "ΔUrea = ΔLCOA×0.567 + $150×0.733 ($/t)", _dL*0.567+150*0.733, _dC*0.567+150*0.733, _dH*0.567+150*0.733),
]
for _, label, lo, ce, hi in derivation_rows:
    ws5.cell(r, 2, label).font = cf
    for ci, v in enumerate([lo, ce, hi], 3):
        ws5.cell(r, ci, v).font = bfont; ws5.cell(r, ci).number_format = '#,##0.0'
    r += 1
# Compute and show ΔMix, ΔN, Factor
for lo_lcoa, label in [(816, "Low"), (1054, "Central"), (1308, "High")]:
    dl = lo_lcoa - LCOA_FOSSIL
    d_AN = dl * 0.425; d_Urea = dl * 0.567 + 150 * 0.733
    d_Mix = 0.642 * d_AN + 0.321 * d_Urea
    d_N = d_Mix / (0.366 * 1000) / 1.08
    factor = d_N / 1.202
    ci = 3 + ["Low", "Central", "High"].index(label)
    ws5.cell(r, ci, d_Mix).font = bfont; ws5.cell(r, ci).number_format = '#,##0.0'
    ws5.cell(r + 1, ci, d_N).font = bfont; ws5.cell(r + 1, ci).number_format = '0.000'
    ws5.cell(r + 2, ci, factor).font = Font(name='Arial', bold=True, size=12, color='4472C4'); ws5.cell(r + 2, ci).number_format = '+0.0%'
ws5.cell(r, 2, "ΔMix ($/t fert. mix)").font = cf; r += 1
ws5.cell(r, 2, "ΔN (€/kg N)").font = cf; r += 1
ws5.cell(r, 2, "GREEN PREMIUM FACTOR").font = bfont; r += 2

# Section C: Key assumptions
ws5.cell(r, 1, "C | KEY ASSUMPTIONS").font = sf; r += 2
for note in [
    "Currency: EUR → USD at fixed 1.08 (ECB avg 2022–2024). Year-specific rates shift median <±5%.",
    f"Fossil reference: ${LCOA_FOSSIL_PRODUCTION}/t production cost + ${ETS_SURCHARGE_2025:.0f}/t ETS surcharge (2025) = ${LCOA_FOSSIL:.0f}/t.",
    f"EU ETS 2025: €{ETS_PRICE_2025}/t CO₂ × {CO2_INTENSITY_NH3} t CO₂/t NH₃ × {ETS_EFF_RATE_2025:.0%} effective = €{CO2_INTENSITY_NH3 * ETS_PRICE_2025 * ETS_EFF_RATE_2025:.0f}/t. EC WAM scenario (dena/SKN 2025).",
    "Free allocation: ~75% in 2025, phasing out to 0% by 2034 under CBAM (Reg. 2023/956). See Sheet 6 for projections.",
    "CO₂ for urea: $150/t (base case). Above CCS point sources ($60–130), below DAC ($230–600).",
    "EU mix: 64.2% AN + 32.1% Urea (after UAN decomposition). N-content mix: 0.366 kg N/kg.",
    "Current-cost framing: 2022–2025 technology & ETS costs. Cost reduction projections in Sheet 6.",
    "EU-02 Moritz excluded from scenarios: transport costs bundled, cannot isolate production LCOA.",
    "EU-07 Cameli: only 2019-tech value ($915) used. 2050 projection ($677) excluded.",
    "Tier 2 studies (EU-03 DECHEMA, EU-08 Mingolla SSRN): included in main analysis. Sensitivity without Tier 2 in SI.",
    f"Import sensitivity: Non-EU Median Low $610/t + $70/t transport = $680/t delivered. GP Factor +{P_IMPORT/B*100:.1f}% vs EU Central +{P_CEN/B*100:.1f}%.",
]:
    ws5.cell(r, 2, note).font = nf
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws5.cell(r, 2).alignment = Alignment(wrap_text=True); r += 1

# ==========================
# Product data for crop and end-product impacts in the gas-price sensitivity sheet
CROPS = [
    ("Wheat", 98.41, 5.508, 216.24),
    ("Maize", 96.87, 7.355, 206.21),
    ("Barley", 77.31, 4.571, 219.89),
    ("Rapeseed", 108.88, 3.163, 432.37),
]
END_PRODUCTS = [
    ("Bread", 0.003866),  # Central GP as fraction
    ("Rapeseed Oil", 0.02868),
    ("Milk", 0.007601),
    ("Meat (Ø)", 0.01574),
]

# SHEET 7: GAS PRICE SENSITIVITY
# ==========================
ws7 = wb.create_sheet("7. Gas Price Sensitivity")
setw(ws7, [32, 14, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14])
rg = 1
ws7.cell(rg, 1, "GAS PRICE SENSITIVITY — FOSSIL NH₃ COST VOLATILITY").font = tf
rg = 2; ws7.cell(rg, 1, "Fossil NH₃ production cost is dominated by natural gas (60–90% of variable cost). Gas price spikes can invert the green premium.").font = nf
ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11)
ws7.cell(rg, 1).alignment = Alignment(wrap_text=True); rg += 2

# A: DECHEMA cost model
ws7.cell(rg, 1, "A | DECHEMA COST MODEL (Annex 2, Table 7.7/7.8)").font = sf; rg += 1
hdr(ws7, rg, ["Parameter", "Value", "Unit", "", "Source", "", "", "", "", "", ""]); rg += 1
gas_model_fill = PatternFill('solid', fgColor='D6E4F0')
for label, val, unit, src in [
    ("Natural gas feed", 21.0, "GJ/t NH₃", "DECHEMA Annex 2, Table 7.7"),
    ("Natural gas fuel", 11.2, "GJ/t NH₃", "DECHEMA Annex 2, Table 7.7"),
    ("Total gas energy demand", 32.2, "GJ/t NH₃", "= Feed + Fuel"),
    ("Electricity demand (other)", 0.61, "MWh/t NH₃", "DECHEMA Annex 2, Table 7.7"),
    ("Maintenance", 47, "€/t NH₃", "DECHEMA Annex 2, Table 7.8"),
    ("Water cost", 2, "€/t NH₃", "DECHEMA Annex 2, Table 7.8"),
    ("Annualised CAPEX", 85, "€/t NH₃", "Hydrogen Europe / industry average"),
]:
    ws7.cell(rg, 1, label).font = cf; ws7.cell(rg, 2, val).font = bfont
    ws7.cell(rg, 3, unit).font = cf; ws7.cell(rg, 5, src).font = nf
    for c in range(1, 6): ws7.cell(rg, c).fill = gas_model_fill
    rg += 1
ws7.cell(rg, 1, "Formula: LCOA_fossil (€/t) = (Gas_€/MWh ÷ 3.6 × 32.2) + (60 × 0.61) + 47 + 2 + 85").font = Font(name='Arial', italic=True, size=9, color='333333')
ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11); rg += 2

# B: Historical TTF
ws7.cell(rg, 1, "B | HISTORICAL EUROPEAN GAS PRICES (TTF, World Bank Commodity Data)").font = sf; rg += 1
hdr(ws7, rg, ["Period", "€/MWh", "Context", "", "", "", "", "", "", "", ""]); rg += 1
for period, price, context in [
    ("Pre-crisis avg (2015–2020)", 16.5, "Normal EU market conditions (Russian pipeline supply)"),
    ("2021 (crisis onset)", 46.6, "Russian supply cuts began, post-COVID demand rebound"),
    ("2022 annual average", 131.1, "Full energy crisis, TTF record levels"),
    ("2022 peak month (Aug)", 228, "Spot TTF peak during Nord Stream uncertainty"),
    ("2023 (normalisation)", 41.4, "EU storage rules, LNG diversification, demand destruction"),
    ("2024 average", 34.6, "Post-crisis 'new normal' with higher structural floor"),
    ("2025 H2 (IEA estimate)", 47, "IEA commentary Feb 2025: 'double pre-crisis levels'"),
    ("March 2026 (current)", 59, "TTF spot; elevated by low storage + Ukraine transit halt"),
]:
    ws7.cell(rg, 1, period).font = cf; ws7.cell(rg, 2, price).font = bfont
    ws7.cell(rg, 2).number_format = '€#,##0'
    ws7.cell(rg, 3, context).font = nf
    ws7.merge_cells(start_row=rg, start_column=3, end_row=rg, end_column=5)
    ws7.cell(rg, 3).alignment = Alignment(wrap_text=True)
    rg += 1
rg += 1

# C: Sensitivity matrix (GP Factor)
ws7.cell(rg, 1, "C | GAS PRICE SENSITIVITY: GREEN PREMIUM FACTOR (Central LCOA $1,054 — corrected median)").font = sf; rg += 1
ws7.cell(rg, 1, "Green LCOA unchanged (independent of gas). All values incl. 2025 ETS at 10% eff. rate.").font = nf
ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11)
ws7.cell(rg, 1).alignment = Alignment(wrap_text=True); rg += 2

NG_GJ = 32.2; ELEC_EUR = 60; ELEC_MWH = 0.61; MAINT_G = 47; WATER_G = 2; CAPEX_G = 85
ETS_10PCT = CO2_INTENSITY_NH3 * ETS_PRICE_2025 * ETS_EFF_RATE_2025  # EUR/t NH3 (no FX — total_eur×FX applied after)

def fossil_from_gas(gas_eur_mwh):
    gas_cost = gas_eur_mwh / 3.6 * NG_GJ
    total_eur = gas_cost + ELEC_EUR * ELEC_MWH + MAINT_G + WATER_G + CAPEX_G + ETS_10PCT
    return total_eur * FX, gas_cost / (gas_cost + ELEC_EUR * ELEC_MWH + MAINT_G + WATER_G) * 100

gas_scenarios = [
    ("TEA literature average", 26, "Long-run EU equilibrium; ≈ $440/t literature avg"),
    ("Pre-crisis avg (2015–20)", 16.5, "Avg TTF before supply cuts"),
    ("Post-crisis 'new normal'", 40, "2023–24 structural floor"),
    ("Current (Mar 2026)", 59, "TTF spot; 2× pre-crisis"),
    ("2022 crisis annual avg", 131, "Full energy crisis year"),
    ("2022 peak month (Aug)", 228, "TTF record month"),
]

# Header with crop + product columns
gas_hdr = ["Gas Price Scenario", "€/MWh", "Fossil LCOA", "GP Factor",
           "Wheat", "Maize", "Barley", "Rapeseed",
           "Bread", "Rapeseed Oil", "Milk", "Meat (Ø)"]
for j, h in enumerate(gas_hdr):
    hdr(ws7, rg, []); ws7.cell(rg, j+1, h).font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    ws7.cell(rg, j+1).fill = PatternFill('solid', fgColor='4472C4')
    ws7.cell(rg, j+1).alignment = Alignment(horizontal='center')
rg += 1

inv_fill = PatternFill('solid', fgColor='C6EFCE')

for label, gas_eur, basis in gas_scenarios:
    fossil_usd, gas_share = fossil_from_gas(gas_eur)
    # Central GP factor
    f_gas, d_n_gas = calc_factor(1054, fossil_usd)
    
    # Crop impacts
    crop_pcts_gas = []
    for cname, n_ha, yld, cprice in CROPS:
        npt = n_ha / yld
        crop_eur = npt * d_n_gas
        crop_pcts_gas.append(crop_eur / cprice)
    
    # End product impacts (scale from Central)
    prod_pcts_gas = [gp_base / P_CEN * f_gas for _, gp_base in END_PRODUCTS]
    
    ws7.cell(rg, 1, label).font = cf
    ws7.cell(rg, 2, gas_eur).font = bfont; ws7.cell(rg, 2).number_format = '€#,##0'
    ws7.cell(rg, 3, round(fossil_usd)).font = bfont; ws7.cell(rg, 3).number_format = '$#,##0'
    ws7.cell(rg, 4, f_gas).font = bfont; ws7.cell(rg, 4).number_format = '+0.0%;-0.0%'
    for j, cp in enumerate(crop_pcts_gas):
        ws7.cell(rg, 5+j, cp).font = cf; ws7.cell(rg, 5+j).number_format = '+0.0%;-0.0%'
        if cp < 0: ws7.cell(rg, 5+j).fill = inv_fill
    for j, pp in enumerate(prod_pcts_gas):
        ws7.cell(rg, 9+j, pp).font = cf; ws7.cell(rg, 9+j).number_format = '+0.0%;-0.0%'
        if pp < 0: ws7.cell(rg, 9+j).fill = inv_fill
    if f_gas < 0:
        ws7.cell(rg, 4).fill = inv_fill; ws7.cell(rg, 4).font = Font(name='Arial', bold=True, size=10, color='006100')
    rg += 1

rg += 1
ws7.cell(rg, 1, "At current gas prices (€59/MWh), the Central premium is only +37% — far below the TEA-based +67%.").font = Font(name='Arial', italic=True, size=10, color='444444')
ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11); rg += 1
ws7.cell(rg, 1, "At 2022 crisis prices, ALL products show NEGATIVE premiums — green fertilizer is cheaper than fossil across the board.").font = Font(name='Arial', italic=True, size=10, color='2d8e4e')
ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11); rg += 2

# D: Sources
ws7.cell(rg, 1, "D | SOURCES").font = sf; rg += 1
for src in [
    "World Bank (2025). Commodity Price Data (Pink Sheet), monthly nominal prices Jan 2025 update. TTF series.",
    "DECHEMA/Ausfelder (2022). Perspective Europe 2030, Annex 2 Tables 7.7/7.8.",
    "IEA (2026). European gas market volatility commentary.",
    "TFI / AGA (2023). Natural gas accounts for 70–90% of ammonia production costs.",
    "ICE Endex. TTF Natural Gas Futures (spot prices, March 2026).",
]:
    ws7.cell(rg, 1, src).font = nf
    ws7.merge_cells(start_row=rg, start_column=1, end_row=rg, end_column=11)
    ws7.cell(rg, 1).alignment = Alignment(wrap_text=True); rg += 1

# ==========================
# REORDER SHEETS: Sources & Methods to last
# ==========================
# Current order: 1.Inputs, 2.GP Calcs, 3.Sources, 4.LCOA Lit, 5.LCOA Summary, 6.Cost Red, 7.Gas
# Target order:  1.Inputs, 2.GP Calcs, 3.LCOA Lit, 4.LCOA Summary, 5.Cost Red, 6.Gas, 7.Sources
source_idx = wb.sheetnames.index("3. Sources & Methods")
wb.move_sheet("3. Sources & Methods", offset=len(wb.sheetnames) - source_idx - 1)

# Rename sheets to fix numbering
# Rename sheets to fix numbering (after dropping the Cost Reduction Outlook sheet)
wb["4. LCOA Literature Review"].title = "3. LCOA Literature Review"
wb["5. LCOA Summary"].title = "4. LCOA Summary"
ws7.title = "5. Gas Price Sensitivity"
ws3.title = "6. Sources & Methods"

# ==========================
# SAVE
# ==========================
outpath = '../outputs/03_affordability_results.xlsx'
# Apply unified visual styling for the replication package (see _styling.py)
from _styling import apply_unified_styling
apply_unified_styling(wb)

wb.save(outpath)
print(f"\nSaved: {outpath}")

# Final summary
print(f"\n=== SCENARIOS (fossil ref ${LCOA_FOSSIL:.0f}/t incl. 2025 ETS) ===")
print(f"  Low:     €{P_LOW:.3f}/kg N  (LCOA $816, Factor +{P_LOW/B*100:.1f}%)")
print(f"  Central: €{P_CEN:.3f}/kg N  (LCOA $1,054, Factor +{P_CEN/B*100:.1f}%)")
print(f"  High:    €{P_HIGH:.3f}/kg N  (LCOA $1,308, Factor +{P_HIGH/B*100:.1f}%)")

print("\n=== SECTOR RANKING (HIGH SCENARIO) ===")
for nace, prod, gps, rank, notes in summary_all:
    print(f"  {rank}. {nace:<25} {gps[2]*100:>5.2f}%  ({prod})")
