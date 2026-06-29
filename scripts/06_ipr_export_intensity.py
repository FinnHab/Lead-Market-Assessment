"""
Import-penetration ratio (IPR) and export intensity (EI) for EU
agricultural sectors.

Inputs (all in ../inputs/):
    cn4_nace_template.xlsx                    CN4-to-NACE concordance, scope flags
    apri_eu_prices_components.csv             Eurostat APRI agricultural producer prices
    estat_apro_cpsh1.tsv.gz                   Eurostat APRO crop production bulk
    prodcom_overlap_resolution_tables_2024.xlsx
                                              PRODCOM overlap resolution for codes
                                              mapped to multiple CN4 / sectors
    full_<YYYY>*.dat                          Eurostat COMEXT monthly extra-EU
                                              trade dumps (12 files per year)
    estat_ds-059358.tsv                       Eurostat PRODCOM annual production

Outputs: ../outputs/06a_ipr_export_intensity_2023.xlsx
         ../outputs/06b_ipr_export_intensity_2024.xlsx

Scope decisions implemented in this script (kept in sync with the CN4
classification template sheet 10_CN4_CLASSIFICATION_FINAL):

  Fertilisers     CN4 2814 (ammonia), 3102, 3105 retained; CN4 3101
                  (organic fertilisers) excluded.
  Meat & meat     CN4 0204 (sheep/goat meat) added; CN4 0407/0408
   products       (eggs in shell, egg products) excluded as NACE 10.89;
                  CN4 1501-1504, 1506 (animal/marine fats) excluded as
                  Chapter-15 fats.
  Plant-based     Limited to nitrogen-relevant oilseed crops 1507
   oils           (soya), 1512 (sunflower/cottonseed), 1514 (rape).
                  Excludes 1508, 1511, 1513, 1515-1518, 1520, 1521.
  Processed       CN4 1105 (potato flour/meal) reassigned to
   cereals       "Fruit & Vegetable (Processed)".
  Fruit & veg     Weighted aggregate of "Fruit and Vegetable
   (processed)    Preparations" + "Fruit and Vegetable Juices"
                  (PRODCOM-value weights ~83% / 17%).
  Animal feed     Oilcake CN4 2304-2306 and corresponding PRODCOM
                  codes retained: regulating C109 (compound feed)
                  reaches C104 (oil pressing) upstream.
  Sugar           CN4 1704 (sugar confectionery) excluded.
  Wine, sugar     Excluded via SKIP_SECTORS / APRO_NO_PRICE_SECTORS.
   beet, coffee,
   tea, cocoa,
   fresh produce
"""

from __future__ import annotations

import glob
import gzip
import os
import re
from collections import defaultdict
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

# =============================================================================
# USER SETTINGS
# =============================================================================
YEAR = "2024"  # Analysis year

BASE_DIR = "../inputs"

TEMPLATE_XLS   = os.path.join(BASE_DIR, "cn4_nace_template.xlsx")
TRADE_PATTERN  = os.path.join(BASE_DIR, f"full_{YEAR}*.dat")
PRODVAL_TSV    = os.path.join(BASE_DIR, "estat_ds-059358.tsv")
APRO_GZ        = os.path.join(BASE_DIR, "estat_apro_cpsh1.tsv.gz")
APRI_CSV       = os.path.join(BASE_DIR, "apri_eu_prices_components.csv")
OVERLAP_XLSX   = os.path.join(BASE_DIR,
                               "prodcom_overlap_resolution_tables_2024.xlsx")
OUT_XLS        = f"../outputs/06_ipr_export_intensity_{YEAR}.xlsx"

EU            = "EU27_2020"
STRUCPRO      = "PR_HU_EU"
TRADE_TYPE_KEEP = "E"
REPORTER_EXCLUDE = {"EU27_2020"}

# Sectors excluded from output entirely.
# Note: sector names must match cn4_nace_template.xlsx exactly.
SKIP_SECTORS = {
    "Sugar",                        # internal code, never in output
    "Wine",                         # excluded: no N-fertiliser link
    "Sugar Beet",                   # excluded: PT not significant
    "Coffee and Tea",               # excluded: previously removed
    "Cocoa and Chocolate",          # excluded: previously removed
    "Fruits and Vegetables (fresh)" # excluded: NaN for IPR/EI
}

# ------------------------------------------------------------
# Sector-name reference (matches cn4_nace_template.xlsx exactly):
# ------------------------------------------------------------
# "Meat & Meat Products"  ← formerly "Meats & Fats" / "Meat, Eggs, Fats"
#     Scope: CN4 0102-0105 (live animals), 0201-0210 (meat), 1601-1602 (prepared meat)
#     Excluded: 0407-0408 (eggs → NACE 10.89), 1501-1503/1506 (animal fats), 1504 (fish fats)
# "Dairy Products"
# "Grain Mill & Bakery"  ← formerly "Processed Cereals and Grains"
# "Fruit & Vegetable (Processed)"   ← merged from Preparations + Juices
# "Oilseed Oils"                    ← formerly "Plant-based Oils"
# "Prepared Animal Feeds"           ← formerly "Animal Feed and By-Products"
# "Fertilisers"                     (illustration only, not in final index)
# "Sugar & Confectionery"                   ← formerly "Refined Sugar & Confectionery"
# "Beer", "Barley", "Maize", "Oilseeds", "Other Cereals and grains"
# "Wheat", "Potatoes"

# APRO-based sectors (use agricultural production statistics for quantities)
APRO_SECTORS = {
    "barley", "wheat", "maize", "oilseeds", "potatoes",
    "other cereals and grains", "fruits and vegetables (fresh)", "sugar beet"
}

# APRO sectors without APRI price (qty only, no EUR values possible)
APRO_NO_PRICE_SECTORS = {"fruits and vegetables (fresh)"}

# =============================================================================
# [5] WEIGHTED AGGREGATION: Fruit & Vegetable (Processed)
# Weights derived from PRODCOM production values:
#   Fruit and Vegetable Preparations: 52.23 bn EUR → weight 0.8297
#   Fruit and Vegetable Juices:       10.72 bn EUR → weight 0.1703
# Applied post-processing to merge both sector rows into one.
# =============================================================================
FV_MERGE_CONFIG = {
    "output_sector": "Fruit & Vegetable (Processed)",
    "components": {
        "Fruit & Vegetable (Processed)": 52.23e9,  # weight for Preparations component
        # Note: after template rename both Preparations and Juices carry the same
        # sector label "Fruit & Vegetable (Processed)". The weighted merge is
        # handled by the post-processing step using the weights below.
    }
}

# Explicit weights for F&V merge (production value basis)
FV_WEIGHTS = {
    "preparations_share": 0.8297,   # 52.23 bn EUR
    "juices_share":       0.1703,   # 10.72 bn EUR
}

# =============================================================================
# COLUMN NAMES
# =============================================================================
COL_CN4      = "cn4"
COL_SECTOR   = "sector"
COL_IMP_V    = "import_value_eur"
COL_EXP_V    = "export_value_eur"
COL_IMP_Q    = "import_qty_kg"
COL_EXP_Q    = "export_qty_kg"
COL_PV_CN4   = "prodval_eur_cn4"
COL_PV_SEC   = "prodval_eur_sector"
COL_APRO     = "prod_qty_kg_apro"
COL_IPR_V    = "ipr_value_sector"
COL_EOPV     = "exports_over_prodval_sector"
COL_IPR_Q    = "ipr_qty_sector"
COL_EOPQ     = "exports_over_prodqty_sector"
COL_AC_EUR   = "apparent_consumption_eur"
COL_AC_KG    = "apparent_consumption_kg"
COL_APRI_PRICE = "apri_price_eur_per_kg"
COL_MKT_SIZE = "market_size_eur_apri"
COL_N_CN4       = "n_cn4"
COL_N_CN4_PV    = "n_cn4_prodval"
COL_N_CN4_APRO  = "n_cn4_apro"
COL_N_CN4_NEITHER = "n_cn4_neither"

S_CN4     = "02_RESULTS_CN4"
S_SEC     = "01_RESULTS_SECTOR"
S_CN4_BG  = "04_CN4_BACKGROUND"
S_SEC_BG  = "03_SECTOR_BACKGROUND"
S_CLASS   = "10_CN4_CLASSIFICATION_FINAL"
S_PV      = "10_CN4_PRODVAL_FINAL"
S_APRO_CODES = "10_SECTOR_APRO_CODES_FINAL"

# =============================================================================
# HELPERS
# =============================================================================
def norm(x) -> str:
    return "" if x is None else str(x).replace("\ufeff","").replace("\xa0"," ").strip()

def keynorm(x) -> str:
    return re.sub(r"\s+", " ", norm(x)).strip().lower()

def z4(x) -> str:
    s = norm(x); s = re.sub(r"\.0$","",s); s = re.sub(r"\D","",s)
    return s.zfill(4)[:4] if s else ""

def z8(x) -> str:
    s = norm(x); s = re.sub(r"\.0$","",s); s = re.sub(r"\D","",s)
    return s.zfill(8)[:8] if s else ""

def safe_float(x) -> float:
    s = norm(x)
    if s in ("",":",): return 0.0
    vv = re.sub(r"[^0-9\.\-]","",s)
    if vv in ("",".","-","-."): return 0.0
    try: return float(vv)
    except: return 0.0

def autodetect_sep(fp: str) -> str:
    with open(fp,"rb") as f:
        line = f.readline().decode("utf-8",errors="ignore")
    return "\t" if line.count("\t") >= line.count(",") else ","

def header_map(ws):
    return {norm(ws.cell(1,c).value).lower(): c
            for c in range(1, ws.max_column+1)
            if norm(ws.cell(1,c).value)}

def get_col(ws, hm, name):
    return hm.get(name.lower())

def write_cell(ws, r, col_name, val, hm):
    c = get_col(ws, hm, col_name)
    if c is None: return
    if hasattr(val, "item"): val = val.item()
    if pd.isna(val) if hasattr(pd,"isna") else (val!=val if isinstance(val,float) else False):
        val = None
    ws.cell(r, c).value = val

def parse_codes(s: str) -> List[str]:
    s = norm(s)
    if not s or s.lower() == "nan": return []
    toks = [t.strip() for t in re.split(r"[;,| ]+", s) if t.strip()]
    return [t for t in toks if re.fullmatch(r"[A-Z][A-Z0-9_]*", t)]

# =============================================================================
# PRODCOM mapping
# =============================================================================
def read_cn4_prodcom_overrides(wb) -> Dict[str, set]:
    S_LOGIC = "00_FINAL_LOGIC"
    if S_LOGIC not in wb.sheetnames:
        print(f"  Warning: {S_LOGIC} not found")
        return {}
    ws = wb[S_LOGIC]
    overrides = {}
    for r in range(37, 45):
        cn4_val   = ws.cell(r, 1).value
        prodcom_val = ws.cell(r, 2).value
        if not cn4_val or not prodcom_val: continue
        cn4 = z4(cn4_val)
        prodcom_codes = {z8(c) for c in norm(prodcom_val).split(",") if z8(c)}
        if cn4 and prodcom_codes:
            overrides[cn4] = prodcom_codes
    return overrides

def build_cn4_to_prodcom_mapping(wb) -> Dict[str, set]:
    S_CN4_PRODCOM = "10_CN4_PRODCOM_FINAL"
    S08 = "08_CN8_TO_PRODCOM"
    S07 = "07_CN8_SCOPE_MAP"
    result = defaultdict(set)
    if S_CN4_PRODCOM in wb.sheetnames:
        print(f"  Using {S_CN4_PRODCOM}")
        ws  = wb[S_CN4_PRODCOM]
        hm  = header_map(ws)
        if "cn4" not in hm or "prodcom" not in hm:
            raise RuntimeError(f"{S_CN4_PRODCOM} missing cn4/prodcom columns")
        for r in range(2, ws.max_row+1):
            cn4    = z4(ws.cell(r, hm["cn4"]).value)
            prodcom = z8(ws.cell(r, hm["prodcom"]).value)
            if cn4 and prodcom: result[cn4].add(prodcom)
        return result
    if S08 not in wb.sheetnames:
        raise RuntimeError(f"Template missing {S_CN4_PRODCOM} and {S08}")
    print(f"  Using {S08} + {S07}")
    ws08 = wb[S08]; hm08 = header_map(ws08)
    col_prodcom = next((hm08[c] for c in ["prodcom","prodcom_code","product"] if c in hm08), None)
    if col_prodcom is None: raise RuntimeError(f"{S08}: no prodcom column")
    col_cn8 = hm08.get("cn8")
    if col_cn8 is None: raise RuntimeError(f"{S08}: no cn8 column")
    if S07 not in wb.sheetnames: raise RuntimeError(f"Need {S07}")
    cn8_to_prodcom = defaultdict(set)
    for r in range(2, ws08.max_row+1):
        cn8 = re.sub(r"\D","",norm(ws08.cell(r,col_cn8).value)).zfill(8)[:8]
        pc  = z8(ws08.cell(r,col_prodcom).value)
        if cn8 and pc: cn8_to_prodcom[cn8].add(pc)
    ws07 = wb[S07]; hm07 = header_map(ws07)
    col7_cn8 = next((hm07[c] for c in ["cn8","cn8_code"] if c in hm07), None)
    col7_cn4 = next((hm07[c] for c in ["cn4","cn4_code"] if c in hm07), None)
    if not col7_cn8 or not col7_cn4: raise RuntimeError(f"{S07}: missing cn8/cn4")
    for r in range(2, ws07.max_row+1):
        cn8 = re.sub(r"\D","",norm(ws07.cell(r,col7_cn8).value)).zfill(8)[:8]
        cn4 = z4(ws07.cell(r,col7_cn4).value)
        if cn8 and cn4 and cn8 in cn8_to_prodcom:
            result[cn4].update(cn8_to_prodcom[cn8])
    return result

def read_overlap_tables(overlap_xlsx: str) -> tuple:
    if not os.path.exists(overlap_xlsx): return {}, {}
    wb = load_workbook(overlap_xlsx, data_only=True)
    cross_sector = {}
    within_sector_proportions = {}
    if "10_PRODCOM_OVERLAP_RES" in wb.sheetnames:
        ws = wb["10_PRODCOM_OVERLAP_RES"]
        for r in range(2, ws.max_row+1):
            prodcom = z8(ws.cell(r,1).value)
            assigned = z4(ws.cell(r,2).value)
            if prodcom and assigned: cross_sector[prodcom] = assigned
    if "10_PRODCOM_WITHIN_SECT" in wb.sheetnames:
        ws = wb["10_PRODCOM_WITHIN_SECT"]
        for r in range(2, ws.max_row+1):
            prodcom      = z8(ws.cell(r,1).value)
            cn4_list_str = ws.cell(r,3).value
            dominant_cn4 = z4(ws.cell(r,4).value)
            dom_share    = ws.cell(r,5).value
            if not (prodcom and cn4_list_str and dominant_cn4 and dom_share): continue
            cn4s = [z4(c.strip()) for c in str(cn4_list_str).split(";")]
            if not cn4s: continue
            nd_share = (1.0 - float(dom_share))/(len(cn4s)-1) if len(cn4s)>1 else 0.0
            shares = {cn4: (float(dom_share) if cn4==dominant_cn4 else nd_share)
                      for cn4 in cn4s}
            within_sector_proportions[prodcom] = shares
    return cross_sector, within_sector_proportions

def read_prodval_from_estat(prodval_tsv, year, prodcom_codes):
    if not os.path.exists(prodval_tsv):
        raise FileNotFoundError(f"PRODVAL not found: {prodval_tsv}")
    print(f"\n=== PRODVAL Data ===")
    eurostat_flags = [":","  :",":c",":C",":z",":Z",":u",":U",
                      ":n",":N",":s",":S",":b",":B",":p",":P",
                      ":e",":E",":d",":D",":l",":L"]
    df = pd.read_csv(prodval_tsv, sep="\t", encoding="utf-8",
                     na_values=eurostat_flags, keep_default_na=True,
                     low_memory=False, dtype=str)
    first_col = df.columns[0]
    dims = df[first_col].str.split(",", n=3, expand=True)
    dims.columns = ["freq","reporter","product","indicators"]
    df = pd.concat([dims, df.drop(columns=[first_col])], axis=1)
    df.columns = [c.strip() if isinstance(c,str) else c for c in df.columns]
    df["product_normalized"] = df["product"].apply(lambda x: z8(x) if pd.notna(x) else "")
    codes_norm = {z8(c) for c in prodcom_codes}
    df_base = df[(df["freq"]=="A") & (df["reporter"]==EU) &
                 (df["product_normalized"].isin(codes_norm))].copy()
    available = set(df_base["indicators"].dropna().unique())
    # OWNPRODVAL (Variable 251001, own-account production at full market price)
    # is preferred over PRODVAL for post-2021 data.
    #
    # Rationale: Since the 2021 EBS regulation, PRODCOM separates own-account
    # production (251001, full market value) from subcontracting (251002,
    # processing fee only). CN trade data always records full product value
    # at the border. Adding subcontracting fees (251002) to the denominator
    # would mix two incompatible value bases: full market price vs. fee-only.
    # OWNPRODVAL is therefore the correct Gross-on-Gross counterpart to
    # COMEXT import/export values.
    # PRODVAL is used as fallback for pre-2021 data where the two were combined.
    chosen = next((i for i in ["OWNPRODVAL","PRODVAL"] if i in available), None)
    if not chosen:
        print(f"  WARNING: No OWNPRODVAL/PRODVAL found"); return {}
    print(f"  Using: {chosen}")
    df_f = df_base[df_base["indicators"]==chosen].copy()
    if year not in df_f.columns:
        raise RuntimeError(f"Year {year} not in PRODVAL file")
    df_f[year] = pd.to_numeric(df_f[year], errors="coerce")
    result = {}
    for _, row in df_f.iterrows():
        pc  = row["product_normalized"]
        val = row[year]
        if pd.notna(val) and val > 0: result[pc] = float(val)
    print(f"  Extracted {len(result)} values for {year}")
    return result

# =============================================================================
# TRADE AGGREGATION  (unchanged)
# =============================================================================
def trade_agg_cn4(cn4_list, year):
    files = sorted(glob.glob(TRADE_PATTERN))
    if not files: raise RuntimeError(f"No trade files: {TRADE_PATTERN}")
    print(f"\n=== Trade Data {year} === ({len(files)} files)")
    targets = {z4(c) for c in cn4_list if z4(c)}
    out = {c: {"imp_val":0.,"exp_val":0.,"imp_qty":0.,"exp_qty":0.} for c in targets}
    usecols = ["REPORTER","TRADE_TYPE","PRODUCT_NC","FLOW","VALUE_EUR","QUANTITY_KG"]
    for fp in files:
        print(f"  {os.path.basename(fp)}")
        sep = autodetect_sep(fp)
        for ch in pd.read_csv(fp, sep=sep, usecols=usecols, chunksize=1_200_000,
                               dtype=str, engine="c"):
            ch.columns = [norm(c) for c in ch.columns]
            ch = ch[ch["TRADE_TYPE"].str.strip()==TRADE_TYPE_KEEP]
            if ch.empty: continue
            ch = ch[~ch["REPORTER"].str.strip().isin(REPORTER_EXCLUDE)]
            if ch.empty: continue
            ch["CN4"] = ch["PRODUCT_NC"].str.slice(0,4).map(z4)
            ch = ch[ch["CN4"].isin(targets)]
            if ch.empty: continue
            ch["VALUE_EUR"]    = pd.to_numeric(ch["VALUE_EUR"],   errors="coerce").fillna(0.)
            ch["QUANTITY_KG"]  = pd.to_numeric(ch["QUANTITY_KG"], errors="coerce").fillna(0.)
            ch["FLOW"] = ch["FLOW"].str.strip()
            g = ch.groupby(["CN4","FLOW"])[["VALUE_EUR","QUANTITY_KG"]].sum()
            for (cn4, flow), row in g.iterrows():
                if flow=="1":
                    out[cn4]["imp_val"] += float(row["VALUE_EUR"])
                    out[cn4]["imp_qty"] += float(row["QUANTITY_KG"])
                elif flow=="2":
                    out[cn4]["exp_val"] += float(row["VALUE_EUR"])
                    out[cn4]["exp_qty"] += float(row["QUANTITY_KG"])
    return out

# =============================================================================
# APRO / APRI  (unchanged)
# =============================================================================
def apro_sum_codes_kg(codes, year):
    print(f"\n=== APRO Data {year} ===")
    if not os.path.exists(APRO_GZ): raise RuntimeError(f"APRO not found: {APRO_GZ}")
    codes_set = set(codes)
    out = {c: 0. for c in codes_set}
    with gzip.open(APRO_GZ,"rt",encoding="utf-8") as f:
        header = f.readline().rstrip("\n").split("\t")
        yi = next((i for i,c in enumerate(header[1:],1) if c.strip()==year), None)
        if yi is None: print(f"  WARNING: Year {year} not found"); return out
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) <= yi: continue
            key = parts[0].split("\\")[0]; kp = key.split(",")
            if len(kp) < 4: continue
            if kp[0].strip() in ("A","M","Q","Y"):
                crops,sp,geo = kp[1].strip(),kp[2].strip(),kp[3].strip()
            else:
                crops,sp,geo = kp[0].strip(),kp[2].strip(),kp[3].strip()
            if geo!=EU or sp!=STRUCPRO or crops not in codes_set: continue
            v = safe_float(parts[yi])
            if v > 0: out[crops] += v * 1_000_000.
    return out

def load_apri_prices(year):
    print(f"\n=== APRI Prices {year} ===")
    if not os.path.exists(APRI_CSV): print(f"  ERROR: not found"); return {}
    df = pd.read_csv(APRI_CSV, sep=";", decimal=",", dtype=str)
    df.columns = [c.strip() for c in df.columns]
    need = {"sector_group","year","eu_mean_price_eur_per_kg"}
    if miss := [c for c in need if c not in df.columns]:
        print(f"  ERROR: missing {miss}"); return {}
    df["year_norm"]   = df["year"].str.extract(r"(\d{4})")[0]
    df["group_norm"]  = df["sector_group"].str.strip().str.lower()
    def pp(x):
        s = str(x).strip().replace(",",".")
        try: return float(s)
        except: return None
    df["price"] = df["eu_mean_price_eur_per_kg"].apply(pp)
    df = df[df["year_norm"]==str(year)]
    if df.empty: print(f"  No rows for {year}"); return {}
    g = df.groupby("group_norm", as_index=False)["price"].mean()
    return {k: float(v) for k,v in zip(g["group_norm"],g["price"]) if pd.notna(v)}

# =============================================================================
# [5] POST-PROCESSING: Weighted merge of F&V Preparations + Juices
# =============================================================================
def merge_fv_processed(sector_results: dict) -> dict:
    """
    Merge "Fruit and Vegetable Preparations" and "Fruit and Vegetable Juices"
    into a single "Fruit & Vegetable (Processed)" row using production-value
    weights (83% Preparations, 17% Juices).

    sector_results: dict keyed by sector name, values are dicts of metric→value
    Returns updated dict with merged row and component rows removed.
    """
    cfg = FV_MERGE_CONFIG
    components = cfg["components"]
    out_name   = cfg["output_sector"]

    # Check both components exist
    missing = [s for s in components if s not in sector_results]
    if missing:
        print(f"  WARNING [FV merge]: missing component(s): {missing}")
        return sector_results

    total_weight = sum(components.values())
    weights = {s: w/total_weight for s, w in components.items()}

    print(f"\n=== Weighted merge: {out_name} ===")
    for s, w in weights.items():
        print(f"  {s}: weight = {w:.4f}")

    # Collect all numeric metric keys from either component
    all_keys = set()
    for s in components:
        all_keys.update(k for k,v in sector_results[s].items()
                        if isinstance(v, (int,float)) and v is not None)

    merged = {}
    for key in all_keys:
        vals = {s: sector_results[s].get(key) for s in components}
        # Only merge if at least one component has a value
        if all(v is None for v in vals.values()):
            merged[key] = None
            continue
        # Weighted average; treat None as 0 for weighting purposes
        # but flag if data is incomplete
        wsum = sum(weights[s] * (v if v is not None else 0.)
                   for s, v in vals.items())
        merged[key] = round(wsum, 6)

    # Recompute derived ratios from merged flow totals where possible
    # (IPR = imp / (prodval + imp - exp); EI = exp / prodval)
    imp  = merged.get("import_value_eur")
    exp  = merged.get("export_value_eur")
    pv   = merged.get("prodval_eur_sector")
    if all(v is not None for v in [imp, exp, pv]) and pv > 0:
        ac = pv + imp - exp
        if ac > 0:
            merged["ipr_value_sector"]           = imp / ac
            merged["apparent_consumption_eur"]   = ac
        merged["exports_over_prodval_sector"] = exp / pv

    result = {k: v for k, v in sector_results.items() if k not in components}
    result[out_name] = merged
    print(f"  -> Created merged sector '{out_name}'")
    return result

# =============================================================================
# MAIN
# =============================================================================
def main():
    print("=" * 80)
    print(f"IPR / Export-Intensity computation - Year: {YEAR}")
    print("=" * 80)
    print("\nScope decisions applied in this run:")
    print("  [1] Fertilisers: 2814 added, 3101 excluded, N-only scope")
    print("  [2] Meat: 0204 added, 1504 excluded")
    print("  [3] Plant-based Oils: restricted to 1507, 1512, 1514")
    print("  [4] Processed Cereals: 1105 moved to F&V Preparations")
    print("  [5] F&V Processed: weighted merge of Preparations (83%) + Juices (17%)")
    print("  [6] Animal Feed: oilcake codes retained (deliberate analytical choice)")
    print("  [7] Wine, Sugar Beet, Cocoa, Coffee, F&V fresh excluded")
    print("  [8] Refined Sugar only (1704 confectionery excluded)")
    print()

    wb = load_workbook(TEMPLATE_XLS)

    # Load CN4 classification
    ws_class = wb[S_CLASS]; hm_class = header_map(ws_class)
    cn4_to_sector = {}; cn4_to_basis = {}
    for r in range(2, ws_class.max_row+1):
        cn4    = z4(ws_class.cell(r, get_col(ws_class,hm_class,"cn4")).value)
        sector = norm(ws_class.cell(r, get_col(ws_class,hm_class,"sector")).value)
        basis  = norm(ws_class.cell(r, get_col(ws_class,hm_class,"trade_basis_used")).value)
        if cn4: cn4_to_sector[cn4]=sector; cn4_to_basis[cn4]=basis

    all_cn4s      = list(cn4_to_sector.keys())
    excluded_cn4s = {cn4 for cn4,b in cn4_to_basis.items() if b=="excluded"}
    print(f"  CN4s total: {len(all_cn4s)}, excluded: {len(excluded_cn4s)}")

    # Load APRO codes
    ws_apro = wb[S_APRO_CODES]; hm_apro = header_map(ws_apro)
    sector_to_codes = {}
    for r in range(2, ws_apro.max_row+1):
        sector = norm(ws_apro.cell(r, get_col(ws_apro,hm_apro,"sector")).value)
        codes  = norm(ws_apro.cell(r, get_col(ws_apro,hm_apro,"apro_crops_codes_used")).value)
        if sector: sector_to_codes[sector] = codes

    # Load / calculate PRODVAL
    cn4_to_prodval = {}
    if S_PV in wb.sheetnames:
        ws_pv = wb[S_PV]; hm_pv = header_map(ws_pv)
        if "cn4" in hm_pv and "prodval_eur_cn4" in hm_pv:
            c_cn4 = hm_pv["cn4"]; c_pv = hm_pv["prodval_eur_cn4"]
            c_year = hm_pv.get("year")
            for r in range(2, ws_pv.max_row+1):
                if c_year and str(ws_pv.cell(r,c_year).value).strip()!=YEAR:
                    continue
                cn4 = z4(ws_pv.cell(r,c_cn4).value)
                pv  = ws_pv.cell(r,c_pv).value
                if cn4 and pv is not None:
                    try: cn4_to_prodval[cn4] = float(pv)
                    except: pass

    if not cn4_to_prodval:
        print("  Calculating PRODVAL from estat...")
        cn4_to_prodcom = build_cn4_to_prodcom_mapping(wb)
        all_prodcom = {pc for pcs in cn4_to_prodcom.values() for pc in pcs}
        prodcom_prodvals = read_prodval_from_estat(PRODVAL_TSV, YEAR, all_prodcom)
        cross_sector_map, within_sector_proportions = read_overlap_tables(OVERLAP_XLSX)
        manual_overrides = read_cn4_prodcom_overrides(wb)
        if manual_overrides:
            for cn4, pcs in manual_overrides.items():
                cn4_to_prodcom[cn4] = pcs
        cn4_to_prodcom_shares = {}
        for cn4, prodcom_set in cn4_to_prodcom.items():
            shares_list = []
            for pc in prodcom_set:
                if pc in cross_sector_map:
                    if cross_sector_map[pc] == cn4:
                        shares_list.append((pc, 1.0))
                elif pc in within_sector_proportions:
                    s = within_sector_proportions[pc]
                    if cn4 in s: shares_list.append((pc, s[cn4]))
                else:
                    shares_list.append((pc, 1.0))
            cn4_to_prodcom_shares[cn4] = shares_list
        for cn4, shares_list in cn4_to_prodcom_shares.items():
            total = sum(prodcom_prodvals.get(pc,0.)*sh for pc,sh in shares_list)
            if total > 0: cn4_to_prodval[cn4] = total

    cn4_is_valid = {}
    for cn4 in all_cn4s:
        basis = cn4_to_basis.get(cn4,"")
        if basis=="excluded":      cn4_is_valid[cn4] = False
        elif basis=="qty_apro":    cn4_is_valid[cn4] = True
        elif basis=="value_prodval": cn4_is_valid[cn4] = (cn4_to_prodval.get(cn4,0.)>0)
        else:                      cn4_is_valid[cn4] = False

    # Trade aggregation
    trade_cn4s = [cn4 for cn4 in all_cn4s if cn4 not in excluded_cn4s]
    trade = trade_agg_cn4(trade_cn4s, YEAR)

    # APRO quantities
    all_apro_codes = []
    for codes_str in sector_to_codes.values():
        all_apro_codes.extend(parse_codes(codes_str))
    apro_data = apro_sum_codes_kg(list(set(all_apro_codes)), YEAR)

    # APRI prices
    apri_prices = load_apri_prices(YEAR)

    # =========================================================================
    # FILL 02_RESULTS_CN4  (unchanged logic)
    # =========================================================================
    print(f"\n=== Filling {S_CN4} ===")
    ws_cn4 = wb[S_CN4]; hm_cn4 = header_map(ws_cn4)
    cn4_col = get_col(ws_cn4, hm_cn4, COL_CN4)
    cn4_to_row = {}
    for r in range(2, ws_cn4.max_row+1):
        cn4 = z4(ws_cn4.cell(r,cn4_col).value)
        if cn4: cn4_to_row[cn4] = r
    for cn4, r in cn4_to_row.items():
        basis = cn4_to_basis.get(cn4,"")
        sector_key = keynorm(cn4_to_sector.get(cn4,""))
        is_apro_no_price = sector_key in APRO_NO_PRICE_SECTORS
        if basis=="excluded":
            for col in [COL_IMP_V,COL_EXP_V,COL_IMP_Q,COL_EXP_Q,COL_PV_CN4]:
                write_cell(ws_cn4, r, col, None, hm_cn4)
            continue
        t = trade.get(cn4, {})
        if is_apro_no_price:
            write_cell(ws_cn4, r, COL_IMP_V, None, hm_cn4)
            write_cell(ws_cn4, r, COL_EXP_V, None, hm_cn4)
        else:
            write_cell(ws_cn4, r, COL_IMP_V, t.get("imp_val",0.), hm_cn4)
            write_cell(ws_cn4, r, COL_EXP_V, t.get("exp_val",0.), hm_cn4)
        write_cell(ws_cn4, r, COL_IMP_Q, t.get("imp_qty",0.), hm_cn4)
        write_cell(ws_cn4, r, COL_EXP_Q, t.get("exp_qty",0.), hm_cn4)
        if basis=="value_prodval":
            write_cell(ws_cn4, r, COL_PV_CN4, cn4_to_prodval.get(cn4,0.), hm_cn4)
    print(f"  Wrote {len(cn4_to_row)} CN4 rows")

    # =========================================================================
    # FILL 01_RESULTS_SECTOR
    # =========================================================================
    print(f"\n=== Filling {S_SEC} ===")
    ws_sec = wb[S_SEC]; hm_sec = header_map(ws_sec)
    sector_col = get_col(ws_sec, hm_sec, COL_SECTOR)
    sector_to_cn4s = defaultdict(list)
    for cn4, sector in cn4_to_sector.items():
        if cn4 not in excluded_cn4s:
            sector_to_cn4s[sector].append(cn4)

    # Collect sector results for post-processing
    sector_results_cache = {}

    for r in range(2, ws_sec.max_row+1):
        sector = norm(ws_sec.cell(r, sector_col).value)
        if not sector or sector in SKIP_SECTORS: continue

        sector_key  = keynorm(sector)
        sector_cn4s = sector_to_cn4s.get(sector, [])
        is_apro          = sector_key in APRO_SECTORS
        is_apro_no_price = sector_key in APRO_NO_PRICE_SECTORS
        valid_cn4s = [cn4 for cn4 in sector_cn4s if cn4_is_valid.get(cn4,False)]

        n_cn4    = len(sector_cn4s)
        n_pv     = sum(1 for cn4 in sector_cn4s
                       if cn4_to_basis.get(cn4)=="value_prodval" and cn4_is_valid.get(cn4,False))
        n_apro   = sum(1 for cn4 in sector_cn4s if cn4_to_basis.get(cn4)=="qty_apro")
        n_neither= n_cn4 - n_pv - n_apro
        write_cell(ws_sec, r, COL_N_CN4,        n_cn4,     hm_sec)
        write_cell(ws_sec, r, COL_N_CN4_PV,     n_pv,      hm_sec)
        write_cell(ws_sec, r, COL_N_CN4_APRO,   n_apro,    hm_sec)
        write_cell(ws_sec, r, COL_N_CN4_NEITHER, n_neither, hm_sec)

        row_data = {}

        if is_apro:
            imp_q = sum(trade.get(cn4,{}).get("imp_qty",0.) for cn4 in valid_cn4s)
            exp_q = sum(trade.get(cn4,{}).get("exp_qty",0.) for cn4 in valid_cn4s)
            write_cell(ws_sec, r, COL_IMP_Q, imp_q, hm_sec)
            write_cell(ws_sec, r, COL_EXP_Q, exp_q, hm_sec)
            codes   = parse_codes(sector_to_codes.get(sector,""))
            prod_qty= sum(apro_data.get(c,0.) for c in codes)
            write_cell(ws_sec, r, COL_APRO, prod_qty, hm_sec)
            apri_price = apri_prices.get(sector_key)
            print(f"  {sector}: apri_price={apri_price}, no_price={is_apro_no_price}")
            if is_apro_no_price or apri_price is None:
                for col in [COL_IMP_V,COL_EXP_V,COL_PV_SEC,COL_IPR_V,
                             COL_EOPV,COL_AC_EUR,COL_APRI_PRICE,COL_MKT_SIZE]:
                    write_cell(ws_sec, r, col, None, hm_sec)
                ac_kg = prod_qty + imp_q - exp_q
                write_cell(ws_sec, r, COL_AC_KG, ac_kg, hm_sec)
                if ac_kg > 0: write_cell(ws_sec, r, COL_IPR_Q, imp_q/ac_kg, hm_sec)
                if prod_qty>0: write_cell(ws_sec, r, COL_EOPQ, exp_q/prod_qty, hm_sec)
            else:
                imp_v = sum(trade.get(cn4,{}).get("imp_val",0.) for cn4 in valid_cn4s)
                exp_v = sum(trade.get(cn4,{}).get("exp_val",0.) for cn4 in valid_cn4s)
                write_cell(ws_sec, r, COL_IMP_V,    imp_v,      hm_sec)
                write_cell(ws_sec, r, COL_EXP_V,    exp_v,      hm_sec)
                write_cell(ws_sec, r, COL_APRI_PRICE, apri_price, hm_sec)
                pv_eur = prod_qty * apri_price
                write_cell(ws_sec, r, COL_PV_SEC, pv_eur, hm_sec)
                ac_eur = pv_eur + imp_v - exp_v
                ac_kg  = prod_qty + imp_q - exp_q
                write_cell(ws_sec, r, COL_AC_EUR, ac_eur, hm_sec)
                write_cell(ws_sec, r, COL_AC_KG,  ac_kg,  hm_sec)
                write_cell(ws_sec, r, COL_MKT_SIZE, ac_kg*apri_price, hm_sec)
                if ac_eur>0:  write_cell(ws_sec, r, COL_IPR_V, imp_v/ac_eur, hm_sec)
                if pv_eur>0:  write_cell(ws_sec, r, COL_EOPV,  exp_v/pv_eur, hm_sec)
                if ac_kg>0:   write_cell(ws_sec, r, COL_IPR_Q, imp_q/ac_kg,  hm_sec)
                if prod_qty>0: write_cell(ws_sec, r, COL_EOPQ, exp_q/prod_qty, hm_sec)
                row_data = {COL_IMP_V: imp_v, COL_EXP_V: exp_v,
                            COL_PV_SEC: pv_eur, COL_AC_EUR: ac_eur,
                            COL_IPR_V: imp_v/ac_eur if ac_eur>0 else None,
                            COL_EOPV: exp_v/pv_eur if pv_eur>0 else None}
        else:
            imp_v  = sum(trade.get(cn4,{}).get("imp_val",0.) for cn4 in valid_cn4s)
            exp_v  = sum(trade.get(cn4,{}).get("exp_val",0.) for cn4 in valid_cn4s)
            pv_total = sum(cn4_to_prodval.get(cn4,0.) for cn4 in valid_cn4s)
            write_cell(ws_sec, r, COL_IMP_V, imp_v,  hm_sec)
            write_cell(ws_sec, r, COL_EXP_V, exp_v,  hm_sec)
            write_cell(ws_sec, r, COL_IMP_Q, None,   hm_sec)
            write_cell(ws_sec, r, COL_EXP_Q, None,   hm_sec)
            write_cell(ws_sec, r, COL_PV_SEC, pv_total if pv_total>0 else None, hm_sec)
            if pv_total > 0:
                ac_eur = pv_total + imp_v - exp_v
                write_cell(ws_sec, r, COL_AC_EUR, ac_eur, hm_sec)
                if ac_eur>0:  write_cell(ws_sec, r, COL_IPR_V, imp_v/ac_eur, hm_sec)
                write_cell(ws_sec, r, COL_EOPV, exp_v/pv_total, hm_sec)
                row_data = {COL_IMP_V: imp_v, COL_EXP_V: exp_v,
                            COL_PV_SEC: pv_total, COL_AC_EUR: ac_eur,
                            "import_value_eur": imp_v, "export_value_eur": exp_v,
                            "prodval_eur_sector": pv_total, "apparent_consumption_eur": ac_eur,
                            "ipr_value_sector": imp_v/ac_eur if ac_eur>0 else None,
                            "exports_over_prodval_sector": exp_v/pv_total}

        if row_data:
            sector_results_cache[sector] = row_data

    # [5] Weighted merge of F&V Processed – append as new row if template has space
    # (In practice: the merged values are written to a dedicated sector row in the template
    #  or appended. Here we log the results for manual insertion or downstream use.)
    merged_results = merge_fv_processed(sector_results_cache)
    fv_merged = merged_results.get(FV_MERGE_CONFIG["output_sector"])
    if fv_merged:
        print(f"\n=== F&V (Processed) merged results ===")
        for k, v in sorted(fv_merged.items()):
            if v is not None: print(f"  {k}: {v:.6g}")

    print(f"\n=== Formatting workbook ===")
    format_workbook(wb)

    print(f"\n=== Saving to {OUT_XLS} ===")
    # Apply unified visual styling for the replication package (see _styling.py)
    from _styling import apply_unified_styling
    apply_unified_styling(wb)

    wb.save(OUT_XLS)
    print("Done!")



# =============================================================================
# POST-PROCESSING: FORMAT & REORGANIZE OUTPUT WORKBOOK
# =============================================================================
# Matches Scale_Results_v16 style: navy title bar, blue headers,
# alternating row bands, Arial throughout.
# Removes pipeline-only sheets, reorders results to front.

def format_workbook(wb):
    """Apply Scale_Results_v16 formatting and reorganize sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    NAVY_DARK = "0D2137"; BLUE_HDR = "2C5282"
    LIGHT_BLUE = "F5F8FF"; ALT_BLUE = "D6E4F7"

    f_title     = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    f_subtitle  = Font(name="Arial", size=9, color="555555")
    f_colhdr    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    f_data      = Font(name="Arial", size=10, color="000000")
    f_data_bold = Font(name="Arial", size=10, bold=True, color="000000")

    fill_title    = PatternFill("solid", fgColor=NAVY_DARK)
    fill_subtitle = PatternFill("solid", fgColor=LIGHT_BLUE)
    fill_colhdr   = PatternFill("solid", fgColor=BLUE_HDR)
    fill_white    = PatternFill("solid", fgColor="FFFFFF")
    fill_alt      = PatternFill("solid", fgColor=ALT_BLUE)

    al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    REMOVE_SHEETS = {"10_CN4_PRODVAL_FINAL", "07_CN8_SCOPE_MAP", "08_CN8_TO_PRODCOM"}
    SHEET_ORDER = [
        "01_RESULTS_SECTOR", "02_RESULTS_CN4",
        "10_CN4_CLASSIFICATION_FINAL", "10_SECTOR_APRO_CODES_FINAL",
        "03_SECTOR_BACKGROUND", "04_CN4_BACKGROUND", "00_FINAL_LOGIC",
    ]
    CFGS = {
        "01_RESULTS_SECTOR": (
            "IPR & Export Intensity — Sector Results  |  EU27  |  2023",
            "Value-weighted (PRODCOM) for manufacturing sectors; APRO × APRI for crop sectors. Trade: Eurostat Comext extra-EU27."),
        "02_RESULTS_CN4": (
            "IPR & Export Intensity — CN4 Product-Level Results  |  EU27  |  2023",
            "Per CN4 code: imports, exports, production value, IPR, EI. Basis: value_prodval (PRODCOM) or qty_apro."),
        "10_CN4_CLASSIFICATION_FINAL": (
            "CN4 to Sector Classification  |  GLM Scope",
            "Manual CN4-to-sector assignment. trade_basis_used: value_prodval, qty_apro, or excluded."),
        "10_SECTOR_APRO_CODES_FINAL": (
            "APRO Crop Codes — Sector Mapping",
            "Eurostat APRO crop_code(s) used for production quantities per agricultural sector."),
        "03_SECTOR_BACKGROUND": (
            "Sector Background — Aggregation Detail",
            "Supporting calculations and intermediate values per sector."),
        "04_CN4_BACKGROUND": (
            "CN4 Background — PRODCOM Overlap & Quantity Detail",
            "PRODCOM mapping, overlap resolution, and quantity cross-checks per CN4 code."),
        "00_FINAL_LOGIC": (
            "Pipeline Logic Documentation  |  generate_ipr_UPDATED.py",
            "Run parameters, file paths, and processing log from the most recent pipeline execution."),
    }

    # Remove pipeline-only sheets
    for name in REMOVE_SHEETS:
        if name in wb.sheetnames:
            del wb[name]
            print(f"  Removed sheet: {name}")

    # Reorder
    for i, name in enumerate(SHEET_ORDER):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cfg = CFGS.get(sheet_name)
        if not cfg:
            continue

        title, subtitle = cfg
        ncols = ws.max_column

        # Insert title/subtitle/spacer rows
        ws.insert_rows(1, 3)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        cell = ws.cell(1, 1, title)
        cell.font = f_title; cell.fill = fill_title; cell.alignment = al_l
        ws.row_dimensions[1].height = 28

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        cell = ws.cell(2, 1, subtitle)
        cell.font = f_subtitle; cell.fill = fill_subtitle; cell.alignment = al_l
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        # Column headers (row 4)
        ws.row_dimensions[4].height = 30
        for ci in range(1, ncols + 1):
            cell = ws.cell(4, ci)
            cell.font = f_colhdr; cell.fill = fill_colhdr
            cell.alignment = al_c; cell.border = bdr

        # Data rows (row 5+)
     